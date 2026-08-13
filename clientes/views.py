from django.utils import timezone
from django.db.models import Q, Count, OuterRef, Subquery
from django.core.paginator import Paginator
from django.shortcuts import render,redirect, get_object_or_404
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
import threading
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.urls import reverse
from modelo_equipamento.models import Modelo_equipamento
from funcao_equipamento.models import Funcao_equipamento
from django.http import JsonResponse
from .models import Cliente, Acesso, Documento, ArquivoVPN, ImagemTopologia, Categoria, Chamado, ComentarioChamado, BackupLog,  BackupTemplate, ComentarioAcesso, OpenVPNConfig
from .models import AcaoL2vpn
from .models import ProxyServer
from .models import AcessoSessao, AcessoComando, TerminalLinkExterno
from .proxy_engine import ProxyEngine
from .decorators import admin_required, cliente_login_required
from usuario import perms as _perms
from .decorators import (
    cliente_login_required,
    admin_required,
    cliente_or_admin_required,
    cliente_can_view_cliente,
    modulo_habilitado_required,
    backoffice_required,
)
from django.http import HttpResponseRedirect
import logging
logger = logging.getLogger(__name__)
from django.http import HttpResponse
import json
import re
import pexpect
import telnetlib
import threading
import ipaddress
import logging
import paramiko
import socket
import os
import hashlib
from pathlib import Path
from datetime import datetime, timedelta
from django.conf import settings
from django.http import FileResponse
from django.http import Http404
import time
from .models import BackupTemplate
from .models import BlocoIP, ValidacaoRPKI_IRR_Log
import requests
import ipaddress
from .models import BlocoIP, ValidacaoRPKI_IRR_Log
# Instalar: pip install netmiko
from netmiko import ConnectHandler
from netmiko.exceptions import NetmikoTimeoutException, NetmikoAuthenticationException


@login_required(login_url='login')
@cliente_can_view_cliente  # ✅ NOVO: Validar se cliente pode ver este cliente
def listar_clientes(request):
    """
    View que lista acessos e dados do cliente.
    - Clientes podem ver APENAS seus próprios dados
    - Admins podem ver qualquer cliente
    """
    id_cliente = request.GET.get('id')

    if not id_cliente:
        messages.error(request, 'Cliente não especificado.')
        return redirect('quadro_geral')

    cliente = get_object_or_404(Cliente, id=id_cliente)

    # ✅ VALIDAÇÃO: Verificar permissão (redundante com o decorator acima,
    # mantido como segunda camada — mesma lógica central para todo mundo)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        messages.error(request, 'Você não possui permissão para visualizar este cliente.')
        if _perms.is_admin(request.user):
            return redirect('quadro_geral')
        if _perms.is_backoffice(request.user):
            return redirect('cadastrar_cliente')
        return redirect('login')

    # Restante do código existente...
    funcao_selecionada = request.GET.get('funcao')
    modelos = Modelo_equipamento.objects.all()
    funcao_equipamentos = Funcao_equipamento.objects.all()

    funcoes = cliente.acessos.values_list('funcao', flat=True).distinct()

    if funcao_selecionada:
        acessos = cliente.acessos.filter(funcao=funcao_selecionada)
    else:
        acessos = cliente.acessos.all()

    documentos = Documento.objects.filter(cliente=cliente).order_by('-data_upload')
    arquivos_vpn = ArquivoVPN.objects.filter(cliente=cliente).order_by('-data_upload')
    imagens_topologia = ImagemTopologia.objects.filter(cliente=cliente).order_by('-data_upload')
    proxies = ProxyServer.objects.filter(cliente=cliente).order_by('-ativo', 'nome')

    # ✅ NOVO: Adicionar flag de tipo de usuário ao contexto
    # is_admin aqui significa "não é o portal do cliente final" (equipe:
    # Administrador, Consultor ou Operador) — controla a UI operacional do
    # dashboard. is_superuser continua restrito ao Administrador de fato.
    is_cliente = False
    is_admin = _perms.is_backoffice(request.user)
    # Distinto de is_admin acima: aqui é só o Administrador de fato — usado
    # para o bypass das abas/ferramentas, que para Consultor/Operador deve
    # respeitar o que a instância tem liberado (InstanciaFerramenta),
    # exatamente como para o portal do cliente final dele.
    is_admin_puro = _perms.is_admin(request.user)
    is_superuser = request.user.is_superuser
    try:
        if not is_admin and Cliente.objects.get_by_usuario_vinculado(request.user).id == cliente.id:
            is_cliente = True
    except:
        pass

    # ── Detectar acessos com erro de backup ─────────────────────────────
    acessos_com_erro_backup = []
    for ac in cliente.acessos.filter(backup_habilitado=True):
        ultimo_backup = BackupLog.objects.filter(acesso=ac).order_by('-data_backup').first()
        if ultimo_backup and ultimo_backup.status == 'ERRO':
            acessos_com_erro_backup.append({
                'tipo': ac.tipo,
                'host': ac.host,
                'mensagem': ultimo_backup.mensagem or 'Erro desconhecido',
                'data': ultimo_backup.data_backup.strftime('%d/%m/%Y %H:%M'),
            })

    # ── Detectar blocos com erro de validação RPKI/IRR ───────────────────
    blocos_rpki_invalidos_cliente = cliente.blocos_ip.filter(
        Q(rpki_valido=False) | Q(rpki_status__in=['Invalid', 'Unknown', 'Error', 'NotChecked'])
    )
    blocos_irr_invalidos_cliente = cliente.blocos_ip.filter(
        Q(irr_valido=False) | Q(irr_status__in=['NotFound', 'ASN_Mismatch', 'Error'])
    )
    total_blocos_rpki_irr_invalidos_cliente = blocos_rpki_invalidos_cliente.count() + blocos_irr_invalidos_cliente.count()

    modulos_habilitados = _perms.modulos_habilitados_dict_para_listagem(request.user, cliente)

    response = render(request, 'listar.html', {
        'cliente': cliente,
        'funcoes': funcoes,
        'acessos': acessos,
        'funcao_selecionada': funcao_selecionada,
        'modelos': modelos,
        'funcao_equipamentos': funcao_equipamentos,
        'documentos': documentos,
        'arquivos_vpn': arquivos_vpn,
        'imagens_topologia': imagens_topologia,
        'proxies': proxies,
        'is_cliente': is_cliente,
        'is_admin': is_admin,
        'is_admin_puro': is_admin_puro,
        'is_superuser': is_superuser,
        'modulos_habilitados': modulos_habilitados,
        'destinos_padrao': DESTINOS_PADRAO,
        'acessos_com_erro_backup': acessos_com_erro_backup,
        'blocos_rpki_invalidos_cliente': blocos_rpki_invalidos_cliente,
        'blocos_irr_invalidos_cliente': blocos_irr_invalidos_cliente,
        'total_blocos_rpki_irr_invalidos_cliente': total_blocos_rpki_irr_invalidos_cliente,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response

def _validar_usuarios_adicionais(usuario_ids, cliente_id_atual=None):
    """
    Um usuário só pode estar vinculado a UM cliente por vez (seja como
    principal ou como adicional) — isso mantém "para qual painel eu vou ao
    logar" sem ambiguidade. Verifica cada id em `usuario_ids` contra
    qualquer outro Cliente (excluindo `cliente_id_atual`, no caso de edição).
    Retorna (lista_de_ids_validos, mensagem_de_erro_ou_None).
    """
    ids_validos = []
    for uid in usuario_ids:
        if not uid:
            continue
        uid = int(uid)
        conflitos = Cliente.objects.filter(Q(usuario_id=uid) | Q(usuarios_adicionais__id=uid))
        if cliente_id_atual:
            conflitos = conflitos.exclude(id=cliente_id_atual)
        cliente_conflitante = conflitos.first()
        if cliente_conflitante:
            usuario_obj = User.objects.filter(id=uid).first()
            nome_usuario = usuario_obj.username if usuario_obj else uid
            return None, (
                f'Erro: O usuário "{nome_usuario}" já está vinculado ao cliente '
                f'"{cliente_conflitante.nome_empresa}".'
            )
        ids_validos.append(uid)
    return ids_validos, None


@login_required(login_url='login')
@backoffice_required
def cadastrar_cliente(request):
    from usuario.perms import is_admin, get_instancia, usuarios_gerenciaveis_por

    if request.method == 'GET':
        clientes = Cliente.objects.visiveis_para(request.user).select_related('usuario__totp_device').prefetch_related('usuarios_adicionais')
        for c in clientes:
            c.usuario_tem_2fa = bool(c.usuario_id and getattr(c.usuario, 'totp_device', None) and c.usuario.totp_device.confirmado)
        if is_admin(request.user):
            usuario = User.objects.all()
        else:
            # Consultor/Operador só podem vincular ao portal do cliente um
            # login que eles mesmos gerenciam (e que ainda não está
            # vinculado a nenhum outro cliente) — evita "roubar" o login de
            # um usuário de fora da instância.
            usuario = usuarios_gerenciaveis_por(request.user).filter(
                cliente__isnull=True, clientes_adicionais__isnull=True
            )
        return render(request, 'cadastrar_cliente.html', {
            'clientes': clientes, 'usuario': usuario})

    elif request.method == 'POST':
        nome_empresa = request.POST.get('nome_empresa')
        email = request.POST.get('email')
        cnpj = request.POST.get('cnpj')
        telefone = request.POST.get('telefone')
        endereco = request.POST.get('endereco')
        cidade = request.POST.get('cidade')
        estado = request.POST.get('estado')
        cep = request.POST.get('cep')
        # Vincular um usuário de login é opcional — cliente pode ser cadastrado
        # e gerido só pela equipe interna, sem acesso próprio ao portal.
        # O campo "usuario" é um input hidden preenchido via JS (dropdown de
        # busca); o atributo HTML `required` não é validado pelo navegador em
        # inputs hidden, então tratamos '' explicitamente como "sem usuário"
        # em vez de deixar chegar no ORM (que rejeitaria '' como PK inválida).
        usuario_id = request.POST.get('usuario') or None
        if usuario_id is not None and not usuario_id.isdigit():
            messages.error(request, 'Erro: Usuário selecionado é inválido.')
            return redirect('cadastrar_cliente')
        usuarios_adicionais_ids = request.POST.getlist('usuarios_adicionais')

        # Consultor/Operador: cliente sempre nasce na própria instância, não
        # é escolha do formulário. Administrador pode opcionalmente atribuir
        # o cliente a uma instância (senão fica "da plataforma").
        if is_admin(request.user):
            instancia_id = request.POST.get('instancia') or None
        else:
            instancia = get_instancia(request.user)
            if not instancia:
                messages.error(request, 'Sua conta não está vinculada a nenhuma instância.')
                return redirect('cadastrar_cliente')
            instancia_id = instancia.id

            # Blindagem contra manipulação direta do POST: só é permitido
            # vincular um login que o próprio Consultor/Operador já gerencia
            # e que ainda não está vinculado a outro cliente.
            if usuario_id and not usuarios_gerenciaveis_por(request.user).filter(
                id=usuario_id, cliente__isnull=True, clientes_adicionais__isnull=True
            ).exists():
                messages.error(request, 'Erro: Usuário selecionado é inválido.')
                return redirect('cadastrar_cliente')

        # Verifica se o email ou telefone já estão cadastrados
        if Cliente.objects.filter(email=email).exists():
            messages.error(request, 'Erro: Já existe um cliente com esse email cadastrado.')
            return redirect('cadastrar_cliente')

        # ✅ NOVA VALIDAÇÃO: Verifica se o usuário já está vinculado a outro cliente
        if usuario_id and Cliente.objects.filter(usuario_id=usuario_id).exists():
            messages.error(request, 'Erro: Este usuário já está vinculado a outro cliente.')
            return redirect('cadastrar_cliente')

        # Usuário principal não pode se repetir na lista de adicionais
        usuarios_adicionais_ids = [
            uid for uid in usuarios_adicionais_ids
            if uid and (not usuario_id or int(uid) != int(usuario_id))
        ]
        ids_validos, erro = _validar_usuarios_adicionais(usuarios_adicionais_ids)
        if erro:
            messages.error(request, erro)
            return redirect('cadastrar_cliente')

        cliente = Cliente(
            nome_empresa=nome_empresa,
            email=email,
            telefone=telefone,
            endereco=endereco,
            cidade=cidade,
            estado=estado,
            cep=cep,
            cnpj=cnpj,
            usuario_id=usuario_id,
            instancia_id=instancia_id,
        )
        cliente.save()
        cliente.usuarios_adicionais.set(ids_validos)
        messages.success(request, 'Cliente cadastrado com sucesso!')
        return redirect('cadastrar_cliente')



@login_required(login_url='login')
@require_http_methods(['POST'])
@modulo_habilitado_required('acessos')
def importar_acessos_crt(request, cliente_id):
    """Importa hosts a partir de um arquivo XML de backup do SecureCRT."""
    import xml.etree.ElementTree as ET

    PROTO_MAP = {
        'SSH2': 'SSH', 'SSH1': 'SSH', 'SSH': 'SSH',
        'Telnet': 'TELNET', 'TELNET': 'TELNET',
        'HTTP': 'HTTP', 'HTTPS': 'HTTPS',
        'FTP': 'FTP', 'FTPS': 'FTPS',
        'WINBOX': 'WINBOX',
    }
    DEFAULT_PORTS = {
        'SSH': 22, 'TELNET': 23, 'HTTP': 80, 'HTTPS': 443,
        'FTP': 21, 'FTPS': 990, 'WINBOX': 8291,
    }

    def _str(el, attr):
        for ch in el:
            if ch.tag == 'string' and ch.get('name') == attr:
                return (ch.text or '').strip()
        return None

    def _dword(el, attr):
        for ch in el:
            if ch.tag == 'dword' and ch.get('name') == attr:
                return (ch.text or '').strip()
        return None

    def _str_in(el, attr):
        conn = el.find('key[@name="Connection"]')
        return _str(conn, attr) if conn is not None else None

    def _dword_in(el, attr):
        conn = el.find('key[@name="Connection"]')
        return _dword(conn, attr) if conn is not None else None

    def _recurse(element, path, out):
        for key in element.findall('key'):
            name = key.get('name', '')
            if name in ('Options',):
                continue
            hostname = _str(key, 'Hostname') or _str_in(key, 'Hostname')
            if hostname:
                proto_raw = _str(key, 'Protocol Name') or _str_in(key, 'Protocol Name') or 'SSH2'
                protocolo = PROTO_MAP.get(proto_raw, 'SSH')
                porta_str = _dword(key, 'Port') or _dword_in(key, 'Port') or ''
                try:
                    porta = int(porta_str)
                except (ValueError, TypeError):
                    porta = DEFAULT_PORTS.get(protocolo, 22)
                usuario = _str(key, 'Username') or _str_in(key, 'Username') or ''
                # path[0] = grupo (ignorar), path[1:] = função(ões), name = nome do host
                grupo = path[0] if path else ''
                funcao_nome = ' / '.join(path[1:]) if len(path) > 1 else ''
                out.append({
                    'nome': name,
                    'grupo': grupo,
                    'funcao_nome': funcao_nome,
                    'host': hostname,
                    'protocolo': protocolo,
                    'porta': porta,
                    'usuario': usuario,
                })
            else:
                _recurse(key, path + [name] if name != 'Sessions' else [], out)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    is_json = 'application/json' in (request.content_type or '')

    if is_json:
        # ── IMPORT: body JSON com lista de sessões selecionadas ──────────
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON inválido.'})

        sessions = data.get('sessions', [])
        created, skipped, errors = [], [], []

        # Cache de funções para evitar N queries
        funcao_cache = {}
        def _resolve_funcao(nome_funcao):
            if not nome_funcao:
                return 13
            key = nome_funcao.upper()
            if key in funcao_cache:
                return funcao_cache[key]
            obj = Funcao_equipamento.objects.filter(descricao__iexact=nome_funcao).first()
            fid = obj.id if obj else 13
            funcao_cache[key] = fid
            return fid

        for s in sessions:
            tipo = (s.get('nome') or '').strip()
            host = (s.get('host') or '').strip()
            protocolo = s.get('protocolo', 'SSH')
            porta = s.get('porta') or DEFAULT_PORTS.get(protocolo, 22)
            usuario = (s.get('usuario') or '').strip()
            funcao_nome = (s.get('funcao_nome') or '').strip()
            funcao_id_front = s.get('funcao_id')
            if not tipo or not host:
                errors.append('Sessão sem nome ou host ignorada.')
                continue
            if Acesso.objects.filter(tipo=tipo, cliente_id=cliente_id).exists():
                skipped.append(tipo)
                continue
            # funcao_id escolhida pelo usuário tem prioridade; 0 = sem função → default 13
            if funcao_id_front and int(funcao_id_front) > 0:
                funcao_id_final = int(funcao_id_front)
            else:
                funcao_id_final = _resolve_funcao(funcao_nome)
            senha = (s.get('senha') or '').strip()
            try:
                Acesso.objects.create(
                    cliente_id=cliente_id,
                    funcao_id=funcao_id_final,
                    tipo=tipo,
                    host=host,
                    porta=porta,
                    protocolo=protocolo,
                    usuario=usuario,
                    senha=senha,
                )
                created.append(tipo)
            except Exception as e:
                errors.append(f'{tipo}: {e}')

        return JsonResponse({'success': True, 'created': len(created), 'skipped': len(skipped), 'errors': errors})

    else:
        # ── PARSE: multipart com arquivo XML ────────────────────────────
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado.'})
        if not arquivo.name.lower().endswith('.xml'):
            return JsonResponse({'success': False, 'error': 'Apenas arquivos .xml são aceitos.'})
        try:
            tree = ET.parse(arquivo)
            root = tree.getroot()
        except ET.ParseError as e:
            return JsonResponse({'success': False, 'error': f'XML inválido: {e}'})

        sessions = []
        sessions_key = root.find('.//key[@name="Sessions"]')
        _recurse(sessions_key if sessions_key is not None else root, [], sessions)

        return JsonResponse({'success': True, 'sessions': sessions, 'total': len(sessions)})


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def template_excel_acessos(request, cliente_id):
    """Gera planilha Excel modelo para importação de acessos."""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    cliente = get_object_or_404(Cliente, id=cliente_id)
    funcoes = list(Funcao_equipamento.objects.all().order_by('descricao'))
    modelos = list(Modelo_equipamento.objects.all().order_by('fabricante', 'nome'))

    wb = openpyxl.Workbook()

    # ── Aba principal ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Acessos'

    # Estilos
    header_fill  = PatternFill('solid', fgColor='1e3a5f')
    required_fill = PatternFill('solid', fgColor='2d1e0f')
    optional_fill = PatternFill('solid', fgColor='1a2a1a')
    header_font  = Font(bold=True, color='FFFFFF', size=10)
    req_font     = Font(bold=True, color='ffa500', size=10)
    opt_font     = Font(bold=True, color='90ee90', size=10)
    center       = Alignment(horizontal='center', vertical='center')
    thin         = Border(
        left=Side(style='thin', color='2a3344'),
        right=Side(style='thin', color='2a3344'),
        top=Side(style='thin', color='2a3344'),
        bottom=Side(style='thin', color='2a3344'),
    )

    COLUNAS = [
        ('Nome do Host (tipo)',  'tipo',       True,  30),
        ('IP / Hostname',        'host',        True,  22),
        ('Protocolo',            'protocolo',   True,  12),
        ('Porta',                'porta',        True,  8),
        ('Usuário',              'usuario',      False, 16),
        ('Senha',                'senha',        False, 16),
        ('Senha ADM',            'senha_adm',    False, 16),
        ('VLAN',                 'vlan',         False, 8),
        ('Função',               'funcao',       False, 22),
        ('Modelo',               'modelo',       False, 28),
        ('Notas Agent NOC',      'notas',        False, 34),
    ]

    # Linha 1: legenda de cores
    ws.merge_cells('A1:K1')
    leg = ws['A1']
    leg.value = (
        f'Planilha de importação — {cliente.nome_empresa}  '
        '|  Laranja = obrigatório  |  Verde = opcional'
    )
    leg.font     = Font(color='cccccc', italic=True, size=9)
    leg.fill     = PatternFill('solid', fgColor='0d1117')
    leg.alignment = center
    ws.row_dimensions[1].height = 18

    # Linha 2: cabeçalhos
    for col_idx, (label, _field, required, width) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill      = required_fill if required else optional_fill
        cell.font      = req_font if required else opt_font
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    # Linha 3: linha de exemplo
    EXEMPLO = [
        'SW-CLIENTE-01', '192.168.0.1', 'SSH', '22',
        'admin', 'senha123', 'adm456', '',
        funcoes[0].descricao if funcoes else '',
        f'{modelos[0].fabricante} {modelos[0].nome}' if modelos else '',
        'Observações extras para o Agent NOC',
    ]
    for col_idx, val in enumerate(EXEMPLO, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font      = Font(color='888888', italic=True, size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin
    ws.row_dimensions[3].height = 18

    # Linhas de dados (4–203 = 200 linhas)
    DATA_START = 4
    DATA_END   = 203
    for row in range(DATA_START, DATA_END + 1):
        for col_idx in range(1, len(COLUNAS) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill   = PatternFill('solid', fgColor='0d1520')
            cell.font   = Font(color='e2e8f0', size=9)
            cell.border = thin

    # Congelar cabeçalho
    ws.freeze_panes = 'A4'

    # ── Validação: Protocolo (coluna C = 3) ───────────────────────
    dv_proto = DataValidation(
        type='list',
        formula1='"SSH,TELNET,HTTP,HTTPS,WINBOX,FTP,FTPS"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_proto.error       = 'Use: SSH, TELNET, HTTP, HTTPS, WINBOX, FTP, FTPS'
    dv_proto.errorTitle  = 'Protocolo inválido'
    dv_proto.showErrorMessage = True
    ws.add_data_validation(dv_proto)
    dv_proto.add(f'C{DATA_START}:C{DATA_END}')

    # ── Aba auxiliar: Funções ──────────────────────────────────────
    ws_func = wb.create_sheet('_Funcoes')
    ws_func.sheet_state = 'hidden'
    for i, f in enumerate(funcoes, start=1):
        ws_func.cell(row=i, column=1, value=f.descricao)

    if funcoes:
        last_func = len(funcoes)
        dv_func = DataValidation(
            type='list',
            formula1=f'_Funcoes!$A$1:$A${last_func}',
            allow_blank=True,
            showDropDown=False,
        )
        dv_func.showErrorMessage = False
        ws.add_data_validation(dv_func)
        dv_func.add(f'I{DATA_START}:I{DATA_END}')

    # ── Aba auxiliar: Modelos ──────────────────────────────────────
    ws_mod = wb.create_sheet('_Modelos')
    ws_mod.sheet_state = 'hidden'
    for i, m in enumerate(modelos, start=1):
        ws_mod.cell(row=i, column=1, value=f'{m.fabricante} — {m.nome}')

    if modelos:
        last_mod = len(modelos)
        dv_mod = DataValidation(
            type='list',
            formula1=f'_Modelos!$A$1:$A${last_mod}',
            allow_blank=True,
            showDropDown=False,
        )
        dv_mod.showErrorMessage = False
        ws.add_data_validation(dv_mod)
        dv_mod.add(f'J{DATA_START}:J{DATA_END}')

    # ── Aba de referência: Funções visíveis ───────────────────────
    ws_ref = wb.create_sheet('Referência')
    ws_ref['A1'] = 'FUNÇÕES DISPONÍVEIS'
    ws_ref['A1'].font = Font(bold=True, color='ffa500', size=11)
    ws_ref['B1'] = 'MODELOS DISPONÍVEIS'
    ws_ref['B1'].font = Font(bold=True, color='90ee90', size=11)
    ws_ref['C1'] = 'FABRICANTE'
    ws_ref['C1'].font = Font(bold=True, color='90ee90', size=11)
    for i, f in enumerate(funcoes, start=2):
        ws_ref.cell(row=i, column=1, value=f.descricao).font = Font(size=9)
    for i, m in enumerate(modelos, start=2):
        ws_ref.cell(row=i, column=2, value=m.nome).font = Font(size=9)
        ws_ref.cell(row=i, column=3, value=m.fabricante).font = Font(size=9, color='888888')
    ws_ref.column_dimensions['A'].width = 22
    ws_ref.column_dimensions['B'].width = 26
    ws_ref.column_dimensions['C'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f'importar_acessos_{cliente.nome_empresa.replace(" ", "_")}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


@login_required(login_url='login')
@require_http_methods(['POST'])
@modulo_habilitado_required('acessos')
def importar_acessos_excel(request, cliente_id):
    """Importa acessos a partir de planilha Excel gerada pelo template."""
    import io
    import openpyxl

    cliente = get_object_or_404(Cliente, id=cliente_id)

    PROTO_VALIDOS = {'SSH', 'TELNET', 'HTTP', 'HTTPS', 'WINBOX', 'FTP', 'FTPS'}
    DEFAULT_PORTS = {
        'SSH': 22, 'TELNET': 23, 'HTTP': 80, 'HTTPS': 443,
        'FTP': 21, 'FTPS': 990, 'WINBOX': 8291,
    }

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado.'})
    if not arquivo.name.lower().endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'error': 'Apenas arquivos .xlsx são aceitos.'})

    senha_padrao = request.POST.get('senha_padrao', '').strip()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(arquivo.read()), data_only=True)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Arquivo inválido: {e}'})

    ws = wb.active

    # Cache de funções e modelos
    funcao_cache = {f.descricao.upper(): f for f in Funcao_equipamento.objects.all()}
    modelo_cache = {}
    for m in Modelo_equipamento.objects.all():
        modelo_cache[f'{m.fabricante} — {m.nome}'.upper()] = m
        modelo_cache[m.nome.upper()] = m

    created, skipped, errors = [], [], []

    # Dados começam na linha 4 (1=legenda, 2=cabeçalho, 3=exemplo)
    for row in ws.iter_rows(min_row=4, values_only=True):
        tipo = str(row[0] or '').strip() if row[0] is not None else ''
        host = str(row[1] or '').strip() if row[1] is not None else ''

        if not tipo or not host:
            continue
        # Pula linha de exemplo (verificação pelo conteúdo)
        if tipo == 'SW-CLIENTE-01' and host == '192.168.0.1':
            continue

        proto_raw = str(row[2] or 'SSH').strip().upper()
        protocolo = proto_raw if proto_raw in PROTO_VALIDOS else 'SSH'

        try:
            porta = int(row[3]) if row[3] is not None else DEFAULT_PORTS.get(protocolo, 22)
        except (ValueError, TypeError):
            porta = DEFAULT_PORTS.get(protocolo, 22)

        usuario  = str(row[4] or '').strip()
        senha    = str(row[5] or '').strip() or senha_padrao
        senha_adm = str(row[6] or '').strip()
        vlan_raw = row[7]
        try:
            vlan = int(vlan_raw) if vlan_raw not in (None, '') else None
        except (ValueError, TypeError):
            vlan = None

        funcao_str = str(row[8] or '').strip()
        modelo_str = str(row[9] or '').strip()
        notas      = str(row[10] or '').strip()

        funcao = funcao_cache.get(funcao_str.upper())
        funcao_id = funcao.id if funcao else 13

        modelo = modelo_cache.get(modelo_str.upper())

        if Acesso.objects.filter(tipo=tipo, cliente_id=cliente_id).exists():
            skipped.append(tipo)
            continue

        try:
            Acesso.objects.create(
                cliente_id=cliente_id,
                funcao_id=funcao_id,
                modelo=modelo,
                tipo=tipo,
                host=host,
                protocolo=protocolo,
                porta=porta,
                usuario=usuario,
                senha=senha,
                senha_adm=senha_adm,
                vlan=vlan,
                notas=notas,
            )
            created.append(tipo)
        except Exception as e:
            errors.append(f'{tipo}: {e}')

    return JsonResponse({
        'success': True,
        'created': len(created),
        'skipped': len(skipped),
        'errors': errors,
    })


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def cadastrar_acesso(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        funcao_id = request.POST.get('funcao')
        modelo_id = request.POST.get('modelo')
        tipo = request.POST.get('tipo')
        host = request.POST.get('hostname')
        host_ipv6 = request.POST.get('hostname_ipv6')
        porta = request.POST.get('porta')
        protocolo = request.POST.get('protocolo')
        usuario = request.POST.get('usuario')
        senha = request.POST.get('senha')
        # Campo só existe no form para is_admin — se ausente (cliente cadastrando/duplicando),
        # usa string vazia em vez de None (o campo não aceita NULL no banco)
        senha_adm = request.POST.get('senha_adm', '')
        vlan = request.POST.get('vlan')
        winbox = request.POST.get('winbox')
        backup_habilitado = request.POST.get('backup_habilitado') == 'on'
        backup_template_id = request.POST.get('backup_template')
        backup_automatico = request.POST.get('backup_automatico') == 'on'

        # ✅ Se funcao_id for vazio ou None, usa o padrão 13
        if not funcao_id or funcao_id == '':
            funcao_id = 13

        # ✅ Tratar VLAN vazia ou inválida
        if vlan == '' or vlan is None:
            vlan = None
        else:
            try:
                vlan = int(vlan)
            except ValueError:
                vlan = None

        # ✅ Tratar WINBOX vazio ou inválido
        if winbox == '' or winbox is None:
            winbox = None
        else:
            try:
                winbox = int(winbox)
            except ValueError:
                winbox = None

        # 🧠 Verifica se já existe um Acesso com o mesmo tipo para o mesmo cliente
        if Acesso.objects.filter(tipo=tipo, cliente_id=cliente_id).exists():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': f'O tipo "{tipo}" já está cadastrado para este cliente.'})
            messages.error(request, f'O tipo "{tipo}" já está cadastrado para este cliente.')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # ✅ Cria o registro normalmente
        acesso = Acesso(
            cliente_id=cliente_id,
            funcao_id=funcao_id,  # ✅ Agora sempre terá um valor (13 por padrão)
            modelo_id=modelo_id,
            tipo=tipo,
            host=host,
            host_ipv6=host_ipv6,
            porta=porta,
            protocolo=protocolo,
            usuario=usuario,
            senha=senha,
            senha_adm=senha_adm,
            vlan=vlan,
            winbox=winbox,
            backup_habilitado=backup_habilitado,
            backup_template_id=backup_template_id if backup_template_id else None,
            backup_automatico=backup_automatico
        )
        acesso.save()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        messages.success(request, 'Acesso cadastrado com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    else:
        return redirect('cadastrar_cliente')


@login_required(login_url='login')
@backoffice_required
def editar_cliente(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('id')
        cliente = get_object_or_404(Cliente, id=cliente_id)

        if not _perms.pode_acessar_cliente(request.user, cliente):
            messages.error(request, 'Você não possui permissão para editar este cliente.')
            return redirect('cadastrar_cliente')

        email = request.POST.get('email')
        telefone = request.POST.get('telefone')

        # Verifica se email já existe em outro cliente
        if Cliente.objects.filter(email=email).exclude(id=cliente_id).exists():
            messages.error(request, 'Erro: Já existe um cliente com esse email cadastrado.')
            return redirect('cadastrar_cliente')

        # Vincular usuário de login é opcional (ver Cliente.usuario). O hidden
        # "usuario" (dropdown de busca via JS) chega como '' quando nada foi
        # selecionado — tratamos como "remover vínculo" em vez de rejeitar.
        usuario_id = request.POST.get('usuario') or None
        if usuario_id is not None and not usuario_id.isdigit():
            messages.error(request, 'Erro: Usuário selecionado é inválido.')
            return redirect('cadastrar_cliente')

        # Atualiza os dados
        cliente.nome_empresa = request.POST.get('nome_empresa')
        cliente.cnpj = request.POST.get('cnpj')
        cliente.cep = request.POST.get('cep')
        cliente.endereco = request.POST.get('endereco')
        cliente.estado = request.POST.get('estado')
        cliente.cidade = request.POST.get('cidade')
        cliente.telefone = telefone
        cliente.email = email
        cliente.usuario_id = usuario_id
        cliente.notas = request.POST.get('notas', '').strip()

        # Usuários adicionais: mesma validação de vínculo único do cadastro
        usuarios_adicionais_ids = request.POST.getlist('usuarios_adicionais')
        usuario_principal_id = cliente.usuario_id
        usuarios_adicionais_ids = [
            uid for uid in usuarios_adicionais_ids
            if uid and (not usuario_principal_id or int(uid) != int(usuario_principal_id))
        ]
        ids_validos, erro = _validar_usuarios_adicionais(usuarios_adicionais_ids, cliente_id_atual=cliente.id)
        if erro:
            messages.error(request, erro)
            return redirect('cadastrar_cliente')

        cliente.save()
        cliente.usuarios_adicionais.set(ids_validos)
        messages.success(request, "Cliente atualizado com sucesso!")
        return redirect('cadastrar_cliente')

    messages.error(request, "Método não permitido.")
    return redirect('cadastrar_cliente')


@login_required(login_url='login')
@backoffice_required
def deletar_cliente(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('id')
        cliente = get_object_or_404(Cliente, id=cliente_id)

        if not _perms.pode_acessar_cliente(request.user, cliente):
            messages.error(request, 'Você não possui permissão para excluir este cliente.')
            return redirect('cadastrar_cliente')

        nome_empresa = cliente.nome_empresa

        # Deleta o cliente (os acessos relacionados serão deletados automaticamente se houver CASCADE)
        cliente.delete()

        messages.success(request, f'Cliente "{nome_empresa}" excluído com sucesso!')
        return redirect('cadastrar_cliente')

    messages.error(request, 'Método não permitido.')
    return redirect('cadastrar_cliente')


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def buscar_acesso(request, acesso_id):
    """
    Validar se cliente pode acessar este acesso
    """
    try:
        acesso = Acesso.objects.get(id=acesso_id)

        # ✅ Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
            return JsonResponse({'error': 'Sem permissão'}, status=403)

        data = {
            'id': acesso.id,
            'tipo': acesso.tipo,
            'host': acesso.host,
            'host_ipv6': acesso.host_ipv6 or '',
            'protocolo': acesso.protocolo,
            'porta': acesso.porta,
            'usuario': acesso.usuario,
            'senha': acesso.senha,
            'senha_adm': (acesso.senha_adm or '') if _perms.is_backoffice(request.user) else '',
            'vlan': acesso.vlan or '',
            'winbox': acesso.winbox or '',
            'funcao_id': acesso.funcao.id if acesso.funcao and hasattr(acesso.funcao, 'id') else '',
            'funcao_nome': acesso.funcao.descricao if acesso.funcao and hasattr(acesso.funcao, 'descricao') else '',
            'modelo_id': acesso.modelo.id if acesso.modelo and hasattr(acesso.modelo, 'id') else '',
            'modelo_nome': acesso.modelo.nome if acesso.modelo and hasattr(acesso.modelo, 'nome') else '',
            'backup_habilitado': acesso.backup_habilitado,
            'backup_template_id': acesso.backup_template.id if acesso.backup_template else '',
            'backup_template_nome': acesso.backup_template.nome if acesso.backup_template else '',
            'backup_automatico': acesso.backup_automatico,
            'notas': acesso.notas or '',
        }

        return JsonResponse(data)

    except Acesso.DoesNotExist:
        return JsonResponse({'error': 'Acesso não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def editar_acesso(request, acesso_id):
    if request.method == 'POST':
        try:
            acesso = get_object_or_404(Acesso, id=acesso_id)

            # Atualiza campos diretos
            acesso.tipo = request.POST.get('tipo')
            acesso.host = request.POST.get('hostname')
            acesso.host_ipv6 = request.POST.get('hostname_ipv6')
            acesso.protocolo = request.POST.get('protocolo')
            acesso.porta = request.POST.get('porta')
            acesso.usuario = request.POST.get('usuario')
            acesso.senha = request.POST.get('senha')
            # Campo só existe no form para is_admin — se ausente (cliente editando),
            # preserva o valor atual em vez de apagar (o campo não aceita NULL no banco)
            acesso.senha_adm = request.POST.get('senha_adm', acesso.senha_adm)
            acesso.backup_habilitado = request.POST.get('backup_habilitado') == 'on'
            template_id = request.POST.get('backup_template')
            acesso.backup_template_id = template_id if template_id else None
            acesso.backup_automatico = request.POST.get('backup_automatico') == 'on'
            acesso.notas = request.POST.get('notas', '').strip()

            # ✅ Tratar WINBOX vazio ou inválido
            winbox = request.POST.get('winbox')
            if winbox == '' or winbox is None:
                acesso.winbox = None
            else:
                try:
                    acesso.winbox = int(winbox)
                except ValueError:
                    acesso.winbox = None  # evita erro se o campo não for numérico

            # ✅ Tratar VLAN vazia ou inválida
            vlan = request.POST.get('vlan')
            if vlan == '' or vlan is None:
                acesso.vlan = None
            else:
                try:
                    acesso.vlan = int(vlan)
                except ValueError:
                    acesso.vlan = None  # evita erro se o campo não for numérico

            # ✅ Atualizar função e modelo apenas se enviados
            funcao_id = request.POST.get('funcao')
            modelo_id = request.POST.get('modelo')

            if funcao_id:
                acesso.funcao = get_object_or_404(Funcao_equipamento, id=funcao_id)
            else:
                acesso.funcao = None

            if modelo_id:
                acesso.modelo = get_object_or_404(Modelo_equipamento, id=modelo_id)
            else:
                acesso.modelo = None

            acesso.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            messages.success(request, 'Acesso atualizado com sucesso!')
            return redirect(f"{reverse('listar_clientes')}?id={acesso.cliente.id}")

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Erro ao editar acesso: {str(e)}')
            return redirect(f"{reverse('listar_clientes')}?id={acesso.cliente.id}")

    return redirect('listar_clientes')



@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def deletar_acesso(request, acesso_id):
    acesso = get_object_or_404(Acesso, id=acesso_id)
    cliente_id = acesso.cliente.id
    tipo_acesso = acesso.tipo

    acesso.delete()

    messages.success(request, f'Acesso "{tipo_acesso}" excluído com sucesso!')
    return redirect(f"{reverse('listar_clientes')}?id={cliente_id}")



@login_required(login_url='login')
@modulo_habilitado_required('documentos')
def upload_documento(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        arquivo = request.FILES.get('arquivo')
        nome = arquivo.name if arquivo else None

        if not arquivo:
            messages.error(request, "Nenhum arquivo selecionado.")
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        Documento.objects.create(
            cliente_id=cliente_id,
            nome=nome,
            arquivo=arquivo
        )
        messages.success(request, f'Documento "{nome}" enviado com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')
    else:
        return redirect('listar_clientes')

@login_required(login_url='login')
@modulo_habilitado_required('documentos')
def deletar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    cliente_id = documento.cliente.id

    # Deleta o arquivo do disco também
    if documento.arquivo and documento.arquivo.storage.exists(documento.arquivo.name):
        documento.arquivo.delete(save=False)

    documento.delete()
    messages.success(request, f'Documento "{documento.nome}" excluído com sucesso!')
    return redirect(reverse('listar_clientes') + f'?id={cliente_id}')


    # ========================================
# VIEWS PARA VPN
# ========================================

@login_required(login_url='login')
@modulo_habilitado_required('vpn')
def upload_vpn(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        arquivo = request.FILES.get('arquivo')
        nome = arquivo.name if arquivo else None
        usuario = request.POST.get('usuario')
        senha = request.POST.get('senha')
        private_key = request.POST.get('private_key')

        if not arquivo:
            messages.error(request, "Nenhum arquivo selecionado.")
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        ArquivoVPN.objects.create(
            cliente_id=cliente_id,
            nome=nome,
            arquivo=arquivo,
            usuario=usuario,
            senha=senha,
            private_key=private_key
        )
        messages.success(request, f'Arquivo VPN "{nome}" enviado com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')
    else:
        return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('vpn')
def deletar_vpn(request, vpn_id):
    vpn = get_object_or_404(ArquivoVPN, id=vpn_id)
    cliente_id = vpn.cliente.id

    # Deleta o arquivo do disco também
    if vpn.arquivo and vpn.arquivo.storage.exists(vpn.arquivo.name):
        vpn.arquivo.delete(save=False)

    vpn.delete()
    messages.success(request, f'Arquivo VPN "{vpn.nome}" excluído com sucesso!')
    return redirect(reverse('listar_clientes') + f'?id={cliente_id}')


# ========================================
# VIEWS PARA TOPOLOGIA
# ========================================

@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def upload_topologia(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        imagem = request.FILES.get('imagem')
        nome = imagem.name if imagem else None
        drawio_url = request.POST.get('drawio_url', '').strip()  # ✅ NOVO

        if not imagem:
            messages.error(request, "Nenhuma imagem selecionada.")
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # Validar se é uma imagem
        valid_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.svg', '.webp']
        if not any(nome.lower().endswith(ext) for ext in valid_extensions):
            messages.error(request, "Apenas imagens são permitidas (JPG, PNG, GIF, SVG, WEBP).")
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # ✅ NOVO: Criar com drawio_url
        ImagemTopologia.objects.create(
            cliente_id=cliente_id,
            nome=nome,
            imagem=imagem,
            drawio_url=drawio_url if drawio_url else None  # Aceita vazio
        )
        messages.success(request, f'Imagem de topologia "{nome}" enviada com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')
    else:
        return redirect('listar_clientes')

@login_required(login_url='login')
def editar_topologia(request, topologia_id):
    """Edita o link DrawIO de uma topologia"""
    if request.method == 'POST':
        topologia = get_object_or_404(ImagemTopologia, id=topologia_id)

        topologia.nome = request.POST.get('nome', topologia.nome)
        topologia.drawio_url = request.POST.get('drawio_url', '').strip()
        topologia.drawio_url = topologia.drawio_url if topologia.drawio_url else None

        topologia.save()

        messages.success(request, 'Topologia atualizada com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={topologia.cliente.id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def deletar_topologia(request, topologia_id):
    topologia = get_object_or_404(ImagemTopologia, id=topologia_id)
    cliente_id = topologia.cliente.id

    # Deleta a imagem do disco também
    if topologia.imagem and topologia.imagem.storage.exists(topologia.imagem.name):
        topologia.imagem.delete(save=False)

    topologia.delete()
    messages.success(request, f'Imagem de topologia "{topologia.nome}" excluída com sucesso!')
    return redirect(reverse('listar_clientes') + f'?id={cliente_id}')


@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def editar_imagem_topologia(request, topologia_id):
    """Edita a imagem de uma topologia (substitui a imagem existente)"""
    if request.method == 'POST':
        topologia = get_object_or_404(ImagemTopologia, id=topologia_id)
        cliente_id = topologia.cliente.id

        # ✅ Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, topologia.cliente):
            messages.error(request, 'Sem permissão para editar esta topologia.')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # ✅ Obter a nova imagem
        imagem = request.FILES.get('imagem')

        if not imagem:
            messages.error(request, 'Nenhuma imagem selecionada.')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # ✅ Validar se é imagem
        nome_arquivo = imagem.name.lower()
        valid_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.svg', '.webp']

        if not any(nome_arquivo.endswith(ext) for ext in valid_extensions):
            messages.error(request, 'Apenas imagens são permitidas (JPG, PNG, GIF, SVG, WEBP).')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # ✅ Deletar imagem antiga se existir
        if topologia.imagem and topologia.imagem.storage.exists(topologia.imagem.name):
            topologia.imagem.delete(save=False)

        # ✅ Atualizar com nova imagem
        topologia.imagem = imagem
        topologia.nome = request.POST.get('nome', imagem.name)
        topologia.save()

        messages.success(request, f'Imagem de topologia atualizada com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('vpn')
def buscar_vpn(request, vpn_id):
    try:
        vpn = ArquivoVPN.objects.get(id=vpn_id)

        data = {
            'id': vpn.id,
            'nome': vpn.nome,
            'usuario': vpn.usuario or '',
            'senha': vpn.senha or '',
            'private_key': vpn.private_key or '',
        }

        return JsonResponse(data)

    except ArquivoVPN.DoesNotExist:
        return JsonResponse({'error': 'VPN não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
@modulo_habilitado_required('vpn')
def editar_vpn(request, vpn_id):
    if request.method == 'POST':
        try:
            vpn = get_object_or_404(ArquivoVPN, id=vpn_id)

            vpn.usuario = request.POST.get('usuario')
            vpn.senha = request.POST.get('senha')
            vpn.private_key = request.POST.get('private_key')

            vpn.save()

            messages.success(request, 'Configuração VPN atualizada com sucesso!')
            return redirect(f"{reverse('listar_clientes')}?id={vpn.cliente.id}")
        except Exception as e:
            messages.error(request, f'Erro ao editar VPN: {str(e)}')
            return redirect(f"{reverse('listar_clientes')}?id={vpn.cliente.id}")

    return redirect('listar_clientes')




    # ========================================
# VIEWS PARA CATEGORIAS
# ========================================

@login_required(login_url='login')
def cadastrar_categoria(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        descricao = request.POST.get('descricao', '')

        if Categoria.objects.filter(nome__iexact=nome).exists():
            return JsonResponse({'error': 'Categoria já existe'}, status=400)

        categoria = Categoria.objects.create(
            nome=nome,
            descricao=descricao
        )

        return JsonResponse({
            'id': categoria.id,
            'nome': categoria.nome,
            'message': 'Categoria cadastrada com sucesso!'
        })

    return JsonResponse({'error': 'Método não permitido'}, status=405)


@login_required(login_url='login')
def buscar_categorias(request):
    query = request.GET.get('q', '')
    categorias = Categoria.objects.filter(nome__icontains=query)[:10]

    results = [{'id': cat.id, 'nome': cat.nome} for cat in categorias]
    return JsonResponse({'results': results})


@login_required(login_url='login')
def listar_chamados_cliente(request):
    """
    Cliente só pode listar seus próprios chamados
    """
    cliente_id = request.GET.get('id')

    # ✅ Verificar permissão
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    chamados = Chamado.objects.filter(cliente=cliente).select_related(
        'categoria', 'responsavel', 'criado_por'
    ).prefetch_related('comentarios')

    return JsonResponse({
        'chamados': [{
            'id': chamado.id,
            'titulo': chamado.titulo,
            'categoria': chamado.categoria.nome if chamado.categoria else '',
            'prioridade': chamado.get_prioridade_display(),
            'status': chamado.get_status_display(),
            'status_code': chamado.status,
            'departamento': chamado.get_departamento_display(),
            'responsavel': chamado.responsavel.get_full_name() or chamado.responsavel.username if chamado.responsavel else 'Não atribuído',
            'data_criacao': chamado.data_criacao.strftime('%d/%m/%Y %H:%M'),
            'total_comentarios': chamado.comentarios.count()
        } for chamado in chamados]
    })



@login_required(login_url='login')
def cadastrar_chamado(request):
    if request.method == 'POST':
        try:
            cliente_id = request.POST.get('cliente')
            categoria_id = request.POST.get('categoria')
            prioridade = request.POST.get('prioridade')
            departamento = request.POST.get('departamento')
            responsavel_id = request.POST.get('responsavel')
            titulo = request.POST.get('titulo')
            descricao = request.POST.get('descricao')
            comentario_inicial = request.POST.get('comentario', '')

            # Validações
            if not all([cliente_id, prioridade, departamento, titulo, descricao]):
                messages.error(request, 'Preencha todos os campos obrigatórios.')
                return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

            # Criar chamado
            chamado = Chamado.objects.create(
                cliente_id=cliente_id,
                categoria_id=categoria_id if categoria_id else None,
                prioridade=prioridade,
                departamento=departamento,
                responsavel_id=responsavel_id if responsavel_id else None,
                criado_por=request.user,
                titulo=titulo,
                descricao=descricao
            )

            # Adicionar comentário inicial se houver
            if comentario_inicial:
                ComentarioChamado.objects.create(
                    chamado=chamado,
                    usuario=request.user,
                    comentario=comentario_inicial
                )

            messages.success(request, f'Chamado #{chamado.id} cadastrado com sucesso!')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        except Exception as e:
            messages.error(request, f'Erro ao cadastrar chamado: {str(e)}')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
def buscar_chamado(request, chamado_id):
    try:
        chamado = Chamado.objects.select_related(
            'categoria', 'cliente', 'responsavel', 'criado_por'
        ).prefetch_related('comentarios__usuario').get(id=chamado_id)

        data = {
            'id': chamado.id,
            'titulo': chamado.titulo,
            'descricao': chamado.descricao,
            'categoria_id': chamado.categoria.id if chamado.categoria else '',
            'categoria_nome': chamado.categoria.nome if chamado.categoria else '',
            'prioridade': chamado.prioridade,
            'departamento': chamado.departamento,
            'status': chamado.status,
            'responsavel_id': chamado.responsavel.id if chamado.responsavel else '',
            'responsavel_nome': chamado.responsavel.get_full_name() or chamado.responsavel.username if chamado.responsavel else '',
            'cliente_id': chamado.cliente.id,
            'cliente_nome': chamado.cliente.nome_empresa,
            'data_criacao': chamado.data_criacao.strftime('%d/%m/%Y %H:%M'),
            'comentarios': [{
                'id': comentario.id,
                'usuario': comentario.usuario.get_full_name() or comentario.usuario.username,
                'comentario': comentario.comentario,
                'data': comentario.data_criacao.strftime('%d/%m/%Y %H:%M'),
                'is_internal': comentario.is_internal
            } for comentario in chamado.comentarios.all()]
        }

        return JsonResponse(data)

    except Chamado.DoesNotExist:
        return JsonResponse({'error': 'Chamado não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def editar_chamado(request, chamado_id):
    if request.method == 'POST':
        try:
            chamado = get_object_or_404(Chamado, id=chamado_id)

            chamado.titulo = request.POST.get('titulo')
            chamado.descricao = request.POST.get('descricao')
            chamado.prioridade = request.POST.get('prioridade')
            chamado.departamento = request.POST.get('departamento')
            chamado.status = request.POST.get('status')

            categoria_id = request.POST.get('categoria')
            chamado.categoria_id = categoria_id if categoria_id else None

            responsavel_id = request.POST.get('responsavel')
            chamado.responsavel_id = responsavel_id if responsavel_id else None

            chamado.save()

            # Adicionar comentário de atualização se houver
            comentario_novo = request.POST.get('comentario_novo')
            if comentario_novo:
                ComentarioChamado.objects.create(
                    chamado=chamado,
                    usuario=request.user,
                    comentario=comentario_novo
                )

            messages.success(request, f'Chamado #{chamado.id} atualizado com sucesso!')
            return redirect(f"{reverse('listar_clientes')}?id={chamado.cliente.id}")

        except Exception as e:
            messages.error(request, f'Erro ao editar chamado: {str(e)}')
            return redirect('listar_clientes')

    return redirect('listar_clientes')


@login_required(login_url='login')
def deletar_chamado(request, chamado_id):
    if request.method == 'POST':
        chamado = get_object_or_404(Chamado, id=chamado_id)
        cliente_id = chamado.cliente.id
        chamado_numero = chamado.id

        chamado.delete()

        messages.success(request, f'Chamado #{chamado_numero} excluído com sucesso!')
        return redirect(f"{reverse('listar_clientes')}?id={cliente_id}")

    return redirect('listar_clientes')


@login_required(login_url='login')
def adicionar_comentario(request, chamado_id):
    if request.method == 'POST':
        try:
            chamado = get_object_or_404(Chamado, id=chamado_id)
            comentario_texto = request.POST.get('comentario')
            is_internal = request.POST.get('is_internal') == 'true'

            if comentario_texto:
                ComentarioChamado.objects.create(
                    chamado=chamado,
                    usuario=request.user,
                    comentario=comentario_texto,
                    is_internal=is_internal
                )
                messages.success(request, 'Comentário adicionado com sucesso!')
            else:
                messages.error(request, 'O comentário não pode estar vazio.')

            return redirect(f"{reverse('listar_clientes')}?id={chamado.cliente.id}")

        except Exception as e:
            messages.error(request, f'Erro ao adicionar comentário: {str(e)}')
            return redirect('listar_clientes')

    return redirect('listar_clientes')


@login_required(login_url='login')
def buscar_usuarios(request):
    query = request.GET.get('q', '')
    usuarios = User.objects.filter(
        Q(username__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query)
    )[:10]

    results = [{
        'id': user.id,
        'nome': user.get_full_name() or user.username,
        'username': user.username
    } for user in usuarios]

    return JsonResponse({'results': results})


@login_required(login_url='login')
def buscar_clientes_chamado(request):
    query = request.GET.get('q', '')
    clientes = Cliente.objects.visiveis_para(request.user).filter(
        Q(nome_empresa__icontains=query) |
        Q(cnpj__icontains=query)
    )[:10]

    results = [{
        'id': cliente.id,
        'nome': cliente.nome_empresa,
        'cnpj': cliente.cnpj
    } for cliente in clientes]

    return JsonResponse({'results': results})


# ========================================
# VIEWS PARA GERENCIAR SERVIDORES PROXY (POR CLIENTE)
# ========================================

@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def cadastrar_proxy(request):
    """Cadastra um novo servidor proxy para um cliente"""
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        nome = request.POST.get('nome')
        host = request.POST.get('host')
        porta = request.POST.get('porta', 22)
        usuario = request.POST.get('usuario')
        senha = request.POST.get('senha')
        ativo = request.POST.get('ativo') == 'on'

        # Validações básicas
        if not all([cliente_id, nome, host, porta, usuario, senha]):
            messages.error(request, 'Preencha todos os campos obrigatórios.')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # Criar proxy
        try:
            ProxyServer.objects.create(
                cliente_id=cliente_id,
                nome=nome,
                host=host,
                porta=int(porta),
                usuario=usuario,
                senha=senha,
                ativo=ativo
            )
            messages.success(request, f'Túnel SSH "{nome}" cadastrado com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao cadastrar túnel: {str(e)}')

        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def buscar_proxy(request, proxy_id):
    """Busca dados de um proxy específico (AJAX)"""
    try:
        proxy = ProxyServer.objects.get(id=proxy_id)

        data = {
            'id': proxy.id,
            'nome': proxy.nome,
            'host': proxy.host,
            'porta': proxy.porta,
            'usuario': proxy.usuario,
            'senha': proxy.senha,
            'ativo': proxy.ativo,
            'data_criacao': proxy.data_criacao.strftime('%d/%m/%Y %H:%M')
        }

        return JsonResponse(data)

    except ProxyServer.DoesNotExist:
        return JsonResponse({'error': 'Túnel SSH não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def editar_proxy(request, proxy_id):
    """Edita um servidor proxy existente"""
    if request.method == 'POST':
        try:
            proxy = get_object_or_404(ProxyServer, id=proxy_id)

            proxy.nome = request.POST.get('nome', proxy.nome)
            proxy.host = request.POST.get('host', proxy.host)
            proxy.porta = int(request.POST.get('porta', 22))
            proxy.usuario = request.POST.get('usuario', proxy.usuario)
            proxy.senha = request.POST.get('senha', proxy.senha)
            proxy.ativo = request.POST.get('ativo') == 'on'

            proxy.save()

            messages.success(request, f'Túnel SSH "{proxy.nome}" atualizado com sucesso!')
            return redirect(reverse('listar_clientes') + f'?id={proxy.cliente.id}')

        except Exception as e:
            messages.error(request, f'Erro ao editar túnel: {str(e)}')
            return redirect('listar_clientes')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def deletar_proxy(request, proxy_id):
    """Deleta um servidor proxy"""
    if request.method == 'POST':
        proxy = get_object_or_404(ProxyServer, id=proxy_id)
        cliente_id = proxy.cliente.id
        nome = proxy.nome

        proxy.delete()

        messages.success(request, f'Túnel SSH "{nome}" excluído com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def testar_proxy(request, proxy_id):
    """Testa a conexão com um servidor proxy (AJAX)"""
    try:
        proxy = ProxyServer.objects.get(id=proxy_id)

        import paramiko

        # Tentar conectar ao proxy
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        ssh_client.connect(
            hostname=proxy.host,
            port=proxy.porta,
            username=proxy.usuario,
            password=proxy.senha,
            timeout=5,
            look_for_keys=False,
            allow_agent=False
        )

        ssh_client.close()

        return JsonResponse({
            'success': True,
            'message': f'✓ Conexão com túnel "{proxy.nome}" bem-sucedida!'
        })

    except paramiko.AuthenticationException:
        return JsonResponse({
            'success': False,
            'message': '✗ Erro de autenticação. Verifique usuário e senha.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'✗ Erro ao conectar: {str(e)}'
        }, status=400)


@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def toggle_proxy_status(request, proxy_id):
    """Ativa/Desativa um servidor proxy (AJAX)"""
    try:
        proxy = ProxyServer.objects.get(id=proxy_id)
        proxy.ativo = not proxy.ativo
        proxy.save()

        status_texto = 'ativado' if proxy.ativo else 'desativado'

        return JsonResponse({
            'success': True,
            'ativo': proxy.ativo,
            'message': f'Túnel SSH "{proxy.nome}" {status_texto} com sucesso!'
        })

    except ProxyServer.DoesNotExist:
        return JsonResponse({'error': 'Túnel SSH não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def cliente_dashboard(request):
    """
    Dashboard exclusivo para clientes
    - Ferramentas de rede
    - Chamados abertos
    - Link para acessos
    """
    if not request.user.is_authenticated:
        return redirect('login')

    # Se for admin/consultor/operador, redireciona para o dashboard correto
    if _perms.is_admin(request.user):
        return redirect('quadro_geral')
    if _perms.is_backoffice(request.user):
        return redirect('cadastrar_cliente')

    # Buscar cliente vinculado
    try:
        cliente = Cliente.objects.get_by_usuario_vinculado(request.user)
    except Cliente.DoesNotExist:
        messages.error(request, 'Você não está vinculado a um cliente.')
        return redirect('login')

    # Buscar chamados abertos do cliente
    chamados_abertos = Chamado.objects.filter(
        cliente=cliente,
        status__in=['aberto', 'em_andamento']
    ).order_by('-data_criacao')[:5]

    return render(request, 'cliente_dashboard.html', {
        'cliente': cliente,
        'chamados_abertos': chamados_abertos,
    })

@login_required(login_url='login')
@modulo_habilitado_required('backups')
def executar_backup_acesso(request, acesso_id):
    """Executa backup manual"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Não autenticado'}, status=401)

    try:
        acesso = Acesso.objects.get(id=acesso_id)

        # Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
            return JsonResponse({'error': 'Sem permissão'}, status=403)

        # Verificar se backup está habilitado
        if not acesso.backup_habilitado:
            return JsonResponse({
                'error': 'Backup não está habilitado para este acesso'
            }, status=400)

        if not acesso.backup_template:
            return JsonResponse({
                'error': 'Template de backup não configurado'
            }, status=400)

        # Executar backup
        resultado = realizar_backup(acesso, request.user)

        if resultado['sucesso']:
            if resultado.get('sem_mudancas'):
                return JsonResponse({
                    'success': True,
                    'sem_mudancas': True,
                    'message': 'Configuracao sem alteracoes — backup nao necessario.',
                    'duracao': resultado['duracao'],
                })
            return JsonResponse({
                'success': True,
                'sem_mudancas': False,
                'message': 'Backup realizado com sucesso!',
                'arquivo': resultado['arquivo'],
                'tamanho': resultado['tamanho'],
                'duracao': resultado['duracao'],
            })
        else:
            return JsonResponse({
                'error': resultado['erro']
            }, status=500)

    except Acesso.DoesNotExist:
        return JsonResponse({'error': 'Acesso não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def _executar_comandos_pexpect(host_conexao, porta_conexao, usuario, senha, comandos):
    """
    Usa pexpect (SSH real) para OLTs com firmware SSH limitado (Parks FIBERLINK).
    Paramiko invoke_shell não é compatível — o firmware fecha o canal imediatamente.
    pexpect spawna o binário ssh do sistema, que negocia corretamente com o OLT.
    """
    ssh_cmd = (
        f"ssh -o StrictHostKeyChecking=no "
        f"-o UserKnownHostsFile=/dev/null "
        f"-o IdentitiesOnly=yes "
        f"-o PubkeyAuthentication=no "
        f"-o PreferredAuthentications=password "
        f"-o ConnectTimeout=15 "
        f"-o ServerAliveInterval=10 "
        f"-o ServerAliveCountMax=3 "
        f"-o LogLevel=ERROR "
        f"-o NumberOfPasswordPrompts=1 "
        f"-o KexAlgorithms=+diffie-hellman-group-exchange-sha1,diffie-hellman-group14-sha1,diffie-hellman-group1-sha1 "
        f"-o HostKeyAlgorithms=+ssh-rsa,ssh-dss "
        f"-o PubkeyAcceptedAlgorithms=+ssh-rsa "
        f"-o Ciphers=+aes128-cbc,aes192-cbc,aes256-cbc,3des-cbc "
        f"-o MACs=+hmac-sha1,hmac-sha2-256,hmac-sha2-512 "
        f"-p {porta_conexao} {usuario}@{host_conexao}"
    )

    print(f"    🔧 pexpect SSH: {usuario}@{host_conexao}:{porta_conexao}")

    process = pexpect.spawn(ssh_cmd, timeout=30, encoding=None, maxread=262144)
    process.setwinsize(50, 200)

    # ✅ Autenticar
    index = process.expect([
        b"password:",
        b"Password:",
        rb".*[#>$\]].*",
        pexpect.TIMEOUT,
        pexpect.EOF
    ], timeout=15)

    if index in (0, 1):
        print(f"    🔐 Enviando senha...")
        process.sendline(senha.encode() if isinstance(senha, str) else senha)
        time.sleep(1)
        try:
            process.expect([rb".*[#>$\]].*", pexpect.TIMEOUT], timeout=10)
        except:
            pass
    elif index == 3:
        raise Exception("Timeout ao aguardar prompt de autenticação SSH")
    elif index == 4:
        raise Exception("Conexão SSH encerrada antes de autenticar")

    print(f"    ✅ Autenticado via pexpect")

    # ✅ Desabilitar paginação
    process.send(b'terminal length 0\r')
    time.sleep(1)
    try:
        process.read_nonblocking(size=65536, timeout=2)
    except:
        pass

    output = ""

    for i, comando in enumerate(comandos, 1):
        if 'terminal length' in comando.lower():
            print(f"  [{i}/{len(comandos)}] {comando} (já enviado — ignorando)")
            continue

        print(f"  [{i}/{len(comandos)}] {comando}")
        output += f"\n{'='*60}\nComando: {comando}\n{'='*60}\n"

        try:
            process.send(comando.encode() + b'\r')

            silencio = 8.0 if any(k in comando.lower() for k in ('show run', 'show tech', 'show version')) else 3.0

            resultado_bytes = b""
            deadline = time.time() + 180
            ultimo_dado = time.time()

            while time.time() < deadline:
                try:
                    chunk = process.read_nonblocking(size=65536, timeout=0.1)
                    if chunk:
                        resultado_bytes += chunk
                        ultimo_dado = time.time()
                except pexpect.exceptions.TIMEOUT:
                    if time.time() - ultimo_dado >= silencio:
                        break
                except (pexpect.exceptions.EOF, OSError):
                    break

            resultado = resultado_bytes.decode('utf-8', errors='replace')
            resultado = limpar_ansi(resultado)
            linhas = resultado.split('\n')
            if linhas and comando.strip() in linhas[0]:
                linhas = linhas[1:]
            resultado = '\n'.join(linhas)
            output += resultado + "\n"
            print(f"    ✅ {len(resultado)} bytes")

        except Exception as e:
            print(f"    ❌ {e}")
            output += f"ERRO: {str(e)}\n"

    try:
        process.send(b'exit\r')
        time.sleep(0.5)
        process.close()
    except:
        pass

    return output

    
def realizar_backup(acesso, usuario=None):
    inicio = time.time()
    ssh_tunnel = None

    try:
        print(f"\n{'='*80}")
        print(f"🔄 INICIANDO BACKUP")
        print(f"{'='*80}")
        print(f"📋 Equipamento: {acesso.tipo}")
        print(f"📡 Host: {acesso.host}:{acesso.porta}")
        print(f"👤 Usuário: {acesso.usuario}")
        print(f"🔧 Modelo: {acesso.modelo}")
        print(f"📝 Template: {acesso.backup_template.nome if acesso.backup_template else 'N/A'}")

        eh_privado = is_private_ip(acesso.host)
        print(f"🔍 IP Privado? {eh_privado}")

        host_conexao = acesso.host
        porta_conexao = int(acesso.porta) if acesso.porta else 22

        # ✅ Detectar fabricante — combina modelo.fabricante + modelo.nome +
        # acesso.tipo (não só modelo.nome) porque o Modelo_equipamento
        # vinculado pode estar cadastrado errado (ex: OLT ZTE com modelo
        # "debian 12" por engano); acesso.tipo costuma estar correto mesmo
        # quando o modelo não está — mesmo fallback já usado em consumers.py.
        _partes_deteccao = []
        if acesso.modelo:
            if getattr(acesso.modelo, 'fabricante', None):
                _partes_deteccao.append(str(acesso.modelo.fabricante))
            if getattr(acesso.modelo, 'nome', None):
                _partes_deteccao.append(str(acesso.modelo.nome))
        if acesso.tipo:
            _partes_deteccao.append(acesso.tipo)
        modelo_nome = ' '.join(_partes_deteccao).lower()

        is_huawei   = 'huawei' in modelo_nome
        is_a10      = 'a10'    in modelo_nome
        is_cisco    = 'cisco'  in modelo_nome
        is_zte      = 'zte'    in modelo_nome
        is_parks    = 'parks'  in modelo_nome
        is_mikrotik = any(k in modelo_nome for k in ('mikrotik', 'routeros', 'routerboard'))

        print(f"🏭 Huawei: {is_huawei} | A10: {is_a10} | Cisco: {is_cisco} | ZTE: {is_zte} | Parks: {is_parks} | MikroTik: {is_mikrotik}")

        # ✅ Criar túnel se IP privado
        if eh_privado:
            print(f"\n{'='*80}")
            print(f"⚠️ IP PRIVADO - CRIANDO TÚNEL SSH")
            print(f"{'='*80}")

            proxy = ProxyServer.objects.filter(cliente=acesso.cliente, ativo=True).first()
            if not proxy:
                # Verificar se VPN WireGuard ativa cobre este IP
                if vpn_cobre_ip(acesso.cliente, acesso.host):
                    print(f"✅ VPN WireGuard cobre {acesso.host} — conectando diretamente")
                    # host_conexao e porta_conexao já apontam para o host diretamente via VPN
                else:
                    raise Exception(
                        "IP privado sem proxy SSH ativo. "
                        "Configure um túnel SSH na aba 'Túneis SSH'."
                    )
            else:
                print(f"✅ Proxy encontrado: {proxy.nome}")

                ssh_tunnel = criar_ssh_tunnel(
                    {
                        'host':    proxy.host,
                        'porta':   proxy.porta,
                        'usuario': proxy.usuario,
                        'senha':   proxy.senha
                    },
                    acesso.host,
                    porta_conexao
                )

                host_conexao = ssh_tunnel['local_host']
                porta_conexao = ssh_tunnel['local_port']

                print(f"✅ Túnel criado: localhost:{porta_conexao} → {acesso.host}:{acesso.porta}")
                time.sleep(1)

        # ✅ Preparar diretório de backup
        backup_dir = preparar_diretorio_backup(acesso.cliente.id, acesso.id)
        print(f"\n📁 Diretório: {backup_dir}")

        comandos = acesso.backup_template.get_comandos_list()
        print(f"🔢 Total de comandos: {len(comandos)}")

        # ✅ Parks usa pexpect (SSH real) — Paramiko invoke_shell não é compatível
        if is_parks:
            print(f"\n{'='*80}")
            print(f"📋 EXECUTANDO COMANDOS (pexpect — Parks FIBERLINK)")
            print(f"{'='*80}")
            output = _executar_comandos_pexpect(
                host_conexao, porta_conexao,
                acesso.usuario, acesso.senha,
                comandos
            )

        else:
            # ✅ Demais fabricantes: Paramiko
            print(f"\n{'='*80}")
            print(f"🔐 CONECTANDO VIA PARAMIKO")
            print(f"{'='*80}")

            _connect_kwargs = dict(
                hostname=host_conexao,
                port=porta_conexao,
                username=acesso.usuario,
                password=acesso.senha,
                timeout=30,
                look_for_keys=False,
                allow_agent=False,
                banner_timeout=30,
                # ZTE OLTs têm timeout de KEX curto — group16/group18/group-exchange
                # demoram 2-5s pra negociar e a conexão cai antes da auth.
                # Desabilitar força o paramiko a usar curve25519/ecdh/group14
                # (rápidos), que a OLT já suporta (ver docs/winbox_vnc.md e memória
                # do projeto). Restrito a is_zte: outros vendors (ex. Huawei NE8000)
                # às vezes só oferecem group-exchange-sha256 — desabilitar geral
                # zera o KEX em comum e quebra a conexão ("no acceptable kex
                # algorithm").
                disabled_algorithms={'kex': [
                    'diffie-hellman-group-exchange-sha256',
                    'diffie-hellman-group-exchange-sha1',
                    'diffie-hellman-group16-sha512',
                    'diffie-hellman-group18-sha512',
                ]} if is_zte else None,
            )
            # Handshake via túnel (proxy → forwarding local por threads) falha
            # de forma intermitente em algumas OLTs Huawei com "No existing
            # session" / "Authentication failed: transport shut down or saw
            # EOF" — visto 3 madrugadas seguidas na OLT-HU-LEAL enquanto o
            # Terminal (canal direto, sem o relay por thread) conectava sem
            # problema nos mesmos horários. Um retry único já resolve, pois a
            # falha é do handshake naquele instante, não das credenciais.
            _erros_transitorios = ('transport shut down or saw EOF', 'No existing session')
            for _tentativa in (1, 2):
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                try:
                    client.connect(**_connect_kwargs)
                    break
                except paramiko.SSHException as _e:
                    if _tentativa == 2 or not any(t in str(_e) for t in _erros_transitorios):
                        raise
                    print(f"⚠️ Handshake SSH falhou ({_e}) — tentando novamente em 3s...")
                    time.sleep(3)
            client.get_transport().set_keepalive(10)
            print(f"✅ Conectado!")

            print(f"\n{'='*80}")
            print(f"📋 EXECUTANDO COMANDOS")
            print(f"{'='*80}")

            # Templates com comando "cd <dir>" (ex: OLTs FiberHome/WOS que navegam
            # entre diretórios da CLI antes do show) exigem sessão com estado
            # persistente — client.exec_command() abre um canal novo e sem
            # memória a cada chamada, então "cd service" de um comando não
            # afeta o "show" de outro. Pior: alguns desses equipamentos (ex.
            # confirmado em teste real numa OLT VSOL/WOS) derrubam a sessão
            # SSH inteira depois do primeiro exec_command, quebrando todos os
            # comandos seguintes do template.
            precisa_shell_persistente = any(c.strip().lower().startswith('cd ') for c in comandos)

            if is_huawei:
                output = _executar_comandos_huawei(client, comandos)
            elif is_a10:
                output = _executar_comandos_a10(client, comandos, acesso.senha_adm)
            elif is_cisco or is_zte:
                output = _executar_comandos_cisco(client, comandos, acesso.usuario, acesso.senha)
            elif is_mikrotik:
                output = _executar_comandos_mikrotik(client, comandos)
            elif precisa_shell_persistente:
                output = _executar_comandos_shell_generico(client, comandos)
            else:
                # Datacom, Juniper, etc.
                output = _executar_comandos_sem_pty(client, comandos)

            client.close()

        if len(output.strip()) < 100:
            raise Exception("Backup vazio ou muito pequeno. Verifique os comandos do template.")

        # ── Normalizar saída e calcular hash ─────────────────────────────
        # Remove ANSI escape codes e linhas que contenham apenas timestamps
        output_norm = re.sub(r'\x1b\[[0-9;]*[mGKHF]', '', output)
        # Remove linhas de paginação comuns (Huawei, Cisco)
        output_norm = re.sub(r'(?m)^[ \t]*--[Mm]ore--|^[ \t]*<--- More --->', '', output_norm)
        output_norm = output_norm.strip()
        hash_novo = hashlib.sha256(output_norm.encode('utf-8', errors='replace')).hexdigest()

        # ── Comparar com último backup bem-sucedido ──────────────────────
        ultimo = BackupLog.objects.filter(
            acesso=acesso,
            status='SUCESSO',
            hash_conteudo=hash_novo,
        ).first()

        duracao = time.time() - inicio

        if ultimo:
            from django.utils import timezone as _tz
            print(f"\n🔁 Conteúdo idêntico ao backup {ultimo.id} — nada a salvar.")
            ultimo.ultima_verificacao = _tz.now()
            ultimo.save(update_fields=['ultima_verificacao'])
            return {
                'sucesso': True,
                'sem_mudancas': True,
                'arquivo': '',
                'tamanho': 0,
                'duracao': f"{duracao:.2f}s",
            }

        # ── Salvar arquivo ───────────────────────────────────────────────
        print(f"\n{'='*80}")
        print(f"💾 SALVANDO ARQUIVO")
        print(f"{'='*80}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # Sanitiza acesso.tipo para uso como nome de arquivo: só espaço virava "_"
        # antes, mas "/" (comum em nomes tipo "BRAS/CGNAT/BORDA - JUNIPER") vira
        # separador de diretório no os.path.join — o(s) subdiretório(s) resultante(s)
        # não existe(m) e o open() abaixo falha com FileNotFoundError. Qualquer
        # caractere fora de letras/números/"-"/"_" vira "_".
        tipo_seguro = re.sub(r'[^A-Za-z0-9_-]+', '_', acesso.tipo).strip('_') or 'backup'
        # MikroTik: .rsc é o formato nativo de script RouterOS (o /export
        # show-sensitive do template já grava as senhas em texto puro, ao
        # contrário do /system backup binário) — cabeçalho abaixo vira
        # comentário "#" para o arquivo continuar sendo um script válido,
        # importável de volta via /import file=....
        extensao = 'rsc' if is_mikrotik else 'txt'
        marcador = '#' if is_mikrotik else ''
        nome_arquivo = f"{tipo_seguro}_{timestamp}.{extensao}"
        arquivo_path = os.path.join(backup_dir, nome_arquivo)

        with open(arquivo_path, 'w', encoding='utf-8') as f:
            def _linha(texto=''):
                prefixo = f"{marcador} " if marcador else ''
                f.write(f"{prefixo}{texto}\n")

            f.write(f"{marcador}{'='*80}\n")
            _linha(f"BACKUP DE CONFIGURAÇÃO")
            f.write(f"{marcador}{'='*80}\n")
            _linha(f"Cliente: {acesso.cliente.nome_empresa}")
            _linha(f"Equipamento: {acesso.tipo}")
            _linha(f"Host: {acesso.host}:{acesso.porta}")
            _linha(f"Acesso: {'VIA PROXY SSH' if eh_privado else 'DIRETO'}")
            _linha(f"Modelo: {acesso.modelo}")
            _linha(f"Template: {acesso.backup_template.nome}")
            _linha(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            _linha(f"Executado por: {usuario.username if usuario else 'Sistema'}")
            f.write(f"{marcador}{'='*80}\n\n")
            f.write(output)

        tamanho = os.path.getsize(arquivo_path)
        arquivo_relativo = os.path.relpath(arquivo_path, settings.MEDIA_ROOT)

        print(f"✅ Arquivo: {nome_arquivo}")
        print(f"📊 Tamanho: {tamanho} bytes")
        print(f"⏱️ Duração: {duracao:.2f}s")

        # ── Registrar log ────────────────────────────────────────────────
        BackupLog.objects.create(
            acesso=acesso,
            cliente=acesso.cliente,
            template=acesso.backup_template,
            arquivo_path=arquivo_relativo,
            tamanho_bytes=tamanho,
            hash_conteudo=hash_novo,
            status='SUCESSO',
            mensagem='Backup realizado com sucesso',
            executado_por=usuario,
            duracao_segundos=duracao,
        )

        print(f"\n{'='*80}")
        print(f"✅ BACKUP CONCLUÍDO COM SUCESSO!")
        print(f"{'='*80}\n")

        return {
            'sucesso': True,
            'sem_mudancas': False,
            'arquivo': nome_arquivo,
            'tamanho': tamanho,
            'duracao': f"{duracao:.2f}s",
        }

    except Exception as e:
        erro = f"Erro: {str(e)}"
        print(f"\n❌ {erro}\n")
        registrar_erro_backup(acesso, usuario, erro, time.time() - inicio)
        return {'sucesso': False, 'erro': erro}

    finally:
        if ssh_tunnel:
            try:
                if 'ssh_client' in ssh_tunnel:
                    ssh_tunnel['ssh_client'].close()
                if 'server_socket' in ssh_tunnel:
                    ssh_tunnel['server_socket'].close()
            except:
                pass

def _executar_comandos_a10(client, comandos, senha_enable):
    """
    A10 Thunder: invoke_shell com enable password.
    Fluxo:
    1. Aguarda prompt >
    2. Envia 'enable'
    3. Envia senha_enable
    4. Aguarda prompt #
    5. Executa comandos do template
    """
    output = ""

    channel = client.invoke_shell(
        term='vt100',
        width=10000,
        height=50
    )
    channel.settimeout(30)

    # Aguardar banner inicial
    time.sleep(2)
    banner = _ler_ate_silencio(channel, silencio=2.0, max_wait=10)
    print(f"    📋 Banner: {banner[-100:]!r}")

    # ✅ Entrar em modo enable
    print(f"    🔐 Enviando enable...")
    channel.send('enable\n')
    time.sleep(1)

    resposta_enable = _ler_ate_silencio(channel, silencio=1.5, max_wait=10)
    print(f"    📋 Após enable: {resposta_enable[-100:]!r}")

    # ✅ Enviar senha de enable (campo senha_adm)
    if 'password' in resposta_enable.lower() or 'Password' in resposta_enable:
        print(f"    🔑 Enviando senha enable...")
        channel.send((senha_enable or '') + '\n')
        time.sleep(1)
        resposta_senha = _ler_ate_silencio(channel, silencio=1.5, max_wait=10)
        print(f"    📋 Após senha: {resposta_senha[-100:]!r}")

        if 'invalid' in resposta_senha.lower() or 'fail' in resposta_senha.lower():
            raise Exception("Senha de enable incorreta no A10 Thunder")
    else:
        # Pode já estar em modo privilegiado
        print(f"    ⚠️ Prompt de senha não apareceu, continuando...")

    # ✅ Executar comandos do template
    for i, comando in enumerate(comandos, 1):
        # 'enable' já foi tratado acima (com a senha_adm) — reenviar aqui
        # forçaria um segundo prompt de senha e o próximo comando do
        # template seria interpretado como a senha, quebrando a sessão.
        if comando.strip().lower() == 'enable':
            print(f"  [{i}/{len(comandos)}] {comando} (já enviado — ignorando)")
            continue

        print(f"  [{i}/{len(comandos)}] {comando}")
        output += f"\n{'='*60}\nComando: {comando}\n{'='*60}\n"

        try:
            channel.send(comando + '\n')

            # show running-config pode demorar mais
            silencio = 5.0 if 'show' in comando.lower() else 2.0
            resultado = _ler_ate_silencio(channel, silencio=silencio, max_wait=120)

            if resultado:
                resultado = limpar_ansi(resultado)

                # Remover eco do comando
                linhas = resultado.split('\n')
                if linhas and comando.strip() in linhas[0]:
                    linhas = linhas[1:]
                resultado = '\n'.join(linhas)

                output += resultado + "\n"
                print(f"    ✅ {len(resultado)} bytes")
            else:
                print(f"    ⚠️ Output vazio")

        except Exception as e:
            print(f"    ❌ {e}")
            output += f"ERRO: {str(e)}\n"

    try:
        channel.send('exit\n')
        time.sleep(0.5)
        channel.close()
    except:
        pass

    return output

def _executar_comandos_sem_pty(client, comandos):
    """
    MikroTik, Cisco e outros: exec_command sem PTY.
    Sem terminal = sem quebra de linha.
    """
    output = ""
    for i, comando in enumerate(comandos, 1):
        print(f"  [{i}/{len(comandos)}] {comando}")
        output += f"\n{'='*60}\nComando: {comando}\n{'='*60}\n"
        try:
            stdin, stdout, stderr = client.exec_command(
                comando, timeout=120, get_pty=False
            )
            resultado = stdout.read().decode('utf-8', errors='replace')
            if resultado:
                resultado = limpar_ansi(resultado)
                output += resultado + "\n"
                print(f"    ✅ {len(resultado)} bytes")
        except Exception as e:
            print(f"    ❌ {e}")
            output += f"ERRO: {str(e)}\n"
    return output


def _erro_sintaxe_routeros(saida):
    """
    Detecta se a saída de um comando RouterOS é, na verdade, um erro de
    sintaxe do CLI (ex: parâmetro não existente nessa versão do RouterOS),
    e não o resultado esperado do comando.
    """
    s = (saida or '').lower()
    return any(marcador in s for marcador in (
        'expected end of command',
        'no such command',
        'bad command name',
        'unknown parameter',
        "no such argument",
    ))


def _executar_comandos_mikrotik(client, comandos):
    """
    MikroTik via exec_command sem PTY. Saída vira um script RouterOS válido
    (separadores de comando como comentário "#") para o backup poder ser
    salvo como .rsc e reimportado via /import se necessário.

    Comandos "/export ... show-sensitive ..." caem automaticamente para a
    mesma linha sem "show-sensitive" quando o RouterOS é antigo demais e não
    reconhece o parâmetro (introduzido no 6.43) — sem esse fallback o backup
    fica vazio/com erro de sintaxe nesses equipamentos mais antigos.
    """
    output = ""
    for i, comando in enumerate(comandos, 1):
        print(f"  [{i}/{len(comandos)}] {comando}")
        cmd_efetivo = comando
        try:
            stdin, stdout, stderr = client.exec_command(comando, timeout=120, get_pty=False)
            resultado = stdout.read().decode('utf-8', errors='replace')

            if 'show-sensitive' in comando and _erro_sintaxe_routeros(resultado):
                cmd_efetivo = re.sub(r'\s*show-sensitive\s*', ' ', comando).strip()
                print(f"    ⚠️ 'show-sensitive' não suportado neste RouterOS — repetindo como: {cmd_efetivo}")
                stdin, stdout, stderr = client.exec_command(cmd_efetivo, timeout=120, get_pty=False)
                resultado = stdout.read().decode('utf-8', errors='replace')

            output += f"\n# {'='*60}\n# Comando: {cmd_efetivo}\n# {'='*60}\n"
            if resultado:
                resultado = limpar_ansi(resultado)
                output += resultado + "\n"
                print(f"    ✅ {len(resultado)} bytes")
        except Exception as e:
            output += f"\n# {'='*60}\n# Comando: {cmd_efetivo}\n# {'='*60}\n"
            print(f"    ❌ {e}")
            output += f"ERRO: {str(e)}\n"
    return output

def _executar_comandos_cisco(client, comandos, usuario='', senha=''):
    """
    Cisco IOS/IOS-XE, ZTE C300 e Parks FIBERLINK.
    Versão diagnóstico: loga tudo que o OLT envia para identificar o problema.
    """
    output = ""

    try:
        client.get_transport().set_keepalive(10)
    except Exception:
        pass

    try:
        channel = client.invoke_shell(term='vt100', width=200, height=50)
    except Exception as e:
        raise Exception(f"Falha ao abrir shell interativo: {e}")

    channel.settimeout(120)

    # ✅ Aguardar mais tempo — Parks pode demorar para exibir o prompt
    time.sleep(5)
    banner = _ler_ate_silencio(channel, silencio=4.0, max_wait=20)
    banner_lower = banner.lower()

    # ✅ LOG COMPLETO — essencial para diagnóstico
    print(f"\n{'='*60}")
    print(f"📋 BANNER COMPLETO DO OLT:")
    print(f"{'='*60}")
    print(repr(banner))
    print(f"{'='*60}")
    print(f"Canal fechado após banner? {channel.closed}")
    print(f"{'='*60}\n")

    # ✅ Responder ao login interativo se detectado
    if any(k in banner_lower for k in ('login:', 'username:', 'user:')):
        print(f"    🔐 Prompt de usuário detectado — enviando usuário...")
        channel.send(usuario + '\n')
        time.sleep(2)
        resposta = _ler_ate_silencio(channel, silencio=3.0, max_wait=15)
        print(f"\n📋 RESPOSTA APÓS USUÁRIO:\n{repr(resposta)}\n")

        if any(k in resposta.lower() for k in ('password:', 'passwd:', 'senha:')):
            print(f"    🔐 Prompt de senha — enviando senha...")
            channel.send(senha + '\n')
            time.sleep(2)
            resposta2 = _ler_ate_silencio(channel, silencio=3.0, max_wait=15)
            print(f"\n📋 RESPOSTA APÓS SENHA:\n{repr(resposta2)}\n")

    elif any(k in banner_lower for k in ('password:', 'passwd:', 'senha:')):
        print(f"    🔐 Prompt de senha direto — enviando senha...")
        channel.send(senha + '\n')
        time.sleep(2)
        resposta = _ler_ate_silencio(channel, silencio=3.0, max_wait=15)
        print(f"\n📋 RESPOSTA APÓS SENHA:\n{repr(resposta)}\n")

    else:
        print(f"    ✅ Sem prompt de login detectado — seguindo...")

    # ✅ Log do estado do canal ANTES de tentar enviar comandos
    print(f"Canal fechado antes dos comandos? {channel.closed}")

    # ✅ NÃO interrompe mais aqui — tenta continuar e loga o que acontece
    # Desabilitar paginação
    try:
        print("  [0] terminal length 0")
        channel.send('terminal length 0\n')
        time.sleep(1)
        resp_tl = _ler_ate_silencio(channel, silencio=2.0, max_wait=10)
        print(f"\n📋 RESPOSTA APÓS terminal length 0:\n{repr(resp_tl)}\n")
    except Exception as e:
        print(f"    ❌ Falha ao enviar terminal length 0: {e}")
        raise Exception(
            f"Falha ao enviar 'terminal length 0': {e}. "
            f"Verifique o log acima para ver o que o OLT enviou."
        )

    for i, comando in enumerate(comandos, 1):
        if 'terminal length' in comando.lower():
            print(f"  [{i}/{len(comandos)}] {comando} (já enviado — ignorando)")
            continue

        print(f"  [{i}/{len(comandos)}] {comando}")
        print(f"  Canal fechado antes de '{comando}'? {channel.closed}")
        output += f"\n{'='*60}\nComando: {comando}\n{'='*60}\n"

        try:
            channel.send(comando + '\n')

            silencio = 8.0 if any(k in comando.lower() for k in ('show run', 'show tech', 'show version')) else 3.0
            resultado = _ler_ate_silencio(channel, silencio=silencio, max_wait=180)

            print(f"\n📋 RESPOSTA DE '{comando}':\n{repr(resultado[:500])}\n")

            if resultado:
                resultado = limpar_ansi(resultado)
                linhas = resultado.split('\n')
                if linhas and comando.strip() in linhas[0]:
                    linhas = linhas[1:]
                resultado = '\n'.join(linhas)
                output += resultado + "\n"
                print(f"    ✅ {len(resultado)} bytes")
            else:
                print(f"    ⚠️ Output vazio")

        except Exception as e:
            err = str(e).lower()
            print(f"    ❌ Erro em '{comando}': {e}")
            output += f"ERRO: {str(e)}\n"
            if any(k in err for k in ('socket', 'closed', 'eof', 'reset', 'broken')):
                raise Exception(
                    f"Conexão perdida em '{comando}': {e}. "
                    f"Verifique o log completo do banner acima."
                )

    try:
        channel.send('exit\n')
        time.sleep(0.5)
        channel.close()
    except Exception:
        pass

    return output



def _executar_comandos_huawei(client, comandos):
    """
    Huawei VRP: invoke_shell com terminal de 10000 colunas.
    - screen-length 0 temporary precisa rodar na mesma sessão
    que display current-configuration
    - Terminal largo evita quebra de linha
    """
    import re

    output = ""

    # ✅ Abrir shell com terminal muito largo
    channel = client.invoke_shell(
        term='vt100',
        width=10000,
        height=50
    )
    channel.settimeout(120)

    # Aguardar o banner de login
    time.sleep(3)
    _ler_ate_silencio(channel, silencio=2.0)

    for i, comando in enumerate(comandos, 1):
        print(f"  [{i}/{len(comandos)}] {comando}")
        output += f"\n{'='*60}\nComando: {comando}\n{'='*60}\n"

        try:
            channel.send(comando + '\n')

            # Aguardar o output completo (silence detection)
            resultado = _ler_ate_silencio(channel, silencio=3.0, max_wait=120)

            if resultado:
                resultado = limpar_ansi(resultado)

                # Remover o eco do comando enviado (primeira linha)
                linhas = resultado.split('\n')
                if linhas and comando.strip() in linhas[0]:
                    linhas = linhas[1:]
                resultado = '\n'.join(linhas)

                output += resultado + "\n"
                print(f"    ✅ {len(resultado)} bytes")
            else:
                print(f"    ⚠️ Output vazio")

        except Exception as e:
            print(f"    ❌ {e}")
            output += f"ERRO: {str(e)}\n"

    try:
        channel.send('quit\n')
        time.sleep(0.5)
        channel.close()
    except:
        pass

    return output


def _executar_comandos_shell_generico(client, comandos):
    """
    Shell interativo genérico (invoke_shell) para equipamentos cujo template
    depende de estado entre comandos — ex: "cd service" seguido de "terminal
    length 0" e "show startup-config" (comum em OLTs FiberHome/firmware WOS
    compartilhada com outras marcas, ex: VSOL).

    client.exec_command() não serve aqui: cada chamada abre um canal novo e
    sem memória do anterior (o "cd" não vale para o próximo comando) e, em
    alguns desses equipamentos, o servidor SSH derruba a sessão inteira
    depois do primeiro exec_command — testado ao vivo numa OLT VSOL/WOS.
    """
    output = ""

    channel = client.invoke_shell(term='vt100', width=200, height=50)
    channel.settimeout(120)

    time.sleep(3)
    _ler_ate_silencio(channel, silencio=2.0)

    for i, comando in enumerate(comandos, 1):
        print(f"  [{i}/{len(comandos)}] {comando}")
        output += f"\n{'='*60}\nComando: {comando}\n{'='*60}\n"

        try:
            channel.send(comando + '\n')
            resultado = _ler_ate_silencio(channel, silencio=3.0, max_wait=120)

            if resultado:
                resultado = limpar_ansi(resultado)
                linhas = resultado.split('\n')
                if linhas and comando.strip() in linhas[0]:
                    linhas = linhas[1:]
                resultado = '\n'.join(linhas)
                output += resultado + "\n"
                print(f"    ✅ {len(resultado)} bytes")
            else:
                print(f"    ⚠️ Output vazio")

        except Exception as e:
            print(f"    ❌ {e}")
            output += f"ERRO: {str(e)}\n"

    try:
        channel.send('exit\n')
        time.sleep(0.5)
        channel.close()
    except Exception:
        pass

    return output


def _ler_ate_silencio(channel, silencio=2.0, max_wait=120):
    """
    Lê do channel paramiko até silêncio por `silencio` segundos
    ou até atingir `max_wait` segundos no total.
    """
    resultado = b""
    inicio = time.time()
    ultimo_dado = time.time()

    while True:
        if time.time() - inicio > max_wait:
            print(f"    ⚠️ max_wait atingido ({max_wait}s)")
            break

        if channel.recv_ready():
            chunk = channel.recv(65536)
            if chunk:
                resultado += chunk
                ultimo_dado = time.time()
        else:
            if time.time() - ultimo_dado >= silencio:
                break
            time.sleep(0.05)

    return resultado.decode('utf-8', errors='replace')


def limpar_ansi(texto):
    """Remove códigos ANSI e normaliza quebras de linha."""
    import re
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    texto = ansi_escape.sub('', texto)
    texto = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', texto)
    texto = texto.replace('[K', '')
    texto = texto.replace('\r\n', '\n').replace('\r', '\n')
    return texto

def ler_saida_comando(ssh_process, silence_timeout=2.0, max_timeout=120,modelo=None):
    """
    ✅ MELHORADO: Lê output até detectar silence + pós-processamento de linhas
    """
    print(f"       🔍 Detectando fim do comando por silence...")

    resultado = ""
    tempo_inicio = time.time()
    ultimo_dado = time.time()
    silence_count = 0
    bytes_totais = 0

    while True:
        tempo_decorrido = time.time() - tempo_inicio

        if tempo_decorrido > max_timeout:
            print(f"       ⚠️ Timeout máximo ({max_timeout}s) atingido")
            break

        try:
            dados = ssh_process.read_nonblocking(timeout=0.1, size=65536)

            if dados:
                resultado += dados
                bytes_totais += len(dados)
                ultimo_dado = time.time()
                silence_count = 0
                print(f"       📥 {len(dados)} bytes ({bytes_totais} total)")
            else:
                silence_count += 1
                tempo_silencio = time.time() - ultimo_dado

                if tempo_silencio >= silence_timeout:
                    print(f"       ✅ Silence detectado ({tempo_silencio:.1f}s) - comando terminou")
                    break

                time.sleep(0.1)

        except pexpect.exceptions.TIMEOUT:
            tempo_silencio = time.time() - ultimo_dado

            if tempo_silencio >= silence_timeout:
                print(f"       ✅ Silence detectado ({tempo_silencio:.1f}s) - comando terminou")
                break

            time.sleep(0.1)

        except Exception as e:
            print(f"       ⚠️ Erro ao ler: {str(e)}")
            break

    print(f"       ✅ Leitura completa: {bytes_totais} bytes, {time.time() - tempo_inicio:.1f}s")

    # ✅ NOVO: Pós-processamento para limpar linhas quebradas
    resultado_limpo = limpar_output_por_fabricante(resultado, modelo)

    return resultado_limpo


def limpar_output_por_fabricante(texto, modelo):
    """
    ✅ NOVA FUNÇÃO: Aplica limpeza específica baseada no fabricante
    """
    # 1. Detectar fabricante
    fabricante = detectar_fabricante(modelo)
    print(f"       🏭 Fabricante detectado: {fabricante}")

    # 2. Aplicar limpeza básica (códigos ANSI) para todos
    texto = remover_codigos_ansi(texto)

    # 3. Aplicar limpeza específica
    if fabricante == 'MIKROTIK':
        return limpar_output_mikrotik(texto)
    elif fabricante in ['CISCO', 'HUAWEI', 'DATACOM', 'JUNIPER', 'EXTREME', 'HP', 'DELL']:
        return limpar_output_generico(texto)
    else:
        return limpar_output_generico(texto)


def detectar_fabricante(modelo):
    """
    Detecta fabricante baseado no modelo do equipamento
    """
    if not modelo:
        return 'DESCONHECIDO'

    modelo_nome = ''
    if hasattr(modelo, 'nome'):
        modelo_nome = modelo.nome.upper()
    elif isinstance(modelo, str):
        modelo_nome = modelo.upper()

    if 'MIKROTIK' in modelo_nome or 'ROUTERBOARD' in modelo_nome:
        return 'MIKROTIK'
    elif 'CISCO' in modelo_nome:
        return 'CISCO'
    elif 'HUAWEI' in modelo_nome:
        return 'HUAWEI'
    elif 'DATACOM' in modelo_nome:
        return 'DATACOM'
    elif 'JUNIPER' in modelo_nome:
        return 'JUNIPER'
    elif 'EXTREME' in modelo_nome:
        return 'EXTREME'
    elif 'HP' in modelo_nome or 'ARUBA' in modelo_nome:
        return 'HP'
    elif 'DELL' in modelo_nome:
        return 'DELL'
    else:
        return 'DESCONHECIDO'


def remover_codigos_ansi(texto):
    """
    Remove códigos ANSI - aplica-se a TODOS os fabricantes
    """
    import re

    # Remover sequências ANSI
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    texto = ansi_escape.sub('', texto)

    # Remover caracteres de controle (exceto \n e \r)
    texto = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', texto)

    # Remover [K
    texto = texto.replace('[K', '')

    return texto


def limpar_output_generico(texto):
    """
    ✅ NOVA FUNÇÃO: Limpeza conservadora para Cisco, Huawei, Datacom, etc.
    MANTÉM quebras de linha originais!
    """
    import re

    print(f"       🔧 Aplicando limpeza genérica (mantém quebras de linha)...")

    # Normalizar line breaks
    texto = texto.replace('\r\n', '\n').replace('\r', '\n')

    linhas = texto.split('\n')
    linhas_limpas = []
    linha_anterior = ''

    for linha in linhas:
        linha_strip = linha.strip()

        # ✅ Manter linhas vazias (não mais de 3 consecutivas)
        if not linha_strip:
            if linha_anterior != '' or (linhas_limpas and linhas_limpas[-1] != ''):
                linhas_limpas.append(linha)
                linha_anterior = ''
            continue

        # Remover apenas prompts duplicados óbvios
        if linha_strip.endswith('#') or linha_strip.endswith('>'):
            if linha_anterior and (linha_anterior.endswith('#') or linha_anterior.endswith('>')):
                continue

        linhas_limpas.append(linha)
        linha_anterior = linha_strip

    texto_final = '\n'.join(linhas_limpas)

    # Remover apenas 4+ linhas vazias consecutivas
    texto_final = re.sub(r'\n{4,}', '\n\n\n', texto_final)

    return texto_final

def limpar_output_mikrotik(texto):
    import re

    # 1. Remover códigos ANSI
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    texto = ansi_escape.sub('', texto)

    # 2. Remover caracteres de controle
    texto = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', texto)
    texto = texto.replace('[K', '')

    # 3. Normalizar quebras de linha
    texto = texto.replace('\r\n', '\n').replace('\r', '\n')

    # 4. Reagrupar linhas quebradas pelo terminal (wrap de 80 colunas do RouterOS)
    #
    # RouterOS quebra linhas longas sem marcador de continuação.
    # Padrão: linha com 79-80 chars → próxima linha NÃO começa com
    # '/', '#', '[' (prompt), ou está vazia.
    linhas = texto.split('\n')
    resultado = []
    i = 0

    while i < len(linhas):
        linha_atual = linhas[i]

        # Continua juntando enquanto parece linha quebrada
        while True:
            proxima_existe = (i + 1) < len(linhas)
            if not proxima_existe:
                break

            proxima = linhas[i + 1]
            proxima_strip = proxima.strip()

            # Comprimento da linha atual sem espaços à direita
            comprimento = len(linha_atual.rstrip())

            # Condições para ser linha quebrada pelo terminal:
            # - Linha atual tem 76-80 chars (wrap típico de 80 colunas)
            # - Linha atual não está vazia
            # - Próxima linha não é um novo comando RouterOS
            # - Próxima linha não é prompt '[usuario@...]'
            # - Próxima linha não é comentário '#'
            # - Próxima linha não está vazia
            eh_quebra = (
                76 <= comprimento <= 80
                and linha_atual.rstrip()
                and proxima_strip
                and not proxima_strip.startswith('/')
                and not proxima_strip.startswith('#')
                and not proxima_strip.startswith('[')
            )

            if not eh_quebra:
                break

            # Juntar sem espaço extra (RouterOS quebra no meio da palavra/token)
            linha_atual = linha_atual.rstrip() + proxima_strip
            i += 1

        resultado.append(linha_atual)
        i += 1

    texto = '\n'.join(resultado)

    # 5. Remover linhas que são apenas o prompt RouterOS
    # Ex: "[tomich@SW - PEREIROS] > "
    texto = re.sub(r'^\[.+?\]\s*>\s*$', '', texto, flags=re.MULTILINE)

    # 6. Remover excesso de linhas vazias
    texto = re.sub(r'\n{3,}', '\n\n', texto)

    return texto.strip()


def conectar_ssh_backup(host, porta, usuario, senha, senha_adm, timeout=120):
    print(f"📤 SSH: Conectando a {host}:{porta}...")

    ssh_path = "/usr/bin/ssh"
    if not os.path.exists(ssh_path):
        ssh_path = "ssh"

    ssh_cmd = (
        f"{ssh_path} -o StrictHostKeyChecking=no "
        f"-o UserKnownHostsFile=/dev/null "
        f"-o ConnectTimeout=30 "
        f"-o ServerAliveInterval=60 "
        f"-o LogLevel=ERROR "
        f"-o KexAlgorithms=+diffie-hellman-group-exchange-sha1,diffie-hellman-group14-sha1,diffie-hellman-group1-sha1 "
        f"-o HostKeyAlgorithms=+ssh-rsa,ssh-dss "
        f"-o PubkeyAcceptedAlgorithms=+ssh-rsa "
        f"-o Ciphers=+aes128-cbc,aes192-cbc,aes256-cbc,3des-cbc "
        f"-o MACs=+hmac-sha1,hmac-sha2-256,hmac-sha2-512 "
        f"-p {porta} {usuario}@{host}"
    )

    ssh_process = None
    tentativas = 3

    for tentativa in range(tentativas):
        try:
            print(f"   [Tentativa {tentativa + 1}/{tentativas}]")

            if tentativa < 2:
                ssh_process = pexpect.spawn(
                    ssh_cmd,
                    timeout=timeout,
                    encoding='utf-8',
                    maxread=65536,
                    cwd=os.path.expanduser('~')
                )
            else:
                ssh_process = pexpect.spawn(
                    '/bin/bash',
                    ['-c', ssh_cmd],
                    timeout=timeout,
                    encoding='utf-8',
                    maxread=65536
                )

            # ✅ CORREÇÃO PRINCIPAL: PTY com largura de 10000 colunas
            # Impede que RouterOS (e qualquer equipamento) quebre linhas longas.
            # Deve ser chamado ANTES da autenticação para que o SSH negocie
            # o tamanho correto do terminal com o equipamento remoto.
            ssh_process.setwinsize(24, 10000)
            print(f"   ✅ Spawn OK! PTY: 24x10000")
            break

        except FileNotFoundError as e:
            print(f"   ❌ FileNotFoundError: {str(e)}")
            if tentativa < tentativas - 1:
                time.sleep(2)
                continue
            else:
                raise Exception(
                    f"❌ SSH não encontrado após {tentativas} tentativas\n"
                    f"Execute no servidor: sudo apt-get install openssh-client"
                )
        except Exception as e:
            print(f"   ❌ Erro inesperado: {str(e)}")
            raise

    try:
        print(f"📤 Aguardando autenticação...")

        index = ssh_process.expect([
            "password:",
            "Password:",
            r".*[#>$\]].*",
        ], timeout=30)

        if index in (0, 1):
            print(f"🔐 Enviando senha...")
            ssh_process.sendline(senha)
            time.sleep(1)
            try:
                ssh_process.read_nonblocking(timeout=1.0, size=65536)
            except:
                pass

        print(f"⏳ Aguardando sistema estabilizar (3s)...")
        time.sleep(3)

        print(f"🧹 Limpando buffer com CTRL+U...")
        ssh_process.send("\x15")
        time.sleep(0.5)

        ssh_process.send("\r")
        time.sleep(0.5)

        try:
            ssh_process.expect([r".*[\#\>\$\]]\s*$", r">", r"\$", r"\]"], timeout=3)
            print(f"✅ Prompt detectado!")
        except pexpect.exceptions.TIMEOUT:
            try:
                ssh_process.read_nonblocking(timeout=0.5, size=65536)
            except:
                pass

        print(f"🔧 Desabilitando paginação...")
        ssh_process.send("terminal length 0\r")
        time.sleep(0.8)
        try:
            ssh_process.read_nonblocking(timeout=1.0, size=65536)
        except:
            pass

        print(f"🎨 Desabilitando cores ANSI do MikroTik...")
        ssh_process.send("set colors=never\r")
        time.sleep(0.8)
        try:
            ssh_process.read_nonblocking(timeout=1.0, size=65536)
        except:
            pass

        ssh_process.send("\r")
        time.sleep(1)
        try:
            ssh_process.read_nonblocking(timeout=1.0, size=65536)
        except:
            pass

        print(f"🔐 SINCRONIZANDO...")
        for tentativa_sync in range(3):
            ssh_process.send("\r")
            time.sleep(0.5)
            try:
                ssh_process.expect([r".*[\#\>\$\]]\s*$"], timeout=2)
                print(f"   ✅ Prompt respondeu!")
                break
            except pexpect.exceptions.TIMEOUT:
                continue

        time.sleep(2)

        try:
            while True:
                dados = ssh_process.read_nonblocking(timeout=0.2, size=65536)
                if not dados:
                    break
        except:
            pass

        print(f"✅✅✅ SSH: 100% PRONTO! ✅✅✅")
        return ssh_process

    except pexpect.exceptions.EOF:
        raise Exception("❌ Conexão SSH encerrada inesperadamente")
    except Exception as e:
        print(f"❌ {str(e)}")
        try:
            ssh_process.close()
        except:
            pass
        raise Exception(f"Erro SSH: {str(e)}")



def conectar_telnet_backup(host, porta, usuario, senha, timeout=120):
    """
    ✅ MEGA CORRIGIDO: Robustez TOTAL para Telnet também
    """
    print(f"📤 Telnet: Conectando a {host}:{porta}...")

    telnet_cmd = f"telnet {host} {porta}"

    telnet_process = pexpect.spawn(
        telnet_cmd,
        timeout=timeout,
        encoding='utf-8',
        maxread=65536  # 64KB
    )

    try:
        print(f"📤 Aguardando login prompt...")

        # ✅ PASSO 1: Aguardar login
        telnet_process.expect([
            "login:",
            "username:",
            "user:",
        ], timeout=15)

        print(f"🔐 Enviando usuário...")
        telnet_process.sendline(usuario)
        time.sleep(0.5)

        # ✅ PASSO 2: Aguardar senha
        telnet_process.expect([
            "password:",
            "Password:",
        ], timeout=10)

        print(f"🔐 Enviando senha...")
        telnet_process.sendline(senha)
        time.sleep(0.5)

        # ✅ PASSO 3: Aguardar prompt
        telnet_process.expect([
            r".*[\#\>\$\]]\s*$",
        ], timeout=15)

        print(f"✅ Autenticado!")

        # ✅ PASSO 4: Estabilizar (3s)
        print(f"⏳ Aguardando estabilizar (3s)...")
        time.sleep(3)

        # ✅ PASSO 5: Limpar
        print(f"🧹 Limpando...")
        telnet_process.send("\x15")  # Ctrl+U
        time.sleep(0.5)
        telnet_process.send("\r")
        time.sleep(0.5)

        try:
            telnet_process.read_nonblocking(timeout=1.0, size=65536)
        except:
            pass

        # ✅ PASSO 6: Sincronizar múltiplas vezes
        print(f"🔐 SINCRONIZANDO...")
        for tentativa in range(3):
            print(f"   Tentativa {tentativa + 1}/3...")
            telnet_process.send("\r")
            time.sleep(0.5)

            try:
                telnet_process.expect([r".*[\#\>\$\]]\s*$"], timeout=2)
                print(f"   ✅ Prompt respondeu!")
                break
            except pexpect.exceptions.TIMEOUT:
                print(f"   ⚠️ Timeout")
                continue

        # ✅ PASSO 7: Aguardar final
        print(f"⏳ Aguardando final (2s)...")
        time.sleep(2)

        # ✅ PASSO 8: Limpar buffer
        print(f"🧹 Limpando buffer...")
        try:
            while True:
                dados = telnet_process.read_nonblocking(timeout=0.2, size=65536)
                if not dados:
                    break
        except:
            pass

        print(f"✅✅✅ Telnet: 100% PRONTO! ✅✅✅")
        return telnet_process

    except pexpect.exceptions.TIMEOUT:
        raise Exception("❌ Timeout ao autenticar Telnet")
    except pexpect.exceptions.EOF:
        raise Exception("❌ Conexão Telnet encerrada")
    except Exception as e:
        print(f"❌ {str(e)}")
        try:
            telnet_process.close()
        except:
            pass
        raise Exception(f"Erro Telnet: {str(e)}")




def detectar_protocolo(porta):
    """Detecta protocolo pela porta"""
    porta_int = int(porta)

    if porta_int == 22:
        return 'ssh'
    elif porta_int == 23:
        return 'telnet'
    elif porta_int in [2222, 8022, 10022, 9022]:
        return 'ssh'
    elif porta_int in [2323, 9023]:
        return 'telnet'
    else:
        return 'ssh'



def preparar_diretorio_backup(cliente_id, acesso_id):
    """
    Cria estrutura de diretórios para backups
    """
    base_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    cliente_dir = os.path.join(base_dir, f'cliente_{cliente_id}')
    acesso_dir = os.path.join(cliente_dir, f'acesso_{acesso_id}')

    os.makedirs(acesso_dir, exist_ok=True)

    return acesso_dir


def mapear_device_type(modelo_nome):
    """
    Mapeia modelo do equipamento para device_type do Netmiko
    """
    modelo_lower = modelo_nome.lower()

    if 'cisco' in modelo_lower:
        if 'ios-xe' in modelo_lower or 'catalyst' in modelo_lower:
            return 'cisco_ios'
        elif 'nexus' in modelo_lower:
            return 'cisco_nxos'
        elif 'asa' in modelo_lower:
            return 'cisco_asa'
        else:
            return 'cisco_ios'

    elif 'huawei' in modelo_lower:
        return 'huawei'

    elif 'mikrotik' in modelo_lower:
        return 'mikrotik_routeros'

    elif 'juniper' in modelo_lower:
        return 'juniper_junos'

    elif 'dell' in modelo_lower:
        return 'dell_os10'

    elif 'hp' in modelo_lower or 'aruba' in modelo_lower:
        return 'hp_procurve'

    elif 'extreme' in modelo_lower:
        return 'extreme'

    else:
        return 'cisco_ios'  # Fallback


def is_private_ip(ip):
    """
    Verifica se IP é privado
    Intervalos privados:
    - 10.0.0.0/8
    - 172.16.0.0/12
    - 192.168.0.0/16
    """
    import ipaddress
    try:
        return ipaddress.ip_address(ip).is_private
    except:
        return False




def vpn_cobre_ip(cliente, host):
    """Verifica se existe VPN WireGuard OU túnel OpenVPN (aba Túneis) com rota
    que cobre o IP do host — em ambos os casos a rota já existe no kernel via
    interface própria, então a conexão pode ser feita direto, sem SSH tunnel."""
    try:
        import ipaddress as _ipa
        from .models import VPNWireGuard, VPNOpenVPN
        host_ip = _ipa.ip_address(host)

        vpns = VPNWireGuard.objects.filter(cliente=cliente, ativo=True, peer_no_servidor=True)
        for vpn in vpns:
            for rede_str in vpn.redes_lista():
                try:
                    if host_ip in _ipa.ip_network(rede_str, strict=False):
                        return True
                except ValueError:
                    pass

        tuneis = VPNOpenVPN.objects.filter(cliente=cliente, ativo=True, cert_emitido=True)
        for tunel in tuneis:
            for rede_str in tunel.redes_lista():
                try:
                    if host_ip in _ipa.ip_network(rede_str, strict=False):
                        return True
                except ValueError:
                    pass
    except Exception:
        pass
    return False


def criar_ssh_tunnel(proxy_server, equipamento_host, equipamento_porta, timeout=10):
    """
    ✅ MESMA FUNÇÃO DO TERMINAL SSH
    Cria túnel com socket forwarding
    """
    print(f"🔧 Criando túnel SSH...")

    try:
        # Conectar ao proxy
        print(f"📤 Conectando ao proxy...")
        ssh_proxy = paramiko.SSHClient()
        ssh_proxy.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        ssh_proxy.connect(
            hostname=proxy_server['host'],
            port=int(proxy_server['porta']),
            username=proxy_server['usuario'],
            password=proxy_server['senha'],
            timeout=timeout,
            look_for_keys=False,
            allow_agent=False
        )

        print(f"✅ Conectado ao proxy!")

        # Encontrar porta local
        sock_temp = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock_temp.bind(('127.0.0.1', 0))
        local_port = sock_temp.getsockname()[1]
        sock_temp.close()

        print(f"📍 Porta local: {local_port}")

        # Criar servidor
        print(f"🔗 Iniciando servidor de forwarding...")
        server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        server_socket.bind(('127.0.0.1', local_port))
        server_socket.listen(5)
        server_socket.settimeout(1)

        print(f"✅ Servidor escutando em 127.0.0.1:{local_port}")

        # Função de forwarding
        def forward_tunnel(client_socket, remote_host, remote_port, transport):
            """Forwarda dados via tunnel"""
            try:
                channel = transport.open_channel(
                    'direct-tcpip',
                    (remote_host, int(remote_port)),
                    ('127.0.0.1', local_port)
                )

                def forward_data(src, dst, direction):
                    """Forwards data"""
                    try:
                        while True:
                            data = src.recv(4096)
                            if not data:
                                break
                            dst.send(data)
                    except:
                        pass
                    finally:
                        try:
                            src.close()
                        except:
                            pass
                        try:
                            dst.close()
                        except:
                            pass

                t1 = threading.Thread(
                    target=forward_data, 
                    args=(client_socket, channel, "C→R")
                )
                t2 = threading.Thread(
                    target=forward_data, 
                    args=(channel, client_socket, "R→C")
                )
                t1.daemon = True
                t2.daemon = True
                t1.start()
                t2.start()

            except Exception as e:
                try:
                    client_socket.close()
                except:
                    pass

        # Thread de aceitação
        def accept_connections(server_socket, transport, remote_host, remote_port):
            """Accepts connections"""
            try:
                while True:
                    try:
                        client_socket, addr = server_socket.accept()
                        thread = threading.Thread(
                            target=forward_tunnel,
                            args=(client_socket, remote_host, remote_port, transport)
                        )
                        thread.daemon = True
                        thread.start()
                    except socket.timeout:
                        continue
                    except:
                        break
            except:
                pass
            finally:
                try:
                    server_socket.close()
                except:
                    pass

        # Iniciar thread
        transport = ssh_proxy.get_transport()
        accept_thread = threading.Thread(
            target=accept_connections,
            args=(server_socket, transport, equipamento_host, equipamento_porta)
        )
        accept_thread.daemon = True
        accept_thread.start()

        print(f"✅ Túnel criado!")
        time.sleep(0.5)

        return {
            'tunnel': None,
            'ssh_client': ssh_proxy,
            'local_host': '127.0.0.1',
            'local_port': local_port,
            'server_socket': server_socket,
            'channel': None,
            'transport': transport,
            'accept_thread': accept_thread
        }

    except Exception as e:
        print(f"❌ Erro: {str(e)}")
        raise


def registrar_erro_backup(acesso, usuario, erro, duracao):
    """Registra erro no log"""
    BackupLog.objects.create(
        acesso=acesso,
        cliente=acesso.cliente,
        template=acesso.backup_template,
        arquivo_path='',
        tamanho_bytes=0,
        status='ERRO',
        mensagem=erro,
        executado_por=usuario,
        duracao_segundos=duracao
    )

@login_required(login_url='login')
@modulo_habilitado_required('backups')
def listar_backups_cliente(request):
    """
    Lista backups de um cliente (AJAX)
    ✅ CORRIGIDO: Verifica se arquivo existe antes de exibir
    """
    cliente_id = request.GET.get('id')

    if not cliente_id:
        return JsonResponse({'error': 'Cliente não especificado'}, status=400)

    # Verificar permissão
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    # Buscar backups
    backups = BackupLog.objects.filter(cliente=cliente).select_related(
        'acesso', 'template', 'executado_por'
    ).order_by('-data_backup')

    # ✅ PASSO 1: Verificar quais arquivos existem e quais são órfãos
    backups_validos = []
    backups_para_deletar = []

    for backup in backups:
        if not backup.arquivo_path:
            # sem arquivo — só mantém se não for SEM_MUDANCAS legado
            if backup.status == 'SEM_MUDANCAS':
                continue  # registros legados SEM_MUDANCAS ignorados
            backups_validos.append(backup)
            continue
        arquivo_path = os.path.join(settings.MEDIA_ROOT, backup.arquivo_path)
        if os.path.exists(arquivo_path):
            backups_validos.append(backup)
        else:
            # Arquivo foi deletado manualmente da VM
            backups_para_deletar.append(backup.id)
            print(f"⚠️ Backup órfão: {backup.arquivo_path}")

    # ✅ PASSO 2: Remover registros órfãos do banco
    if backups_para_deletar:
        BackupLog.objects.filter(id__in=backups_para_deletar).delete()
        print(f"✅ {len(backups_para_deletar)} registro(s) órfão(s) removido(s)")

    # ✅ PASSO 3: Retornar apenas backups válidos
    return JsonResponse({
        'backups': [{
            'id': backup.id,
            'acesso_tipo': backup.acesso.tipo,
            'acesso_host': backup.acesso.host,
            'template': backup.template.nome if backup.template else 'N/A',
            'status': backup.get_status_display(),
            'status_code': backup.status,
            'tamanho': backup.get_tamanho_formatado() if backup.tamanho_bytes else '-',
            'data': backup.data_backup.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M:%S'),
            'duracao': f"{backup.duracao_segundos:.2f}s",
            'executado_por': backup.executado_por.username if backup.executado_por else 'Sistema',
            'mensagem': backup.mensagem or '',
            'arquivo_path': backup.arquivo_path,
            'tem_arquivo': bool(backup.arquivo_path),
            'hash': backup.hash_conteudo or '',
            'ultima_verificacao': backup.ultima_verificacao.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M:%S') if backup.ultima_verificacao else None,
        } for backup in backups_validos]
    })


@login_required(login_url='login')
@login_required
def exportar_senhas_pdf(request, cliente_id):
    """Gera PDF com todos os acessos e credenciais do cliente. Apenas superusuários.
    ?root=1  inclui Senha Root; ?root=0 (padrão) omite."""
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Acesso negado.')

    incluir_root = request.GET.get('root', '0') == '1'

    cliente = get_object_or_404(Cliente, id=cliente_id)
    acessos = Acesso.objects.filter(cliente=cliente).select_related('funcao', 'modelo').order_by('tipo')

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    from django.utils import timezone as tz

    buf = BytesIO()
    pagesize = landscape(A4)
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    cor_accent    = colors.HexColor('#1d4ed8')
    cor_linha_par = colors.HexColor('#f1f5f9')

    style_titulo = ParagraphStyle('titulo', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#0d1829'), spaceAfter=4)
    style_sub = ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748b'), spaceAfter=12)
    style_cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)
    style_mono = ParagraphStyle('mono', parent=styles['Normal'],
        fontSize=8, leading=10, fontName='Courier')

    elements = []

    # Cabeçalho
    tipo_export = 'Com Senha Root' if incluir_root else 'Sem Senha Root'
    elements.append(Paragraph('Credenciais de Acesso', style_titulo))
    elements.append(Paragraph(
        f'{cliente.nome_empresa} &nbsp;|&nbsp; CNPJ: {cliente.cnpj or "—"} &nbsp;|&nbsp; '
        f'Gerado em: {tz.localtime(tz.now()).strftime("%d/%m/%Y %H:%M")} &nbsp;|&nbsp; {tipo_export}',
        style_sub
    ))
    elements.append(Spacer(1, 0.3*cm))

    # Largura útil da página (A4 paisagem, 1.5cm de margem em cada lado) = 26.7cm.
    # As larguras abaixo somam menos que isso para evitar que a tabela seja
    # cortada nas laterais do PDF.
    if incluir_root:
        header     = ['Descrição', 'Host', 'Proto', 'Porta', 'Usuário', 'Senha', 'Senha Root', 'Função']
        col_widths = [4.2*cm, 4.5*cm, 1.6*cm, 1.3*cm, 3.6*cm, 3.8*cm, 3.8*cm, 3.5*cm]
    else:
        header     = ['Descrição', 'Host', 'Proto', 'Porta', 'Usuário', 'Senha', 'Função']
        col_widths = [5.0*cm, 5.5*cm, 1.8*cm, 1.5*cm, 4.3*cm, 4.5*cm, 4.0*cm]

    data = [header]
    for ac in acessos:
        row = [
            Paragraph(ac.tipo or '—', style_cell),
            Paragraph(ac.host or '—', style_mono),
            Paragraph(ac.protocolo or '—', style_cell),
            Paragraph(str(ac.porta) if ac.porta else '—', style_cell),
            Paragraph(ac.usuario or '—', style_mono),
            Paragraph(ac.senha or '—', style_mono),
        ]
        if incluir_root:
            row.append(Paragraph(ac.senha_adm or '—', style_mono))
        row.append(Paragraph(ac.funcao.descricao if ac.funcao else '—', style_cell))
        data.append(row)

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), cor_accent),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING',    (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, cor_linha_par]),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    sufixo = '_com_root' if incluir_root else '_sem_root'
    nome_arquivo = f"senhas_{cliente.nome_empresa.replace(' ', '_')[:35]}{sufixo}.pdf"
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


@login_required(login_url='login')
@login_required
def exportar_senhas_txt(request, cliente_id):
    """Gera arquivo .txt com todos os acessos e credenciais do cliente. Apenas superusuários.
    ?root=1  inclui Senha Root; ?root=0 (padrão) omite."""
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Acesso negado.')

    incluir_root = request.GET.get('root', '0') == '1'

    cliente = get_object_or_404(Cliente, id=cliente_id)
    acessos = Acesso.objects.filter(cliente=cliente).select_related('funcao', 'modelo').order_by('tipo')

    from django.utils import timezone as tz

    tipo_export = 'Com Senha Root' if incluir_root else 'Sem Senha Root'
    linhas = [
        'CREDENCIAIS DE ACESSO',
        f'{cliente.nome_empresa} | CNPJ: {cliente.cnpj or "—"}',
        f'Gerado em: {tz.localtime(tz.now()).strftime("%d/%m/%Y %H:%M")} | {tipo_export}',
        '=' * 70,
        '',
    ]

    for ac in acessos:
        linhas.append(f'Descrição : {ac.tipo or "—"}')
        linhas.append(f'Host      : {ac.host or "—"}')
        linhas.append(f'Protocolo : {ac.protocolo or "—"}')
        linhas.append(f'Porta     : {ac.porta if ac.porta else "—"}')
        linhas.append(f'Usuário   : {ac.usuario or "—"}')
        linhas.append(f'Senha     : {ac.senha or "—"}')
        if incluir_root:
            linhas.append(f'Senha Root: {ac.senha_adm or "—"}')
        linhas.append(f'Função    : {ac.funcao.descricao if ac.funcao else "—"}')
        linhas.append('-' * 70)

    conteudo = '\n'.join(linhas) + '\n'

    sufixo = '_com_root' if incluir_root else '_sem_root'
    nome_arquivo = f"senhas_{cliente.nome_empresa.replace(' ', '_')[:35]}{sufixo}.txt"
    response = HttpResponse(conteudo, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


@modulo_habilitado_required('backups')
def download_backup(request, backup_id):
    """Download de backup"""
    try:
        backup = BackupLog.objects.get(id=backup_id)

        # Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, backup.cliente):
            messages.error(request, 'Sem permissão')
            return redirect('listar_clientes')

        arquivo_path = os.path.join(settings.MEDIA_ROOT, backup.arquivo_path)

        if not os.path.exists(arquivo_path):
            messages.error(request, 'Arquivo não encontrado')
            return redirect('listar_clientes')

        return FileResponse(
            open(arquivo_path, 'rb'),
            as_attachment=True,
            filename=os.path.basename(arquivo_path)
        )

    except BackupLog.DoesNotExist:
        messages.error(request, 'Backup não encontrado')
        return redirect('listar_clientes')
    except Exception as e:
        messages.error(request, f'Erro: {str(e)}')
        return redirect('listar_clientes')

@login_required(login_url='login')
@modulo_habilitado_required('backups')
def deletar_backup(request, backup_id):
    """Deleta backup"""
    if request.method == 'POST':
        try:
            backup = get_object_or_404(BackupLog, id=backup_id)

            if not _perms.pode_acessar_cliente(request.user, backup.cliente):
                messages.error(request, 'Sem permissão')
                return redirect('listar_clientes')

            cliente_id = backup.cliente.id

            # Deletar arquivo
            arquivo_path = os.path.join(settings.MEDIA_ROOT, backup.arquivo_path)
            if os.path.exists(arquivo_path):
                os.remove(arquivo_path)

            # Deletar registro
            backup.delete()

            messages.success(request, 'Backup excluído!')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        except Exception as e:
            messages.error(request, f'Erro: {str(e)}')
            return redirect('listar_clientes')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('backups')
def buscar_templates_backup(request):
    """Busca templates de backup"""
    templates = BackupTemplate.objects.filter(ativo=True).order_by('fabricante', 'nome')

    return JsonResponse({
        'templates': [{
            'id': t.id,
            'nome': t.nome,
            'fabricante': t.get_fabricante_display(),
            'descricao': t.descricao or ''
        } for t in templates]
    })

@login_required(login_url='login')
@modulo_habilitado_required('backups')
def backup_conteudo(request, backup_id):
    """Retorna conteúdo do arquivo de backup para visualização no modal."""
    try:
        backup = BackupLog.objects.select_related('acesso', 'cliente').get(id=backup_id)

        # Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, backup.cliente):
            return JsonResponse({'error': 'Sem permissão'}, status=403)

        if not backup.arquivo_path:
            return JsonResponse({'error': 'Este registro não possui arquivo (sem mudanças)'}, status=404)

        arquivo_path = os.path.join(settings.MEDIA_ROOT, backup.arquivo_path)
        if not os.path.exists(arquivo_path):
            return JsonResponse({'error': 'Arquivo não encontrado no servidor'}, status=404)

        tamanho = os.path.getsize(arquivo_path)
        LIMITE = 512 * 1024  # 512 KB
        with open(arquivo_path, 'r', encoding='utf-8', errors='replace') as f:
            conteudo = f.read(LIMITE)

        truncado = tamanho > LIMITE
        return JsonResponse({
            'conteudo': conteudo,
            'nome': os.path.basename(arquivo_path),
            'tamanho': backup.get_tamanho_formatado(),
            'truncado': truncado,
            'hash': backup.hash_conteudo or '',
        })

    except BackupLog.DoesNotExist:
        return JsonResponse({'error': 'Backup não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def terminal_page(request):
    """Renderiza a página de terminal SSH múltiplo"""
    from wiki.models import ArtigoWiki, CategoriaWiki
    return render(request, 'terminal.html', {
        'wiki_categorias': CategoriaWiki.objects.all(),
        'wiki_fabricantes': ArtigoWiki.FABRICANTES,
    })


@require_http_methods(["GET"])
def terminal_link_externo(request, token):
    """
    Página PÚBLICA (sem login) para um visitante externo — ex: suporte de
    fabricante numa chamada — acessar um terminal compartilhado por um
    usuário do CRM através de um link temporário (TerminalLinkExterno).

    A autorização inteira é o token (UUID imprevisível na URL); não expõe
    nada do resto do CRM (sem sidebar de hosts, sem outros clientes) — só
    o único terminal daquele Acesso, e só enquanto o link for válido.
    """
    try:
        link = TerminalLinkExterno.objects.select_related('acesso').get(id=token)
    except TerminalLinkExterno.DoesNotExist:
        return render(request, 'terminal_link_invalido.html', {'motivo': 'Link inválido.'}, status=404)

    valido, motivo = link.validar()
    if not valido:
        return render(request, 'terminal_link_invalido.html', {'motivo': motivo}, status=410)

    acesso = link.acesso
    return render(request, 'terminal_externo.html', {
        'token': str(link.id),
        'host_label': f'{acesso.tipo or acesso.protocolo} — {acesso.host}',
        'expira_em': link.expira_em.isoformat(),
    })


@login_required(login_url='login')
@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def listar_acessos_terminal(request):
    """Retorna JSON com acessos SSH/Telnet. Filtra por cliente_id se informado."""
    cliente_id = request.GET.get('cliente')

    base_qs = Acesso.objects.select_related('cliente', 'funcao').filter(
        protocolo__iregex=r'^(ssh|telnet|rlogin)$'
    )

    if cliente_id:
        # Filtrar pelo cliente especificado — verificar permissão de acesso
        cliente_obj = get_object_or_404(Cliente, id=cliente_id)
        if not _perms.pode_acessar_cliente(request.user, cliente_obj):
            return JsonResponse({'acessos': []})
        acessos = base_qs.filter(cliente=cliente_obj).order_by('tipo')
    elif _perms.is_admin(request.user):
        acessos = base_qs.order_by('cliente__nome_empresa', 'tipo')
    elif _perms.is_backoffice(request.user):
        # Consultor/Operador: só acessos dos clientes da própria instância.
        acessos = base_qs.filter(
            cliente__in=Cliente.objects.visiveis_para(request.user)
        ).order_by('cliente__nome_empresa', 'tipo')
    else:
        try:
            cliente_obj = Cliente.objects.get_by_usuario_vinculado(request.user)
            acessos = base_qs.filter(cliente=cliente_obj).order_by('tipo')
        except Cliente.DoesNotExist:
            return JsonResponse({'acessos': []})

    data = [
        {
            'id': a.id,
            'tipo': a.tipo,
            'host': a.host,
            'porta': a.porta or 22,
            'protocolo': a.protocolo,
            'usuario': a.usuario,
            'cliente_nome': a.cliente.nome_empresa,
            'cliente_id': a.cliente.id,
            'funcao': a.funcao.descricao if a.funcao else '',
            'winbox': a.winbox or 0,
        }
        for a in acessos
    ]
    return JsonResponse({'acessos': data})


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def winbox_page(request, acesso_id):
    """Renderiza a página WebFig (interface web MikroTik) via proxy"""
    acesso = get_object_or_404(Acesso, id=acesso_id)

    # Verificar permissões
    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    # Determinar porta WebFig (80 = HTTP padrão do RouterOS)
    # O campo 'winbox' armazena a porta do serviço Winbox (8291),
    # mas a interface web (WebFig) roda na porta HTTP (80 por padrão)
    webfig_porta = 80
    
    # Construir URL do proxy interno que já existe no sistema
    webfig_url = f'/clientes/acessos/{acesso.id}/web/{webfig_porta}/http/'
    
    context = {
        'acesso': acesso,
        'webfig_url': webfig_url,
        'webfig_porta': webfig_porta,
        'vnc_mode': 'winbox',
    }
    
    return render(request, 'winbox.html', context)

@modulo_habilitado_required('acessos')
def webfig_vnc_page(request, acesso_id):
    """Renderiza a página WebFig via VNC (Browser no servidor)"""
    acesso = get_object_or_404(Acesso, id=acesso_id)

    # Verificar permissões
    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    context = {
        'acesso': acesso,
        'vnc_mode': 'browser',
    }
    
    return render(request, 'winbox.html', context)




# ============================================
# FUNÇÕES DE PING - ADICIONAR AO views.py
# ============================================

@login_required(login_url='login')
@modulo_habilitado_required('tuneis')
def proxy_ativo_cliente(request):
    """Indica se o cliente tem um ProxyServer SSH ativo — usado para validar
    backup/acesso a equipamentos com IP privado antes de tentar a operação."""
    cliente_id = request.GET.get('cliente_id')
    if not cliente_id:
        return JsonResponse({'tem_proxy_ativo': False})

    proxy = ProxyServer.objects.filter(cliente_id=cliente_id, ativo=True).first()
    return JsonResponse({
        'tem_proxy_ativo': bool(proxy),
        'nome': proxy.nome if proxy else None,
        'host': proxy.host if proxy else None,
    })


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def ping_acesso(request, acesso_id):
    """Realiza ping para um acesso (via proxy se necessário)"""
    try:
        acesso = Acesso.objects.get(id=acesso_id)

        # ✅ Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
            return JsonResponse({'error': 'Sem permissão'}, status=403)

        host = acesso.host
        eh_privado = is_private_ip(host)

        print(f"\n{'='*80}")
        print(f"🔍 PING REQUEST")
        print(f"{'='*80}")
        print(f"Host: {host}")
        print(f"IP Privado? {eh_privado}")

        # ✅ Se IP privado, executar via proxy
        if eh_privado:
            print(f"⚠️ IP PRIVADO - Usando proxy SSH")

            proxy = ProxyServer.objects.filter(
                cliente=acesso.cliente,
                ativo=True
            ).first()

            if not proxy:
                if vpn_cobre_ip(acesso.cliente, host):
                    print(f"✅ VPN WireGuard cobre {host} — ping direto via VPN")
                    resultado = ping_direto(host)
                else:
                    return JsonResponse({
                        'error': 'IP privado sem proxy SSH ativo',
                        'host': host,
                        'status': 'erro'
                    }, status=400)
            else:
                resultado = ping_via_proxy(proxy, host)
        else:
            # ✅ IP público, ping direto
            print(f"✅ IP PÚBLICO - Ping direto")
            resultado = ping_direto(host)

        print(f"{'='*80}")
        print(f"Resultado: {resultado}")
        print(f"{'='*80}\n")

        return JsonResponse(resultado)

    except Acesso.DoesNotExist:
        return JsonResponse({'error': 'Acesso não encontrado'}, status=404)
    except Exception as e:
        print(f"❌ ERRO: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def ping_direto(host, packets=10):
    """
    ✅ Ping direto para IP público
    Compatível com Linux e Windows
    """
    try:
        import subprocess
        import platform

        # ✅ Detectar sistema operacional
        sistema = platform.system()

        if sistema == 'Windows':
            cmd = ['ping', '-n', str(packets), host]
        else:  # Linux/Mac
            cmd = ['ping', '-c', str(packets), host]

        print(f"📤 Executando: {' '.join(cmd)}")

        # ✅ Executar ping com timeout
        resultado = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=30
        )

        output = resultado.stdout
        return_code = resultado.returncode

        print(f"Return code: {return_code}")
        print(f"Output:\n{output}")

        # ✅ Parsear resultado
        return parsear_output_ping(output, host, return_code)

    except subprocess.TimeoutExpired:
        return {
            'host': host,
            'status': 'timeout',
            'mensagem': 'Timeout ao executar ping',
            'packets_enviados': packets,
            'packets_recebidos': 0,
            'pacotes_perdidos': packets
        }
    except Exception as e:
        return {
            'host': host,
            'status': 'erro',
            'mensagem': f'Erro ao executar ping: {str(e)}',
            'packets_enviados': packets,
            'packets_recebidos': 0,
            'pacotes_perdidos': packets
        }


def ping_via_proxy(proxy, host, packets=10):
    """
    ✅ Ping via proxy SSH (para IPs privados)
    """
    try:
        print(f"📡 Ping via proxy SSH")

        # ✅ Conectar ao proxy
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        ssh_client.connect(
            hostname=proxy.host,
            port=proxy.porta,
            username=proxy.usuario,
            password=proxy.senha,
            timeout=10,
            look_for_keys=False,
            allow_agent=False
        )

        print(f"✅ Conectado ao proxy")

        # ✅ Executar ping no servidor remoto
        cmd_ping = f'ping -c {packets} {host}'
        print(f"📤 Executando no proxy: {cmd_ping}")

        stdin, stdout, stderr = ssh_client.exec_command(cmd_ping, timeout=30)

        output = stdout.read().decode('utf-8', errors='ignore')
        error = stderr.read().decode('utf-8', errors='ignore')

        ssh_client.close()

        print(f"Output:\n{output}")
        if error:
            print(f"Error:\n{error}")

        # ✅ Parsear resultado
        return parsear_output_ping(output, host, 0)

    except Exception as e:
        print(f"❌ Erro ping via proxy: {str(e)}")
        return {
            'host': host,
            'status': 'erro',
            'mensagem': f'Erro ao executar ping via proxy: {str(e)}',
            'packets_enviados': packets,
            'packets_recebidos': 0,
            'pacotes_perdidos': packets
        }


def parsear_output_ping(output, host, return_code):
    """
    ✅ Parser universal para output de ping
    Funciona com Linux, Mac e Windows
    """
    try:
        import re

        # ✅ Se return_code é diferente de 0, host não respondeu
        if return_code != 0 and 'transmitted' not in output.lower():
            return {
                'host': host,
                'status': 'inalcancavel',
                'mensagem': f'Host {host} não está alcançável (sem resposta)',
                'packets_enviados': 0,
                'packets_recebidos': 0,
                'pacotes_perdidos': 0,
                'output': output[:500]
            }

        # ✅ Padrão Linux: "10 packets transmitted, 10 received, 0% packet loss"
        linux_pattern = r'(\d+)\s+packets? transmitted[,.]?\s+(\d+)\s+(?:packets? )?received'

        # ✅ Procurar padrão Linux
        match_linux = re.search(linux_pattern, output)
        if match_linux:
            enviados = int(match_linux.group(1))
            recebidos = int(match_linux.group(2))
            perdidos = enviados - recebidos

            # ✅ Procurar tempo
            time_pattern = r'min/avg/max(?:/stddev)?\s*=\s*([0-9.]+)/([0-9.]+)/([0-9.]+)'
            match_time = re.search(time_pattern, output)

            tempos = {}
            if match_time:
                tempos = {
                    'min': float(match_time.group(1)),
                    'avg': float(match_time.group(2)),
                    'max': float(match_time.group(3))
                }

            return {
                'host': host,
                'status': 'sucesso' if recebidos > 0 else 'timeout',
                'packets_enviados': enviados,
                'packets_recebidos': recebidos,
                'pacotes_perdidos': perdidos,
                'percentual_perda': (perdidos / enviados * 100) if enviados > 0 else 100,
                'tempos': tempos,
                'mensagem': f'{recebidos}/{enviados} packets recebidos' if recebidos > 0 else 'Sem resposta'
            }

        # ✅ Se não encontrou padrão
        return {
            'host': host,
            'status': 'desconhecido',
            'mensagem': 'Não foi possível parsear resultado do ping',
            'output': output[:300]
        }

    except Exception as e:
        print(f"❌ Erro ao parsear: {str(e)}")
        return {
            'host': host,
            'status': 'erro',
            'mensagem': f'Erro ao parsear resultado: {str(e)}',
            'output': output[:300]
        }


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def traceroute_acesso(request, acesso_id):
    """Executa traceroute para um acesso (direto ou via proxy SSH)."""
    import subprocess, platform

    try:
        acesso = Acesso.objects.get(id=acesso_id)
    except Acesso.DoesNotExist:
        return JsonResponse({'error': 'Acesso não encontrado'}, status=404)

    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    host = acesso.host
    eh_privado = is_private_ip(host)

    try:
        if eh_privado:
            proxy = ProxyServer.objects.filter(cliente=acesso.cliente, ativo=True).first()
            if not proxy:
                if vpn_cobre_ip(acesso.cliente, host):
                    resultado = _traceroute_direto(host)
                else:
                    return JsonResponse({'error': 'IP privado sem proxy SSH ativo', 'host': host}, status=400)
            else:
                resultado = _traceroute_via_proxy(proxy, host)
        else:
            resultado = _traceroute_direto(host)

        return JsonResponse(resultado)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def _traceroute_direto(host):
    import subprocess, platform, re
    sistema = platform.system()
    if sistema == 'Windows':
        cmd = ['tracert', '-d', '-h', '20', host]
    else:
        # prefere mtr, cai para traceroute
        import shutil
        if shutil.which('mtr'):
            cmd = ['mtr', '--report', '--report-cycles', '3', '--no-dns', host]
        elif shutil.which('traceroute'):
            cmd = ['traceroute', '-n', '-w', '2', '-m', '20', host]
        else:
            cmd = ['tracepath', '-n', host]

    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        output = (proc.stdout + proc.stderr).strip()
    except subprocess.TimeoutExpired:
        return {'host': host, 'status': 'timeout', 'output': 'Timeout ao executar traceroute.'}
    except FileNotFoundError:
        return {'host': host, 'status': 'erro', 'output': 'Ferramenta de traceroute não encontrada no servidor.'}

    return {
        'host': host,
        'status': 'ok',
        'ferramenta': cmd[0],
        'output': output[:4000],
    }


def _traceroute_via_proxy(proxy, host):
    try:
        import paramiko
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=proxy.host, port=proxy.porta,
            username=proxy.usuario, password=proxy.senha,
            timeout=10, look_for_keys=False, allow_agent=False,
        )
        cmd = (
            f'which mtr > /dev/null 2>&1 && '
            f'mtr --report --report-cycles 3 --no-dns {host} 2>&1 || '
            f'traceroute -n -w 2 -m 20 {host} 2>&1'
        )
        _, stdout, stderr = ssh.exec_command(cmd, timeout=60)
        output = (stdout.read() + stderr.read()).decode('utf-8', errors='ignore').strip()
        ssh.close()
        ferramenta = 'mtr' if ('Loss%' in output or 'HOST' in output) else 'traceroute'
        return {'host': host, 'status': 'ok', 'ferramenta': ferramenta, 'output': output[:4000]}
    except Exception as e:
        return {'host': host, 'status': 'erro', 'output': f'Erro via proxy SSH: {e}'}


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def cadastrar_bloco_ip(request):
    """Cadastra um novo bloco de IP"""
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        tipo = request.POST.get('tipo')
        bloco = request.POST.get('bloco')
        asn = request.POST.get('asn')
        irr_registry = request.POST.get('irr_registry')

        # Validações básicas
        if not all([cliente_id, tipo, bloco]):
            messages.error(request, 'Preencha todos os campos obrigatórios.')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # Validar formato do bloco IP
        try:
            ipaddress.ip_network(bloco)
        except ValueError:
            messages.error(request, f'Formato de bloco IP inválido: {bloco}')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        # Criar bloco
        try:
            BlocoIP.objects.create(
                cliente_id=cliente_id,
                tipo=tipo,
                bloco=bloco,
                asn=asn,
                irr_registry=irr_registry
            )
            messages.success(request, f'Bloco {bloco} cadastrado com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao cadastrar bloco: {str(e)}')

        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def buscar_bloco_ip(request, bloco_id):
    """Busca dados de um bloco IP específico (AJAX)"""
    try:
        bloco = BlocoIP.objects.get(id=bloco_id)

        # Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, bloco.cliente):
            return JsonResponse({'error': 'Sem permissão'}, status=403)

        data = {
            'id': bloco.id,
            'tipo': bloco.tipo,
            'bloco': bloco.bloco,
            'asn': bloco.asn or '',
            'irr_registry': bloco.irr_registry or '',
            'rpki_valido': bloco.rpki_valido,
            'irr_valido': bloco.irr_valido,
            'rpki_status': bloco.get_status_rpki_display(),
            'irr_status': bloco.get_status_irr_display(),
            'ultima_validacao': bloco.ultima_validacao.strftime('%d/%m/%Y %H:%M:%S') if bloco.ultima_validacao else 'Nunca',
            'rpki_mensagem': bloco.rpki_mensagem or '',
            'irr_mensagem': bloco.irr_mensagem or ''
        }

        return JsonResponse(data)

    except BlocoIP.DoesNotExist:
        return JsonResponse({'error': 'Bloco não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def editar_bloco_ip(request, bloco_id):
    """Edita um bloco IP existente"""
    if request.method == 'POST':
        try:
            bloco = get_object_or_404(BlocoIP, id=bloco_id)

            # Verificar permissão
            if not _perms.pode_acessar_cliente(request.user, bloco.cliente):
                messages.error(request, 'Sem permissão')
                return redirect('listar_clientes')

            bloco.bloco = request.POST.get('bloco')
            bloco.tipo = request.POST.get('tipo')
            bloco.asn = request.POST.get('asn')
            bloco.irr_registry = request.POST.get('irr_registry')

            # Validar formato do bloco IP
            try:
                ipaddress.ip_network(bloco.bloco)
            except ValueError:
                messages.error(request, f'Formato de bloco IP inválido: {bloco.bloco}')
                return redirect(reverse('listar_clientes') + f'?id={bloco.cliente.id}')

            bloco.save()

            messages.success(request, f'Bloco {bloco.bloco} atualizado com sucesso!')
            return redirect(reverse('listar_clientes') + f'?id={bloco.cliente.id}')

        except Exception as e:
            messages.error(request, f'Erro ao editar bloco: {str(e)}')
            return redirect('listar_clientes')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def deletar_bloco_ip(request, bloco_id):
    """Deleta um bloco IP"""
    if request.method == 'POST':
        bloco = get_object_or_404(BlocoIP, id=bloco_id)
        cliente_id = bloco.cliente.id
        bloco_texto = bloco.bloco

        # Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, bloco.cliente):
            messages.error(request, 'Sem permissão para deletar este bloco IP')
            return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

        bloco.delete()

        messages.success(request, f'Bloco {bloco_texto} excluído com sucesso!')
        return redirect(reverse('listar_clientes') + f'?id={cliente_id}')

    return redirect('listar_clientes')


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def validar_bloco_rpki_irr(request, bloco_id):
    """Executa validação RPKI/IRR manual para um bloco"""
    try:
        bloco = BlocoIP.objects.get(id=bloco_id)

        # Verificar permissão
        if not _perms.pode_acessar_cliente(request.user, bloco.cliente):
            return JsonResponse({'error': 'Sem permissão'}, status=403)

        # Executar validação
        resultado = executar_validacao_rpki_irr(bloco)

        if resultado['sucesso']:
            return JsonResponse({
                'success': True,
                'rpki_status': bloco.get_status_rpki_display(),
                'rpki_valido': bloco.rpki_valido,
                'rpki_mensagem': bloco.rpki_mensagem or '',
                'irr_status': bloco.get_status_irr_display(),
                'irr_valido': bloco.irr_valido,
                'irr_mensagem': bloco.irr_mensagem or '',
                'ultima_validacao': bloco.ultima_validacao.strftime('%d/%m/%Y %H:%M:%S')
            })
        else:
            return JsonResponse({
                'error': resultado['erro']
            }, status=500)

    except BlocoIP.DoesNotExist:
        return JsonResponse({'error': 'Bloco não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def listar_blocos_cliente(request):
    """Lista blocos IP de um cliente (AJAX)"""
    cliente_id = request.GET.get('id')

    if not cliente_id:
        return JsonResponse({'error': 'Cliente não especificado'}, status=400)

    # Verificar permissão
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    # Buscar blocos
    blocos = BlocoIP.objects.filter(cliente=cliente).order_by('tipo', 'bloco')

    return JsonResponse({
        'blocos': [{
            'id': bloco.id,
            'tipo': bloco.get_tipo_display(),
            'tipo_code': bloco.tipo,
            'bloco': bloco.bloco,
            'asn': bloco.asn or 'N/A',
            'irr_registry': bloco.irr_registry or 'N/A',
            'rpki_valido': bloco.rpki_valido,
            'rpki_status': bloco.get_status_rpki_display(),
            'rpki_mensagem': bloco.rpki_mensagem or '',
            'irr_valido': bloco.irr_valido,
            'irr_status': bloco.get_status_irr_display(),
            'irr_mensagem': bloco.irr_mensagem or '',
            'ultima_validacao': bloco.ultima_validacao.strftime('%d/%m/%Y %H:%M:%S') if bloco.ultima_validacao else 'Nunca',
            'data_criacao': bloco.data_criacao.strftime('%d/%m/%Y %H:%M')
        } for bloco in blocos]
    })


# ============================================
# AMPSCAN — VARREDURA DE AMPLIFICAÇÃO DDoS
# ============================================

@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def listar_ampscan_resultados(request):
    """Lista os achados atuais (não resolvidos) da varredura de amplificação
    de um cliente (AJAX)."""
    from .models import AmpScanResultado

    cliente_id = request.GET.get('id')
    if not cliente_id:
        return JsonResponse({'error': 'Cliente não especificado'}, status=400)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    resultados = AmpScanResultado.objects.filter(cliente=cliente, resolvido=False).select_related('bloco_ip').order_by('status', '-ultima_deteccao')

    return JsonResponse({
        'resultados': [{
            'id': r.id,
            'ip': r.ip,
            'porta': r.porta,
            'protocolo': r.protocolo,
            'servico': r.servico,
            'descricao_risco': r.descricao_risco,
            'status': r.status,
            'status_display': r.get_status_display(),
            'tempo_resposta_ms': r.tempo_resposta_ms,
            'bloco': r.bloco_ip.bloco if r.bloco_ip else None,
            'primeira_deteccao': timezone.localtime(r.primeira_deteccao).strftime('%d/%m/%Y %H:%M'),
            'ultima_deteccao': timezone.localtime(r.ultima_deteccao).strftime('%d/%m/%Y %H:%M'),
        } for r in resultados]
    })


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def listar_ampscan_execucoes(request):
    """Últimas execuções da varredura de amplificação de um cliente (AJAX) —
    usado tanto para mostrar 'última varredura' quanto para o front-end
    fazer polling enquanto uma varredura sob demanda está rodando."""
    from .models import AmpScanExecucaoLog

    cliente_id = request.GET.get('id')
    if not cliente_id:
        return JsonResponse({'error': 'Cliente não especificado'}, status=400)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    execucoes = AmpScanExecucaoLog.objects.filter(cliente=cliente).order_by('-iniciado_em')[:5]

    return JsonResponse({
        'execucoes': [{
            'id': e.id,
            'iniciado_em': timezone.localtime(e.iniciado_em).strftime('%d/%m/%Y %H:%M:%S'),
            'finalizado_em': timezone.localtime(e.finalizado_em).strftime('%d/%m/%Y %H:%M:%S') if e.finalizado_em else None,
            'em_andamento': e.finalizado_em is None,
            'total_ips': e.total_ips,
            'total_probes': e.total_probes,
            'total_vulneraveis': e.total_vulneraveis,
            'total_protegidos': e.total_protegidos,
            'total_expostos': e.total_expostos,
            'blocos_ignorados': e.blocos_ignorados,
            'sucesso': e.sucesso,
            'erro_mensagem': e.erro_mensagem,
        } for e in execucoes]
    })


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
@require_http_methods(['POST'])
def ampscan_escanear_agora(request):
    """Dispara a varredura de amplificação sob demanda para um cliente
    (assíncrona via Celery — a varredura de um /24 inteiro pode levar dezenas
    de segundos, não dá pra rodar no ciclo de vida da request)."""
    from .models import BlocoIP
    from .tasks import ampscan_escanear_cliente

    cliente_id = request.POST.get('id')
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    if not BlocoIP.objects.filter(cliente=cliente).exists():
        return JsonResponse({'error': 'Cliente não tem blocos de IP cadastrados (RPKI/IRR).'}, status=400)

    ampscan_escanear_cliente.delay(cliente.id)
    return JsonResponse({'success': True})


# ============================================
# ROTALOOP — DETECÇÃO DE LOOP DE ROTEAMENTO
# ============================================

@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def listar_rotaloop_resultados(request):
    """Lista os loops de roteamento atuais (não resolvidos) de um cliente (AJAX)."""
    from .models import RotaLoopResultado

    cliente_id = request.GET.get('id')
    if not cliente_id:
        return JsonResponse({'error': 'Cliente não especificado'}, status=400)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    resultados = RotaLoopResultado.objects.filter(cliente=cliente, resolvido=False).select_related('bloco_ip').order_by('-ultima_deteccao')

    return JsonResponse({
        'resultados': [{
            'id': r.id,
            'bloco': r.bloco_ip.bloco,
            'ip_alvo': r.ip_alvo,
            'ip_em_loop': r.ip_em_loop,
            'ferramenta': r.ferramenta,
            'hops': r.hops,
            'primeira_deteccao': timezone.localtime(r.primeira_deteccao).strftime('%d/%m/%Y %H:%M'),
            'ultima_deteccao': timezone.localtime(r.ultima_deteccao).strftime('%d/%m/%Y %H:%M'),
        } for r in resultados]
    })


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
def listar_rotaloop_execucoes(request):
    """Últimas execuções do teste de loop de roteamento de um cliente (AJAX)."""
    from .models import RotaLoopExecucaoLog

    cliente_id = request.GET.get('id')
    if not cliente_id:
        return JsonResponse({'error': 'Cliente não especificado'}, status=400)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    execucoes = RotaLoopExecucaoLog.objects.filter(cliente=cliente).order_by('-iniciado_em')[:5]

    return JsonResponse({
        'execucoes': [{
            'id': e.id,
            'iniciado_em': timezone.localtime(e.iniciado_em).strftime('%d/%m/%Y %H:%M:%S'),
            'finalizado_em': timezone.localtime(e.finalizado_em).strftime('%d/%m/%Y %H:%M:%S') if e.finalizado_em else None,
            'em_andamento': e.finalizado_em is None,
            'total_blocos_testados': e.total_blocos_testados,
            'total_loops_detectados': e.total_loops_detectados,
            'sucesso': e.sucesso,
            'erro_mensagem': e.erro_mensagem,
        } for e in execucoes]
    })


@login_required(login_url='login')
@modulo_habilitado_required('rpki_irr')
@require_http_methods(['POST'])
def rotaloop_testar_agora(request):
    """Dispara o teste de loop de roteamento sob demanda para um cliente
    (assíncrono via Celery)."""
    from .models import BlocoIP
    from .tasks import rotaloop_testar_cliente

    cliente_id = request.POST.get('id')
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    if not BlocoIP.objects.filter(cliente=cliente).exists():
        return JsonResponse({'error': 'Cliente não tem blocos de IP cadastrados (RPKI/IRR).'}, status=400)

    rotaloop_testar_cliente.delay(cliente.id)
    return JsonResponse({'success': True})


# ============================================
# FUNÇÕES DE VALIDAÇÃO RPKI/IRR
# ============================================

def executar_validacao_rpki_irr(bloco):
    """
    Executa validação RPKI e IRR para um bloco IP
    """
    import time
    inicio = time.time()

    try:
        print(f"\n{'='*80}")
        print(f"🔍 INICIANDO VALIDAÇÃO RPKI/IRR")
        print(f"{'='*80}")
        print(f"Bloco: {bloco.bloco}")
        print(f"Tipo: {bloco.get_tipo_display()}")
        print(f"ASN: {bloco.asn}")
        print(f"IRR Registry: {bloco.irr_registry}")

        # ✅ VALIDAÇÃO RPKI
        rpki_resultado = validar_rpki(bloco.bloco, bloco.asn)

        # ✅ VALIDAÇÃO IRR
        irr_resultado = validar_irr(bloco.bloco, bloco.asn, bloco.irr_registry)

        # ✅ Atualizar bloco
        bloco.rpki_valido = rpki_resultado['valido']
        bloco.rpki_status = rpki_resultado['status']
        bloco.rpki_mensagem = rpki_resultado['mensagem']

        bloco.irr_valido = irr_resultado['valido']
        bloco.irr_status = irr_resultado['status']
        bloco.irr_mensagem = irr_resultado['mensagem']

        bloco.ultima_validacao = datetime.now()
        bloco.save()

        duracao = time.time() - inicio

        # ✅ Registrar log
        ValidacaoRPKI_IRR_Log.objects.create(
            bloco=bloco,
            rpki_valido=rpki_resultado['valido'],
            rpki_status=rpki_resultado['status'],
            rpki_detalhes=rpki_resultado['detalhes'],
            irr_valido=irr_resultado['valido'],
            irr_status=irr_resultado['status'],
            irr_detalhes=irr_resultado['detalhes'],
            duracao_segundos=duracao
        )

        print(f"\n{'='*80}")
        print(f"✅ VALIDAÇÃO CONCLUÍDA!")
        print(f"{'='*80}")
        print(f"RPKI: {rpki_resultado['status']} - {rpki_resultado['mensagem']}")
        print(f"IRR: {irr_resultado['status']} - {irr_resultado['mensagem']}")
        print(f"Duração: {duracao:.2f}s")
        print(f"{'='*80}\n")

        return {
            'sucesso': True,
            'rpki': rpki_resultado,
            'irr': irr_resultado,
            'duracao': duracao
        }

    except Exception as e:
        erro = f"Erro na validação: {str(e)}"
        print(f"\n❌ {erro}\n")
        return {'sucesso': False, 'erro': erro}

def validar_rpki(bloco, asn):
    """
    ✅ CORRIGIDO v6: Usa RIPE Stat RPKI Validation

    API: https://stat.ripe.net/data/rpki-validation/data.json
    Parâmetros:
    - resource: ASN (ex: 268858)
    - prefix: Bloco IP (ex: 45.174.160.0/23)

    Resposta:
    - status: "valid", "invalid", "unknown"
    - validating_roas: lista de ROAs que cobrem o prefixo

    🔑 Simples, eficiente e funciona!
    """
    print(f"\n{'='*60}")
    print(f"📡 VALIDANDO RPKI: {bloco} (AS{asn})")
    print(f"{'='*60}")

    if not asn:
        return {
            'valido': False,
            'status': 'NotChecked',
            'mensagem': 'ASN não informado',
            'detalhes': 'ASN é necessário para validação'
        }

    asn_limpo = asn.replace('AS', '').replace('as', '').strip()
    print(f"ASN limpo: {asn_limpo}")

    # ==========================
    # API PRIMÁRIA: RIPE Stat RPKI Validation
    # ==========================
    print(f"\n[1/2] 🌐 RIPE Stat RPKI Validation...")
    try:
        url = f"https://stat.ripe.net/data/rpki-validation/data.json?resource=AS{asn_limpo}&prefix={bloco}"
        print(f"      URL: {url}")
        print(f"      Conectando...", end=" ", flush=True)

        resp = requests.get(url, timeout=15, headers={'User-Agent': 'CONEXA-CRM/1.0'})
        print(f"✅ {resp.status_code}")

        if resp.status_code == 200:
            data = resp.json()

            # Extrair informações
            dados = data.get('data', {})
            status = dados.get('status', 'unknown').lower()
            validating_roas = dados.get('validating_roas', [])

            print(f"      📋 Status: {status}")
            print(f"      📋 ROAs encontradas: {len(validating_roas)}")

            # ==========================
            # Processar Status
            # ==========================

            # ✅ VÁLIDO
            if status == 'valid':
                print(f"      ✅ ROA VÁLIDA!")

                # Detalhar ROAs
                detalhes_roas = []
                for roa in validating_roas:
                    origin = roa.get('origin', '?')
                    prefix = roa.get('prefix', '?')
                    max_length = roa.get('max_length', '?')
                    validity = roa.get('validity', '?')
                    detalhes_roas.append(f"{prefix} AS{origin} (/{max_length})")

                detalhes_texto = "\n   ".join(detalhes_roas) if detalhes_roas else "ROA válido"

                return {
                    'valido': True,
                    'status': 'Valid',
                    'mensagem': 'ROA válido encontrado',
                    'detalhes': f'{bloco} é coberto por ROA válida:\n   {detalhes_texto}'
                }

            # ❌ INVÁLIDO
            elif status == 'invalid':
                print(f"      ❌ ROA INVÁLIDA!")

                # Detalhar conflito
                conflitos = []
                if validating_roas:
                    for roa in validating_roas:
                        origin = roa.get('origin', '?')
                        prefix = roa.get('prefix', '?')
                        conflitos.append(f"AS{origin} para {prefix}")
                else:
                    conflitos.append("Conflito não especificado")

                detalhes_texto = " ou ".join(conflitos)

                return {
                    'valido': False,
                    'status': 'Invalid',
                    'mensagem': 'ROA inválido - conflito detectado',
                    'detalhes': f'{bloco} entra em conflito com ROA(s): {detalhes_texto}'
                }

            # ⏳ DESCONHECIDO
            elif status == 'unknown':
                print(f"      ⏳ ROA NÃO ENCONTRADO")

                return {
                    'valido': False,
                    'status': 'Unknown',
                    'mensagem': 'ROA não encontrado',
                    'detalhes': f'{bloco} / AS{asn_limpo} não possui ROA publicado. Publique em: https://my.lacnic.net'
                }

            else:
                print(f"      ❓ Status desconhecido: {status}")
                return {
                    'valido': False,
                    'status': 'Unknown',
                    'mensagem': f'Status desconhecido: {status}',
                    'detalhes': f'Resposta: {data}'
                }

        else:
            print(f"❌ Status HTTP: {resp.status_code}")
            return {
                'valido': False,
                'status': 'Error',
                'mensagem': f'Erro HTTP {resp.status_code}',
                'detalhes': f'RIPE Stat retornou erro: {resp.text[:200]}'
            }

    except requests.exceptions.Timeout:
        # Timeout do RIPE Stat costuma ser pontual/transitório — retornar erro
        # direto aqui (como antes) descartava o fallback do Cloudflare RPKI
        # logo abaixo, que resolveria a mesma consulta sem esperar a próxima
        # rodada agendada. Cai no fallback como qualquer outra exceção.
        print(f"⏱️ Timeout — tentando fallback")
    except Exception as e:
        print(f"❌ {type(e).__name__}: {e}")

    # ==========================
    # FALLBACK: Cloudflare RPKI (se RIPE falhar)
    # ==========================
    print(f"\n[2/2] 🌐 Cloudflare RPKI (fallback)...")
    try:
        url = f"https://rpki.cloudflare.com/api/v1/validity/{asn_limpo}/{bloco}"
        print(f"      URL: {url}")
        print(f"      Conectando...", end=" ", flush=True)

        resp = requests.get(url, timeout=10, headers={'User-Agent': 'CONEXA-CRM/1.0'})
        print(f"✅ {resp.status_code}")

        if resp.status_code == 200:
            data = resp.json()
            validity = data.get('validity', {})
            state = validity.get('state', '').upper()

            print(f"      📋 State: {state}")

            if state == 'VALID':
                print(f"      ✅ ROA VÁLIDA")
                return {
                    'valido': True,
                    'status': 'Valid',
                    'mensagem': 'ROA válido encontrado',
                    'detalhes': f'{bloco} possui ROA válida para AS{asn_limpo} (Cloudflare RPKI)'
                }
            elif state == 'INVALID':
                print(f"      ❌ ROA INVÁLIDA")
                return {
                    'valido': False,
                    'status': 'Invalid',
                    'mensagem': 'ROA inválido',
                    'detalhes': f'ROA conflita para {bloco} e AS{asn_limpo}'
                }

    except Exception as e:
        print(f"      ⚠️ Indisponível")

    # ==========================
    # Nenhuma API funcionou
    # ==========================
    print(f"\n❌ Validação não disponível")
    return {
        'valido': False,
        'status': 'Unknown',
        'mensagem': 'Validação RPKI não disponível',
        'detalhes': f'Não foi possível validar {bloco} / AS{asn_limpo}. Verifique em: https://routinator.lacnic.net'
    }


def consultar_lacnic_whois(bloco, asn):
    """
    Consulta LACNIC whois APENAS para verificar se bloco está registrado
    NÃO para inferir status de ROA (isso é responsabilidade de APIs RPKI)
    """
    print(f"      Consultando whois.lacnic.net:43...")

    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(15)

        sock.connect(('whois.lacnic.net', 43))
        sock.send(f"{bloco}\n".encode())

        response = b""
        sock.settimeout(5)
        try:
            while True:
                data = sock.recv(4096)
                if not data:
                    break
                response += data
        except socket.timeout:
            pass

        sock.close()

        output = response.decode('utf-8', errors='ignore')
        output_lower = output.lower()

        resultado = {
            'encontrado': False,
            'inetnum': None,
            'asn': None,
            'owner': None,
            'raw': output
        }

        if 'inetnum:' in output_lower or 'inet6num:' in output_lower:
            resultado['encontrado'] = True

            # Extrair informações
            for linha in output.split('\n'):
                linha_lower = linha.lower()

                if linha_lower.startswith('inetnum:') or linha_lower.startswith('inet6num:'):
                    resultado['inetnum'] = linha.split(':', 1)[1].strip()

                if linha_lower.startswith('aut-num:'):
                    resultado['asn'] = linha.split(':', 1)[1].strip()

                if linha_lower.startswith('owner:'):
                    resultado['owner'] = linha.split(':', 1)[1].strip()

        return resultado

    except Exception as e:
        print(f"      ⚠️ Erro: {type(e).__name__}")
        return {'encontrado': False}


def descobrir_status_roa(bloco, asn, dados_lacnic):
    """
    Tenta descobrir se a ROA existe baseado em dados do LACNIC

    Lógica:
    1. Se ASN no LACNIC == ASN da query → provável que tenha ROA ou está em processo
    2. Se ASN diferente → ROA pode ter conflito
    3. Se não tem ASN → ainda pode estar em processo
    """
    print(f"\n   Analisando dados LACNIC para ROA...")

    asn_lacnic = dados_lacnic.get('asn', '').replace('AS', '').replace('as', '').strip()

    # Caso 1: ASN coincide
    if asn_lacnic and asn_lacnic == asn:
        print(f"      ✅ ASN coincide: {asn}")
        return {
            'valido': True,
            'status': 'Valid',
            'mensagem': 'ASN confirmado no LACNIC',
            'detalhes': f'{bloco} registrado para AS{asn} no LACNIC. ROA provavelmente válido.'
        }

    # Caso 2: ASN diferente
    elif asn_lacnic:
        print(f"      ⚠️ ASN diferente: esperado AS{asn}, encontrado AS{asn_lacnic}")
        return {
            'valido': False,
            'status': 'Invalid',
            'mensagem': 'ASN não coincide',
            'detalhes': f'{bloco} está registrado para AS{asn_lacnic} no LACNIC, não AS{asn}. ROA pode ter conflito.'
        }

    # Caso 3: Bloco registrado mas sem ASN
    else:
        print(f"      ℹ️ Bloco registrado mas sem ASN específico")
        return {
            'valido': False,
            'status': 'Unknown',
            'mensagem': 'Status ROA indeterminado',
            'detalhes': f'{bloco} está registrado no LACNIC. Se ROA foi criada recentemente, verifique em 24h ou consulte https://my.lacnic.net'
        }


def validar_irr(bloco, asn, irr_registry):
    """
    ✅ VERSÃO FINAL TESTADA: Valida IRR com queries corretas
    - LACNIC usa inetnum (TESTADO - FUNCIONA)
    - Suporta blocos agregados
    """
    try:
        print(f"\n{'='*60}")
        print(f"📡 VALIDANDO IRR")
        print(f"{'='*60}")
        print(f"Bloco: {bloco}")
        print(f"ASN: {asn}")
        print(f"Registry: {irr_registry}")

        if not irr_registry:
            irr_registry = 'LACNIC'

        asn_limpo = asn.replace('AS', '').replace('as', '').strip() if asn else ''

        # Lista de servidores para tentar
        servidores = []

        # ✅ LACNIC tem formato especial (TESTADO)
        if irr_registry.upper() == 'LACNIC':
            servidores.append(('whois.lacnic.net', 'LACNIC', 'inetnum'))

        # Adicionar servidor primário padrão
        servidor_primario = get_whois_server(irr_registry)
        if servidor_primario != 'whois.lacnic.net':
            servidores.append((servidor_primario, irr_registry, 'route'))

        # Adicionar RADB como fallback
        if servidor_primario != 'whois.radb.net':
            servidores.append(('whois.radb.net', 'RADB', 'route'))

        # Adicionar RIPE como segundo fallback
        if servidor_primario != 'whois.ripe.net':
            servidores.append(('whois.ripe.net', 'RIPE', 'route'))

        # Tentar cada servidor
        for idx, (whois_server, registry_name, tipo_query) in enumerate(servidores, 1):
            print(f"\n[{idx}/{len(servidores)}] Tentando {registry_name} ({whois_server})...")

            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(15)
                sock.connect((whois_server, 43))

                # ✅ Queries diferentes para LACNIC vs outros
                if tipo_query == 'inetnum':
                    # LACNIC: consultar inetnum/inet6num
                    queries = [
                        f"{bloco}\n",                    # Query simples (TESTADA - FUNCIONA)
                        f"-T inetnum {bloco}\n",         # Query específica
                        f"-T inet6num {bloco}\n",        # IPv6
                    ]
                else:
                    # Outros IRRs: consultar route objects
                    queries = [
                        f"-r -T route {bloco}\n",      # Query específica para route
                        f"-r -T route6 {bloco}\n",     # Query para route6 (IPv6)
                        f"-i origin AS{asn_limpo}\n" if asn_limpo else f"{bloco}\n",  # Por origin
                        f"{bloco}\n",                   # Query simples
                    ]

                for query_idx, query in enumerate(queries, 1):
                    print(f"   Query {query_idx}: {query.strip()}")

                    sock.send(query.encode())

                    response = b""
                    sock.settimeout(5)

                    try:
                        while True:
                            data = sock.recv(4096)
                            if not data:
                                break
                            response += data
                    except socket.timeout:
                        pass

                    output = response.decode('utf-8', errors='ignore').lower()

                    print(f"   Recebido: {len(output)} bytes")

                    # ✅ Padrões diferentes para LACNIC vs outros
                    if tipo_query == 'inetnum':
                        # LACNIC: procurar inetnum/inet6num
                        padroes = ['inetnum:', 'inet6num:', 'owner:', 'ownerid:', 'aut-num:']
                        encontrou = any(padrao in output for padrao in padroes)

                        if encontrou:
                            print(f"   ✅ Prefixo encontrado no LACNIC!")

                            # Extrair inetnum e ASN
                            linhas = output.split('\n')
                            inetnum_encontrado = None
                            asn_encontrado = None

                            for linha in linhas:
                                if 'inetnum:' in linha or 'inet6num:' in linha:
                                    inetnum_encontrado = linha.split(':', 1)[1].strip()
                                if 'aut-num:' in linha:
                                    asn_encontrado = linha.split(':', 1)[1].strip().replace('as', '')

                            # Verificar se tem AS
                            if asn_limpo and asn_encontrado == asn_limpo:
                                print(f"   ✅ ASN encontrado e confere!")
                                sock.close()
                                return {
                                    'valido': True,
                                    'status': 'Found',
                                    'mensagem': f'Prefixo registrado em {registry_name}',
                                    'detalhes': f'{bloco} encontrado em {registry_name} (bloco: {inetnum_encontrado}) com AS{asn_limpo}'
                                }
                            elif asn_limpo and asn_encontrado and asn_encontrado != asn_limpo:
                                print(f"   ⚠️ ASN diferente! Esperado: AS{asn_limpo}, Encontrado: AS{asn_encontrado}")
                                sock.close()
                                return {
                                    'valido': False,
                                    'status': 'ASN_Mismatch',
                                    'mensagem': f'Prefixo registrado mas ASN diferente',
                                    'detalhes': f'{bloco} está em {inetnum_encontrado} com AS{asn_encontrado}, não AS{asn_limpo}'
                                }
                            else:
                                sock.close()
                                return {
                                    'valido': True,
                                    'status': 'Found',
                                    'mensagem': f'Prefixo registrado em {registry_name}',
                                    'detalhes': f'{bloco} encontrado em {registry_name} (bloco: {inetnum_encontrado}) - ASN não verificado'
                                }
                    else:
                        # Outros IRRs: procurar route objects
                        padroes_route = ['route:', 'route6:', 'origin:']
                        encontrou_route = any(padrao in output for padrao in padroes_route)
                        encontrou_asn = asn_limpo and (f'as{asn_limpo}' in output or asn_limpo in output)

                        if encontrou_route:
                            print(f"   ✅ Route encontrado!")

                            if encontrou_asn:
                                print(f"   ✅ ASN encontrado!")
                                sock.close()
                                return {
                                    'valido': True,
                                    'status': 'Found',
                                    'mensagem': f'Route encontrado em {registry_name}',
                                    'detalhes': f'Route {bloco} com AS{asn_limpo} em {registry_name}'
                                }
                            elif asn_limpo:
                                print(f"   ⚠️ Route encontrado mas ASN diferente")
                                sock.close()
                                return {
                                    'valido': False,
                                    'status': 'ASN_Mismatch',
                                    'mensagem': f'Route encontrado mas ASN não confere',
                                    'detalhes': f'{bloco} existe em {registry_name} mas com ASN diferente de AS{asn_limpo}'
                                }
                            else:
                                sock.close()
                                return {
                                    'valido': True,
                                    'status': 'Found',
                                    'mensagem': f'Route encontrado em {registry_name}',
                                    'detalhes': f'{bloco} encontrado (ASN não verificado)'
                                }

                    # Resetar socket para próxima query
                    if query_idx < len(queries):
                        sock.close()
                        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        sock.settimeout(15)
                        sock.connect((whois_server, 43))

                sock.close()

            except socket.timeout:
                print(f"   ⏱️ Timeout")
                continue
            except socket.error as e:
                print(f"   ❌ Erro de conexão: {str(e)}")
                continue
            except Exception as e:
                print(f"   ❌ Erro: {str(e)}")
                continue

        # Se não encontrou em nenhum servidor
        print(f"\n❌ Route/Inetnum não encontrado em nenhum servidor")
        return {
            'valido': False,
            'status': 'NotFound',
            'mensagem': f'Route não encontrado',
            'detalhes': f'Nenhum route/inetnum encontrado para {bloco} em {irr_registry}, RADB ou RIPE'
        }

    except Exception as e:
        print(f"\n❌ ERRO CRÍTICO: {str(e)}")
        return {
            'valido': False,
            'status': 'Error',
            'mensagem': f'Erro na validação',
            'detalhes': str(e)
        }


def get_whois_server(irr_registry):
    """Retorna servidor whois apropriado"""
    servidores = {
        'LACNIC': 'whois.lacnic.net',
        'RIPE': 'whois.ripe.net',
        'ARIN': 'whois.arin.net',
        'APNIC': 'whois.apnic.net',
        'AFRINIC': 'whois.afrinic.net',
        'RADB': 'whois.radb.net',
    }

    return servidores.get(irr_registry.upper(), 'whois.radb.net')



@login_required(login_url='login')
@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def listar_comentarios_acesso(request, acesso_id):
    """Lista comentários de um acesso específico"""
    acesso = get_object_or_404(Acesso, id=acesso_id)

    # ✅ CORRIGIDO: Verificação de permissão CORRETA
    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    comentarios = acesso.comentarios.all()

    dados = []
    for com in comentarios:
        dados.append({
            'id': com.id,
            'usuario': com.usuario.get_full_name() or com.usuario.username,
            'comentario': com.comentario,
            'data_criacao': com.data_criacao.strftime('%d/%m/%Y %H:%M:%S'),
            'data_atualizacao': com.data_atualizacao.strftime('%d/%m/%Y %H:%M:%S') if com.data_atualizacao != com.data_criacao else None,
        })

    return JsonResponse({
        'success': True,
        'comentarios': dados,
        'total': len(dados)
    })


@modulo_habilitado_required('acessos')
def adicionar_comentario_acesso(request, acesso_id):
    """Adiciona um novo comentário ao acesso"""
    acesso = get_object_or_404(Acesso, id=acesso_id)

    # ✅ CORRIGIDO: Verificação de permissão CORRETA
    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    comentario_texto = request.POST.get('comentario', '').strip()

    if not comentario_texto:
        return JsonResponse({'error': 'Comentário não pode estar vazio'}, status=400)

    if len(comentario_texto) > 5000:
        return JsonResponse({'error': 'Comentário muito longo (máximo 5000 caracteres)'}, status=400)

    comentario = ComentarioAcesso.objects.create(
        acesso=acesso,
        usuario=request.user,
        comentario=comentario_texto
    )

    return JsonResponse({
        'success': True,
        'message': 'Comentário adicionado com sucesso',
        'comentario': {
            'id': comentario.id,
            'usuario': request.user.get_full_name() or request.user.username,
            'comentario': comentario.comentario,
            'data_criacao': comentario.data_criacao.strftime('%d/%m/%Y %H:%M:%S'),
        }
    })


AUDITORIA_SESSOES_POR_PAGINA = 20


@login_required(login_url='login')
@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def listar_sessoes_auditoria(request, acesso_id):
    """Lista sessões de acesso (SSH/Telnet/WinBox) de um Acesso, mais recentes
    primeiro. Aceita filtro opcional por período (?data_inicio=&data_fim=,
    AAAA-MM-DD) e paginação opcional (?pagina=, 20/página) — sem esses
    parâmetros mantém o comportamento antigo (todas as sessões de uma vez),
    usado pelo modal de auditoria por equipamento."""
    acesso = get_object_or_404(Acesso, id=acesso_id)

    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    sessoes = acesso.sessoes_auditoria.select_related('usuario').annotate(
        total_comandos=Count('comandos')
    ).order_by('-iniciada_em')

    data_inicio_str = (request.GET.get('data_inicio') or '').strip()
    data_fim_str = (request.GET.get('data_fim') or '').strip()
    try:
        if data_inicio_str:
            sessoes = sessoes.filter(iniciada_em__date__gte=datetime.strptime(data_inicio_str, '%Y-%m-%d').date())
        if data_fim_str:
            sessoes = sessoes.filter(iniciada_em__date__lte=datetime.strptime(data_fim_str, '%Y-%m-%d').date())
    except ValueError:
        return JsonResponse({'error': 'Data inválida.'}, status=400)

    resposta = {'success': True}

    pagina_str = request.GET.get('pagina')
    if pagina_str is not None:
        try:
            pagina = max(1, int(pagina_str))
        except (TypeError, ValueError):
            pagina = 1
        paginator = Paginator(sessoes, AUDITORIA_SESSOES_POR_PAGINA)
        page_obj = paginator.get_page(pagina)
        sessoes_pagina = page_obj.object_list
        resposta.update({
            'pagina': page_obj.number,
            'total_paginas': paginator.num_pages,
            'tem_anterior': page_obj.has_previous(),
            'tem_proxima': page_obj.has_next(),
        })
    else:
        sessoes_pagina = sessoes

    dados = []
    for s in sessoes_pagina:
        dados.append({
            'id': s.id,
            'tipo': s.tipo,
            'tipo_display': s.get_tipo_display(),
            'usuario': (s.usuario.get_full_name() or s.usuario.username) if s.usuario else '—',
            'ip_origem': s.ip_origem,
            'status': s.status,
            'iniciada_em': s.iniciada_em.strftime('%d/%m/%Y %H:%M:%S'),
            'encerrada_em': s.encerrada_em.strftime('%d/%m/%Y %H:%M:%S') if s.encerrada_em else None,
            'duracao_segundos': s.duracao_segundos,
            'video_url': (settings.MEDIA_URL + s.arquivo_video) if s.arquivo_video else None,
            'total_comandos': s.total_comandos,
            'tem_transcript': bool(s.transcript),
        })

    resposta.update({'sessoes': dados, 'total': sessoes.count()})
    return JsonResponse(resposta)


@login_required(login_url='login')
@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def listar_comandos_sessao(request, sessao_id):
    """Lista os comandos digitados (stdin) numa AcessoSessao SSH/Telnet."""
    sessao = get_object_or_404(AcessoSessao, id=sessao_id)
    acesso = sessao.acesso

    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    dados = [
        {'comando': c.comando, 'executado_em': c.executado_em.strftime('%d/%m/%Y %H:%M:%S')}
        for c in sessao.comandos.all()
    ]

    return JsonResponse({'success': True, 'comandos': dados, 'total': len(dados)})


@login_required(login_url='login')
@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def ver_transcript_sessao(request, sessao_id):
    """Retorna o transcript completo (stdout, ANSI removido) de uma
    AcessoSessao SSH/Telnet — mostra o comando expandido/completo tal como
    o equipamento ecoou, não só o que foi digitado."""
    sessao = get_object_or_404(AcessoSessao, id=sessao_id)
    acesso = sessao.acesso

    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    return JsonResponse({'success': True, 'transcript': sessao.transcript})


@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def auditoria_cliente_view(request, cliente_id):
    """Tela de auditoria do cliente: visão consolidada de todos os hosts
    (Acesso) que já tiveram alguma sessão de acesso gravada, com filtro
    por período/host/usuário. Complementa o modal de auditoria por
    equipamento (aba Acessos), que fica restrito a um host por vez."""
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if not _perms.pode_acessar_cliente(request.user, cliente):
        messages.error(request, 'Você não possui permissão para visualizar este cliente.')
        return redirect('quadro_geral')

    usuarios_com_acesso = User.objects.filter(
        sessoes_acesso__acesso__cliente=cliente
    ).distinct().order_by('first_name', 'username')

    return render(request, 'auditoria_cliente.html', {
        'cliente': cliente,
        'usuarios_com_acesso': usuarios_com_acesso,
    })


@login_required(login_url='login')
@require_http_methods(["GET"])
@modulo_habilitado_required('acessos')
def auditoria_cliente_hosts(request, cliente_id):
    """Resumo por host das sessões de auditoria de um cliente: um item por
    Acesso que tem ao menos 1 sessão no período filtrado, com totais e
    dados da última sessão. Aceita ?data_inicio=&data_fim= (AAAA-MM-DD),
    ?busca= (host/função) e ?usuario_id=."""
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    data_inicio_str = (request.GET.get('data_inicio') or '').strip()
    data_fim_str = (request.GET.get('data_fim') or '').strip()
    busca = (request.GET.get('busca') or '').strip()
    usuario_id = (request.GET.get('usuario_id') or '').strip()

    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else None
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else None
    except ValueError:
        return JsonResponse({'error': 'Data inválida.'}, status=400)

    # Filtro aplicado tanto na contagem (via relação reversa sessoes_auditoria)
    # quanto na subquery da última sessão (direto no model AcessoSessao) —
    # precisa das duas formas porque uma atravessa o related_name e a outra não.
    filtro_periodo = Q()
    filtro_periodo_sessao = Q()
    if data_inicio:
        filtro_periodo &= Q(sessoes_auditoria__iniciada_em__date__gte=data_inicio)
        filtro_periodo_sessao &= Q(iniciada_em__date__gte=data_inicio)
    if data_fim:
        filtro_periodo &= Q(sessoes_auditoria__iniciada_em__date__lte=data_fim)
        filtro_periodo_sessao &= Q(iniciada_em__date__lte=data_fim)
    if usuario_id:
        filtro_periodo &= Q(sessoes_auditoria__usuario_id=usuario_id)
        filtro_periodo_sessao &= Q(usuario_id=usuario_id)

    ultima_sessao_qs = AcessoSessao.objects.filter(
        filtro_periodo_sessao, acesso=OuterRef('pk')
    ).order_by('-iniciada_em')

    acessos = Acesso.objects.filter(cliente=cliente).annotate(
        total_sessoes=Count('sessoes_auditoria', filter=filtro_periodo, distinct=True),
        ultima_sessao_em=Subquery(ultima_sessao_qs.values('iniciada_em')[:1]),
        ultima_sessao_usuario=Subquery(ultima_sessao_qs.values('usuario__username')[:1]),
        ultima_sessao_usuario_nome=Subquery(ultima_sessao_qs.values('usuario__first_name')[:1]),
        ultima_sessao_status=Subquery(ultima_sessao_qs.values('status')[:1]),
    ).filter(total_sessoes__gt=0).select_related('funcao')

    if busca:
        acessos = acessos.filter(Q(host__icontains=busca) | Q(tipo__icontains=busca) | Q(funcao__descricao__icontains=busca))

    acessos = acessos.order_by('-ultima_sessao_em')

    hosts = []
    for a in acessos:
        hosts.append({
            'acesso_id': a.id,
            'host': a.host,
            'tipo': a.tipo,
            'funcao': a.funcao.descricao if a.funcao else '—',
            'total_sessoes': a.total_sessoes,
            'ultima_sessao_em': a.ultima_sessao_em.strftime('%d/%m/%Y %H:%M:%S') if a.ultima_sessao_em else None,
            'ultima_sessao_usuario': (a.ultima_sessao_usuario_nome or a.ultima_sessao_usuario) or '—',
            'sessao_ativa': a.ultima_sessao_status == 'ativa',
        })

    return JsonResponse({
        'success': True,
        'hosts': hosts,
        'total_hosts': len(hosts),
        'total_sessoes': sum(h['total_sessoes'] for h in hosts),
    })


@login_required(login_url="login")
@require_http_methods(["POST", "DELETE"])
@modulo_habilitado_required('acessos')
def deletar_comentario_acesso(request, comentario_id):
    """Deleta um comentário do acesso"""
    comentario = get_object_or_404(ComentarioAcesso, id=comentario_id)

    # Verificar permissão - apenas o autor ou admin pode deletar
    if request.user != comentario.usuario and not _perms.is_backoffice(request.user):
        return JsonResponse({'error': 'Sem permissão para deletar'}, status=403)

    acesso = comentario.acesso
    comentario.delete()

    return JsonResponse({
        'success': True,
        'message': 'Comentário deletado com sucesso'
    })


@login_required(login_url="login")
@require_http_methods(["POST"])
@modulo_habilitado_required('acessos')
def editar_comentario_acesso(request, comentario_id):
    """Edita um comentário existente"""
    comentario = get_object_or_404(ComentarioAcesso, id=comentario_id)

    # Verificar permissão - apenas o autor ou admin pode editar
    if request.user != comentario.usuario and not _perms.is_backoffice(request.user):
        return JsonResponse({'error': 'Sem permissão para editar'}, status=403)

    novo_texto = request.POST.get('comentario', '').strip()

    if not novo_texto:
        return JsonResponse({'error': 'Comentário não pode estar vazio'}, status=400)

    if len(novo_texto) > 5000:
        return JsonResponse({'error': 'Comentário muito longo (máximo 5000 caracteres)'}, status=400)

    comentario.comentario = novo_texto
    comentario.save()

    return JsonResponse({
        'success': True,
        'message': 'Comentário atualizado com sucesso',
        'comentario': {
            'id': comentario.id,
            'comentario': comentario.comentario,
            'data_atualizacao': comentario.data_atualizacao.strftime('%d/%m/%Y %H:%M:%S'),
        }
    })


# ─────────────────────────────────────────────────────────────────────────────
# TESTES DE REDE — ping + MTR via SSH no proxy do cliente
# ─────────────────────────────────────────────────────────────────────────────

DESTINOS_PADRAO = [
    '8.8.8.8',
    '1.1.1.1',
    'whatsapp.com',
    'instagram.com',
    'uol.com.br',
    'nytimes.com',
]

DNS_PUBLICOS = [
    {'nome': 'Google',      'ip': '8.8.8.8'},
    {'nome': 'Cloudflare',  'ip': '1.1.1.1'},
    {'nome': 'OpenDNS',     'ip': '208.67.222.222'},
    {'nome': 'Quad9',       'ip': '9.9.9.9'},
    {'nome': 'AdGuard',     'ip': '94.140.14.14'},
]


def _ssh_exec(proxy, cmd, timeout=45):
    """Executa um comando via SSH no proxy. Retorna (stdout, stderr, return_code)."""
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname=proxy.host,
        port=int(proxy.porta),
        username=proxy.usuario,
        password=proxy.senha,
        timeout=15,
        look_for_keys=False,
        allow_agent=False,
        banner_timeout=15,
    )
    try:
        stdin, stdout, stderr = client.exec_command(cmd, timeout=timeout)
        out = stdout.read().decode('utf-8', errors='replace')
        err = stderr.read().decode('utf-8', errors='replace')
        rc  = stdout.channel.recv_exit_status()
        return out, err, rc
    finally:
        client.close()


def _exec_local(cmd, timeout=45):
    """Executa comando localmente via subprocess. Retorna (stdout, stderr, return_code)."""
    import subprocess
    try:
        proc = subprocess.run(
            cmd, shell=True, capture_output=True, text=True, timeout=timeout
        )
        return proc.stdout, proc.stderr, proc.returncode
    except subprocess.TimeoutExpired:
        return '', 'Timeout', -1
    except Exception as e:
        return '', str(e), -1


def _parse_ping_output(output, host):
    """Extrai estatísticas do output de ping (Linux)."""
    result = {'host': host, 'alcancavel': False, 'enviados': 0,
              'recebidos': 0, 'perdidos': 0, 'perda_pct': 100,
              'rtt_min': None, 'rtt_avg': None, 'rtt_max': None,
              'output': output.strip()[:2000]}
    m = re.search(r'(\d+)\s+packets? transmitted,\s*(\d+)\s+received', output)
    if m:
        result['enviados']  = int(m.group(1))
        result['recebidos'] = int(m.group(2))
        result['perdidos']  = result['enviados'] - result['recebidos']
        result['perda_pct'] = round(result['perdidos'] / max(result['enviados'], 1) * 100)
        result['alcancavel'] = result['recebidos'] > 0
    m2 = re.search(r'min/avg/max[^=]*=\s*([\d.]+)/([\d.]+)/([\d.]+)', output)
    if m2:
        result['rtt_min'] = m2.group(1)
        result['rtt_avg'] = m2.group(2)
        result['rtt_max'] = m2.group(3)
    return result


@login_required(login_url='login')
@require_http_methods(['POST'])
@modulo_habilitado_required('testes_rede')
def teste_rede_cliente(request, cliente_id):
    """
    Executa ping + mtr/traceroute para os destinos através do proxy SSH.
    Aceita JSON body:
      ipv            : 'v4' | 'v6' | 'ambos'  (padrão: 'v4')
      packets        : int 1-100               (padrão: 5)
      wait           : int 1-10 segundos       (padrão: 3)
      destinos_extras: list[str]               (destinos adicionais)
      destino_unico  : str | null              (se informado, testa APENAS este destino)
    """
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if not _perms.is_backoffice(request.user) or not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    proxy = ProxyServer.objects.filter(cliente=cliente, ativo=True).first()
    usar_vpn_local = False
    if not proxy:
        from .models import VPNWireGuard
        from . import vpn_manager as _wgm
        vpns_ativas = VPNWireGuard.objects.filter(cliente=cliente, ativo=True, peer_no_servidor=True)
        if vpns_ativas.exists():
            peers = _wgm.get_peers_status()
            usar_vpn_local = any(
                peers.get(v.cliente_public_key, {}).get('conectado') for v in vpns_ativas
            )
        if not usar_vpn_local:
            return JsonResponse({'error': 'Nenhum proxy SSH ativo configurado para este cliente.'}, status=400)

    # Executor: SSH no proxy ou local via VPN
    def _exec(cmd, timeout=45):
        if usar_vpn_local:
            return _exec_local(cmd, timeout)
        return _ssh_exec(proxy, cmd, timeout)

    # ── Parâmetros ────────────────────────────────────────────────────────
    try:
        body = json.loads(request.body) if request.body else {}
    except Exception:
        body = {}

    ipv             = body.get('ipv', 'v4')
    packets         = max(1, min(100, int(body.get('packets', 5))))
    wait            = max(1, min(10,  int(body.get('wait', 3))))
    destino_unico   = (body.get('destino_unico') or '').strip()
    destinos_extras = [d.strip() for d in body.get('destinos_extras', []) if d.strip()]

    # Montar lista final de destinos
    if destino_unico:
        # Modo destino único — ignora padrão e extras
        todos_destinos = [destino_unico]
    else:
        todos_destinos = list(DESTINOS_PADRAO)
        for d in destinos_extras:
            if d not in todos_destinos:
                todos_destinos.append(d)

    # ── Opções de versão IP ───────────────────────────────────────────────
    # ping: -4/-6  |  mtr: -4/-6  |  traceroute: -4/-6
    mtr_flag_v4 = '-4'
    mtr_flag_v6 = '-6'

    # SSH timeout total = (packets * wait) + margem
    ssh_ping_timeout = packets * wait + 15

    def ping_cmd(destino, flag):
        """Monta comando ping com flag de versão IP."""
        return f'ping {flag} -c {packets} -W {wait} {destino} 2>&1'

    def mtr_cmd(destino, ip_flag):
        """mtr -4/-6 com fallback para traceroute -4/-6."""
        return (
            f'which mtr > /dev/null 2>&1 && '
            f'mtr --report --report-cycles {packets} --no-dns {ip_flag} {destino} 2>&1 || '
            f'traceroute -n -w 2 -m 20 {ip_flag} {destino} 2>&1'
        )

    def testar_destino(destino):
        resultado = {
            'destino':  destino,
            'extra':    destino not in DESTINOS_PADRAO,
            'ping_v4':  None,
            'ping_v6':  None,
            'mtr_v4':   None,
            'mtr_v6':   None,
            'erro':     None,
        }
        try:
            if ipv in ('v4', 'ambos'):
                out, _, _ = _exec(ping_cmd(destino, '-4'), timeout=ssh_ping_timeout)
                resultado['ping_v4'] = _parse_ping_output(out, destino)

            if ipv in ('v6', 'ambos'):
                out6, _, _ = _exec(ping_cmd(destino, '-6'), timeout=ssh_ping_timeout)
                resultado['ping_v6'] = _parse_ping_output(out6, destino)

            # MTR — só o protocolo primário para não duplicar
            mtr_flag = mtr_flag_v4 if ipv in ('v4', 'ambos') else mtr_flag_v6
            out_mtr, err_mtr, _ = _exec(mtr_cmd(destino, mtr_flag), timeout=90)
            mtr_raw = (out_mtr + err_mtr).strip()[:4000]
            ferr    = 'mtr' if ('Loss%' in mtr_raw or 'HOST' in mtr_raw) else 'traceroute'

            if ipv in ('v4', 'ambos'):
                resultado['mtr_v4'] = {'output': mtr_raw, 'ferramenta': ferr}
            if ipv == 'v6':
                out_mtr6, err_mtr6, _ = _exec(mtr_cmd(destino, mtr_flag_v6), timeout=90)
                mtr_raw6 = (out_mtr6 + err_mtr6).strip()[:4000]
                resultado['mtr_v6'] = {'output': mtr_raw6, 'ferramenta': ferr}
            elif ipv == 'ambos':
                out_mtr6, err_mtr6, _ = _exec(mtr_cmd(destino, mtr_flag_v6), timeout=90)
                mtr_raw6 = (out_mtr6 + err_mtr6).strip()[:4000]
                resultado['mtr_v6'] = {'output': mtr_raw6, 'ferramenta': ferr}

        except Exception as e:
            resultado['erro'] = str(e)
        return resultado

    # Executar em paralelo
    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
        futuros = {pool.submit(testar_destino, d): d for d in todos_destinos}
        resultados = []
        for futuro in concurrent.futures.as_completed(futuros):
            resultados.append(futuro.result())

    # Ordenar: padrão primeiro, extras depois (na ordem em que foram inseridos)
    ordem = {d: i for i, d in enumerate(todos_destinos)}
    resultados.sort(key=lambda r: ordem.get(r['destino'], 99))

    proxy_info = {'nome': proxy.nome, 'host': proxy.host} if proxy else {'nome': 'VPN WireGuard (local)', 'host': 'servidor CRM'}
    return JsonResponse({
        'ok':        True,
        'proxy':     proxy_info,
        'via_vpn':   usar_vpn_local,
        'ipv':       ipv,
        'packets':   packets,
        'wait':      wait,
        'resultados': resultados,
    })


# ─────────────────────────────────────────────────────────────────────────────
# TESTE DE DNS — nslookup comparativo com DNS do cliente e DNS públicos
# ─────────────────────────────────────────────────────────────────────────────

# Domínios que naturalmente retornam IPs diferentes por geo-DNS / CDN
KNOWN_CDN_DOMAINS = [
    'whatsapp.com', 'instagram.com', 'facebook.com', 'fb.com', 'fbcdn.net',
    'twitter.com', 'x.com', 't.co',
    'google.com', 'google.com.br', 'youtube.com', 'googleapis.com', 'gstatic.com',
    'netflix.com', 'nflxso.net', 'nflxvideo.net',
    'tiktok.com', 'byteoversea.com', 'musical.ly',
    'amazon.com', 'amazonaws.com', 'cloudfront.net', 'awsstatic.com',
    'akamai.net', 'akamaiedge.net', 'akamaitechnologies.com',
    'fastly.net', 'fastlylb.net',
    'cloudflare.com', 'cloudflare.net',
    'microsoft.com', 'office.com', 'live.com', 'msftncsi.com',
    'apple.com', 'icloud.com',
    'telegram.org', 'snapchat.com',
]


def _is_cdn_domain(dominio):
    """Verifica se o domínio é conhecido por usar geo-DNS / CDN."""
    d = dominio.lower()
    return any(cdn in d for cdn in KNOWN_CDN_DOMAINS)


def _same_slash16(ips):
    """/16 prefix heuristic: se todos os IPs compartilham /16, provavelmente é CDN."""
    if len(ips) <= 1:
        return True
    prefixes = {'.'.join(ip.split('.')[:2]) for ip in ips if ip.count('.') >= 1}
    return len(prefixes) == 1


def _get_asn_for_ip(ip):
    """
    Resolve o ASN de um IP via Team Cymru DNS (gratuito, sem API key).
    Exemplo: dig +short 8.8.8.8.origin.asn.cymru.com TXT
      → '15169 | 8.8.8.0/24 | US | arin | 1992-12-01'
    Retorna dict com 'asn', 'prefix', 'country' ou None em caso de falha.
    """
    try:
        import subprocess as _sp
        rev = '.'.join(reversed(ip.split('.')))
        r = _sp.run(
            ['dig', '+short', '+time=4', '+tries=1', f'{rev}.origin.asn.cymru.com', 'TXT'],
            capture_output=True, text=True, timeout=6,
        )
        txt = r.stdout.strip().strip('"').strip("'")
        if not txt or 'NXDOMAIN' in txt:
            return None
        parts = [p.strip() for p in txt.split('|')]
        return {
            'asn':     parts[0] if parts else None,
            'prefix':  parts[1] if len(parts) > 1 else None,
            'country': parts[2] if len(parts) > 2 else None,
        }
    except Exception:
        return None

def _nslookup_ssh(proxy, dominio, dns_ip, timeout=20):
    """
    Resolve dominio@dns_ip via SSH no proxy.
    Tenta dig primeiro; se não disponível, usa nslookup.
    """
    # Comando único: tenta dig, senão usa nslookup como fallback
    cmd = (
        f'if command -v dig > /dev/null 2>&1; then '
        f'  dig +short +time=5 +tries=2 A {dominio} @{dns_ip} 2>&1; '
        f'  echo "---IPV6---"; '
        f'  dig +short +time=5 +tries=2 AAAA {dominio} @{dns_ip} 2>&1; '
        f'else '
        f'  echo "---NSLOOKUP---"; '
        f'  nslookup {dominio} {dns_ip} 2>&1; '
        f'fi'
    )
    try:
        if proxy is not None:
            out, err_out, _ = _ssh_exec(proxy, cmd, timeout=timeout)
        else:
            out, err_out, _ = _exec_local(cmd, timeout=timeout)
        raw = (out + err_out).strip()

        ipv4_list = []
        ipv6_list = []
        ferramenta = 'dig'

        if '---NSLOOKUP---' in raw:
            # Parsing de nslookup
            ferramenta = 'nslookup'
            body = raw.split('---NSLOOKUP---', 1)[-1]
            for line in body.splitlines():
                # nslookup retorna: "Address: 1.2.3.4" ou "Address: ::1"
                m = re.match(r'\s*Address(?:\s*\d+)?:\s*(\S+)', line)
                if m:
                    addr = m.group(1)
                    if ':' in addr:
                        ipv6_list.append(addr)
                    elif re.match(r'^\d+\.\d+\.\d+\.\d+$', addr):
                        # Ignorar o próprio IP do servidor DNS (primeira linha do nslookup)
                        if addr != dns_ip:
                            ipv4_list.append(addr)
        elif '---IPV6---' in raw:
            # Parsing de dig
            ferramenta = 'dig'
            partes = raw.split('---IPV6---')
            a_part    = partes[0]
            aaaa_part = partes[1] if len(partes) > 1 else ''

            # dig +short pode retornar CNAMEs antes do IP — filtrar apenas IPs
            for line in a_part.splitlines():
                line = line.strip()
                if re.match(r'^\d+\.\d+\.\d+\.\d+$', line):
                    ipv4_list.append(line)
            for line in aaaa_part.splitlines():
                line = line.strip()
                if ':' in line and not line.startswith(';'):
                    ipv6_list.append(line)
        else:
            # Resposta inesperada — tentar extrair IPs diretamente
            for line in raw.splitlines():
                line = line.strip()
                if re.match(r'^\d+\.\d+\.\d+\.\d+$', line) and line != dns_ip:
                    ipv4_list.append(line)

        # Remover duplicatas preservando ordem
        seen = set()
        ipv4_unique = [ip for ip in ipv4_list if ip not in seen and not seen.add(ip)]
        seen6 = set()
        ipv6_unique = [ip for ip in ipv6_list if ip not in seen6 and not seen6.add(ip)]

        return {
            'dns':        dns_ip,
            'ipv4':       ipv4_unique,
            'ipv6':       ipv6_unique,
            'ferramenta': ferramenta,
            'raw':        raw[:500],  # debug
            'erro':       None,
        }
    except Exception as e:
        return {'dns': dns_ip, 'ipv4': [], 'ipv6': [], 'ferramenta': '?', 'raw': '', 'erro': str(e)}


@login_required(login_url='login')
@require_http_methods(['POST'])
@modulo_habilitado_required('testes_rede')
def teste_dns_cliente(request, cliente_id):
    """
    Realiza nslookup de um domínio usando:
    1. DNS recursivo do cliente (configurado no campo dns_cliente)
    2. DNS públicos (Google, Cloudflare, OpenDNS, Quad9, AdGuard)
    Retorna comparativo das respostas.
    """
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if not _perms.is_backoffice(request.user) or not _perms.pode_acessar_cliente(request.user, cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Body inválido'}, status=400)

    dominio      = body.get('dominio', '').strip().lower()
    dns_cliente  = body.get('dns_cliente', '').strip()

    if not dominio:
        return JsonResponse({'error': 'Domínio não informado.'}, status=400)

    # Remover protocolo se vier
    dominio = re.sub(r'^https?://', '', dominio).split('/')[0]

    proxy = ProxyServer.objects.filter(cliente=cliente, ativo=True).first()
    usar_vpn_local_dns = False
    if not proxy:
        from .models import VPNWireGuard
        from . import vpn_manager as _wgm
        vpns_ativas = VPNWireGuard.objects.filter(cliente=cliente, ativo=True, peer_no_servidor=True)
        if vpns_ativas.exists():
            peers = _wgm.get_peers_status()
            usar_vpn_local_dns = any(
                peers.get(v.cliente_public_key, {}).get('conectado') for v in vpns_ativas
            )
        if not usar_vpn_local_dns:
            return JsonResponse({'error': 'Nenhum proxy SSH ativo configurado para este cliente.'}, status=400)

    todos_dns = []
    if dns_cliente:
        todos_dns.append({'nome': 'DNS do Cliente', 'ip': dns_cliente, 'is_cliente': True})
    for pub in DNS_PUBLICOS:
        todos_dns.append({**pub, 'is_cliente': False})

    resultados = []

    with __import__('concurrent.futures', fromlist=['ThreadPoolExecutor']).ThreadPoolExecutor(max_workers=8) as pool:
        futuros = {
            pool.submit(_nslookup_ssh, proxy, dominio, d['ip']): d
            for d in todos_dns
        }
        for futuro in __import__('concurrent.futures', fromlist=['as_completed']).as_completed(futuros):
            info = futuros[futuro]
            res  = futuro.result()
            res['nome']       = info['nome']
            res['is_cliente'] = info['is_cliente']
            resultados.append(res)

    # Ordenar: cliente primeiro, depois públicos na ordem original
    resultados.sort(key=lambda r: (0 if r['is_cliente'] else 1, todos_dns.index(next(d for d in todos_dns if d['ip'] == r['dns']))))

    # ── Coletar todos os IPs únicos ────────────────────────────────────────
    todos_ipv4 = set()
    for r in resultados:
        todos_ipv4.update(r['ipv4'])

    responderam      = [r for r in resultados if r['ipv4']]
    qtd_responderam  = len(responderam)
    is_cdn_domain    = _is_cdn_domain(dominio)

    # ── Lookup ASN para cada IP único (em paralelo, servidor Django) ────────
    asn_map = {}  # {ip: {'asn': '15169', 'prefix': '8.8.8.0/24', ...}}
    if todos_ipv4:
        import concurrent.futures as _cf
        with _cf.ThreadPoolExecutor(max_workers=min(8, len(todos_ipv4))) as pool:
            futures = {pool.submit(_get_asn_for_ip, ip): ip for ip in todos_ipv4}
            for fut in _cf.as_completed(futures):
                ip = futures[fut]
                asn_map[ip] = fut.result()

    # Enriquecer cada resultado com ASN por IP
    for r in resultados:
        r['asn_info'] = {
            ip: asn_map.get(ip)
            for ip in r['ipv4']
        }

    # ── Análise de consistência inteligente ────────────────────────────────
    unique_asns   = {v['asn'] for v in asn_map.values() if v and v.get('asn')}
    same_asn      = len(unique_asns) <= 1
    same_prefix16 = _same_slash16(list(todos_ipv4))

    if qtd_responderam == 0:
        status_geral     = 'sem_resposta'
        consistente_geral = None
        for r in resultados:
            r['consistente'] = None

    elif len(todos_ipv4) <= 1:
        # Todos os servidores retornaram o mesmo IP único
        status_geral      = 'ok'
        consistente_geral = True
        for r in resultados:
            r['consistente'] = True if r['ipv4'] else None

    elif same_asn:
        # IPs diferentes mas mesmo ASN → geo-DNS / anycast / CDN
        status_geral      = 'cdn'
        consistente_geral = True
        for r in resultados:
            r['consistente'] = True if r['ipv4'] else None

    elif same_prefix16 or is_cdn_domain:
        # Mesmo /16 ou domínio conhecido de CDN → provavelmente normal
        status_geral      = 'cdn'
        consistente_geral = True
        for r in resultados:
            r['consistente'] = True if r['ipv4'] else None

    elif qtd_responderam == 1:
        status_geral      = 'parcial'
        consistente_geral = None
        for r in resultados:
            r['consistente'] = True if r['ipv4'] else None

    else:
        # IPs e ASNs diferentes → inconsistência real
        ref_ipv4          = responderam[0]['ipv4']
        conjuntos         = set()
        for r in resultados:
            if r['ipv4']:
                r['consistente'] = set(r['ipv4']) == set(ref_ipv4)
                conjuntos.add(frozenset(r['ipv4']))
            else:
                r['consistente'] = None
        consistente_geral = len(conjuntos) == 1
        status_geral      = 'ok' if consistente_geral else 'inconsistente'

    proxy_info_dns = {'nome': proxy.nome, 'host': proxy.host} if proxy else {'nome': 'VPN WireGuard (local)', 'host': 'servidor CRM'}
    return JsonResponse({
        'ok':               True,
        'dominio':          dominio,
        'proxy':            proxy_info_dns,
        'via_vpn':          usar_vpn_local_dns,
        'resultados':       resultados,
        'todos_ips_distintos': sorted(todos_ipv4),
        'consistente_geral':   consistente_geral,
        'status_geral':        status_geral,   # 'ok'|'cdn'|'inconsistente'|'parcial'|'sem_resposta'
        'qtd_responderam':     qtd_responderam,
        'is_cdn_domain':       is_cdn_domain,
        'same_asn':            same_asn,
        'same_prefix16':       same_prefix16,
        'unique_asns':         sorted(unique_asns),
        'asn_map':             {ip: asn_map[ip] for ip in asn_map if asn_map[ip]},
    })



@csrf_exempt
@login_required(login_url='login')
@modulo_habilitado_required('acessos')
def proxy_web_acesso(request, acesso_id, porta=None, scheme=None, path=''):
    from urllib.parse import urlparse as _up, urljoin as _uj
    acesso = get_object_or_404(Acesso, id=acesso_id)

    # ── Permissão ─────────────────────────────────────────────────────
    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return HttpResponse('Sem permissao', status=403)

    # ── Normalização de URL ───────────────────────────────────────────
    if porta is None or scheme is None:
        porta_web = int(request.GET.get('porta', 80))
        path_qs   = request.GET.get('path', '/')
        if path_qs.startswith('/'): path_qs = path_qs[1:]

        # Só infere o scheme pela porta quando o chamador não informou um
        # explicitamente — caso contrário um acesso HTTPS em porta "não
        # padrão" (ex: 8006) seria forçado de volta para HTTP aqui.
        if 'scheme' in request.GET:
            scheme = request.GET['scheme']
        elif porta_web in (443, 8443, 4443):
            scheme = 'https'
        else:
            scheme = 'http'

        redirect_url = f'/clientes/acessos/{acesso_id}/web/{porta_web}/{scheme}/{path_qs}'
        return HttpResponseRedirect(redirect_url)

    porta_web = int(porta)
    if not path: path = '/'
    if not path.startswith('/'): path = '/' + path

    qs = request.GET.urlencode()
    full_path = path + ('?' + qs if qs else '')

    target_host = acesso.host.strip()
    try:
        if '://' in target_host:
            target_host = _up(target_host).hostname
    except Exception:
        pass

    proxy_base = f'/clientes/acessos/{acesso_id}/web/{porta_web}/{scheme}'
    target_url = f"{scheme}://{target_host}:{porta_web}{full_path}"

    # ── Proxy SSH se IP privado ───────────────────────────────────────
    proxy_srv = None
    if ProxyEngine.is_private_ip(target_host):
        proxy_srv = ProxyServer.objects.filter(cliente=acesso.cliente, ativo=True).first()
        if not proxy_srv:
            if vpn_cobre_ip(acesso.cliente, target_host):
                pass  # VPN ativa cobre este IP — acesso direto sem proxy
            else:
                return HttpResponse(
                    ProxyEngine.get_error_page(
                        f'IP privado (<code>{target_host}</code>) sem proxy SSH ativo.<br>'
                        f'Configure um túnel SSH para este cliente.'
                    ),
                    content_type='text/html', status=400
                )

    # ── Executar Requisição via ProxyEngine ───────────────────────────
    engine = ProxyEngine(proxy_srv,
                         device_username=acesso.usuario,
                         device_password=acesso.senha)

    _django_cookies = {'sessionid', 'csrftoken', 'messages'}
    raw_cookie = request.META.get('HTTP_COOKIE', '')
    # Isolate cookies per-acesso: browser stores device cookies as "a{id}_NAME=value".
    # Only forward cookies belonging to THIS acesso (stripped of prefix).
    # This prevents AIROS_SESSIONID (or any session cookie) from device A
    # bleeding into requests to device B.
    cookie_prefix = f'a{acesso_id}_'
    filtered_cookie_parts = []
    for c in raw_cookie.split(';'):
        c = c.strip()
        if not c or '=' not in c:
            continue
        name, val = c.split('=', 1)
        name = name.strip()
        if name in _django_cookies:
            continue
        if name.startswith(cookie_prefix):
            filtered_cookie_parts.append(f'{name[len(cookie_prefix):]}={val.strip()}')
    filtered_cookies = '; '.join(filtered_cookie_parts)

    req_headers = {}
    if filtered_cookies:
        req_headers['Cookie'] = filtered_cookies
    if request.META.get('CONTENT_TYPE'):
        req_headers['Content-Type'] = request.META['CONTENT_TYPE']
    if request.META.get('HTTP_AUTHORIZATION'):
        req_headers['Authorization'] = request.META['HTTP_AUTHORIZATION']
    if request.META.get('HTTP_ACCEPT'):
        req_headers['Accept'] = request.META['HTTP_ACCEPT']
    # Repassar headers customizados que SPAs enviam (X-Auth-Token, X-CSRF-Token, etc.)
    for meta_key, meta_val in request.META.items():
        if meta_key.startswith('HTTP_X_'):
            header_name = meta_key[5:].replace('_', '-').title()
            req_headers[header_name] = meta_val
    # Proxmox VE exige CSRFPreventionToken em todas as requisições POST/PUT/DELETE
    # Esse header NÃO começa com X-, então não é capturado pelo loop acima
    if 'HTTP_CSRFPREVENTIONTOKEN' in request.META:
        req_headers['CSRFPreventionToken'] = request.META['HTTP_CSRFPREVENTIONTOKEN']

    body = request.body if request.method in ('POST', 'PUT', 'PATCH') else None

    try:
        resp = engine.do_request(
            method=request.method,
            url=target_url,
            headers=req_headers,
            body=body,
        )

        if resp is None:
            proxy_info = (f' via proxy <code>{proxy_srv.nome}</code> ({proxy_srv.host})'
                          if proxy_srv else '')
            return HttpResponse(
                ProxyEngine.get_error_page(
                    f'Sem resposta de <code>{scheme}://{target_host}:{porta_web}</code>{proxy_info}.<br><br>'
                    f'Verifique:<br>'
                    f'&bull; Se o equipamento est&aacute; acess&iacute;vel pela rede do proxy<br>'
                    f'&bull; Se a porta {porta_web} e o protocolo {scheme.upper()} est&atilde;o corretos<br>'
                    f'&bull; Os logs do servidor para mais detalhes'
                ),
                content_type='text/html', status=502
            )

        # ── Tratar Redirects cross-port (ex: http→https) ─────────────
        if resp.status_code in (301, 302, 303, 307, 308):
            location = resp.headers.get('Location', '/')
            # Location pode vir relativo (ex: "zabbix.php?action=..." sem "/" na
            # frente) — resolver contra a URL real requisitada (target_url) garante
            # um path absoluto normalizado. Sem isso, um Location relativo virava
            # concatenação direta com o proxy_base (ex: ".../web/80/http" +
            # "zabbix.php" = ".../web/80/httpzabbix.php", 404).
            p = _up(_uj(target_url, location))

            redir_host   = p.hostname or target_host
            redir_scheme = p.scheme or scheme
            if p.port:
                redir_port = p.port
            elif redir_scheme == 'https':
                redir_port = 443
            else:
                redir_port = 80

            if not p.netloc or redir_host == target_host:
                redir_path = p.path or '/'
                redir_qs   = ('?' + p.query) if p.query else ''
                new_loc = (f'/clientes/acessos/{acesso_id}/web/{redir_port}/{redir_scheme}'
                           f'{redir_path}{redir_qs}')
                response = HttpResponse(status=resp.status_code)
                response['Location'] = new_loc
            else:
                response = HttpResponse(status=resp.status_code)
                response['Location'] = location

            for cookie_str in getattr(resp, 'cookies_raw', []):
                parts = cookie_str.split(';')
                if parts:
                    nv = parts[0].split('=', 1)
                    if len(nv) == 2:
                        response.set_cookie(
                            cookie_prefix + nv[0].strip(), nv[1].strip(),
                            path='/', samesite='Lax'
                        )
            return response

        # ── Processar Conteúdo ────────────────────────────────────────
        content_type = resp.headers.get('Content-Type', '')
        content      = resp.content

        # Se o dispositivo não enviou Content-Type, inferir pelo conteúdo:
        # bytes que começam com '<' provavelmente são HTML; caso contrário, não reescrever.
        if not content_type:
            stripped = content.lstrip()
            content_type = 'text/html' if stripped.startswith(b'<') else 'application/octet-stream'

        if 'text/html' in content_type or 'text/css' in content_type:
            content = engine.rewrite_content(content, content_type, proxy_base, target_host,
                                             cookie_prefix=cookie_prefix)
            if 'text/html' in content_type:
                content_type = 'text/html; charset=utf-8'

        # Alguns devices (ex: firmware Mimosa/Airspan) devolvem um campo
        # "https":false no JSON de login/status e o próprio JS deles compara
        # isso com location.protocol pra decidir se navega pra http:// ou
        # https:// (ver guard _isSchemeSwapNoop no script injetado abaixo,
        # que trata os casos em que esse JS usa location.href/assign/replace
        # — mas nem todo device passa por esses três, e um script.write ou
        # <a>.click() ainda escapariam). Neutralizar na origem: dentro do
        # proxy, o scheme real do browser é sempre o do CRM (https),
        # independente do scheme com que falamos com o device (que já está
        # embutido no path, ex: ".../web/80/http/") — então, quando estamos
        # falando HTTP com o device, forçar esse campo pra "true" evita que o
        # JS dele tente "corrigir" um scheme que não é dele decidir.
        if scheme == 'http' and 'json' in content_type:
            content = re.sub(rb'"https"\s*:\s*false', b'"https":true', content)

        django_resp = HttpResponse(content, content_type=content_type, status=resp.status_code)

        for k, v in resp.headers.items():
            if k.lower() not in ProxyEngine.STRIP_HEADERS:
                try:
                    django_resp[k] = v
                except Exception:
                    pass

        for h in ('X-Frame-Options', 'Content-Security-Policy'):
            if h in django_resp:
                del django_resp[h]


        for cookie_str in getattr(resp, 'cookies_raw', []):
            parts = cookie_str.split(';')
            if parts:
                nv = parts[0].split('=', 1)
                if len(nv) == 2:
                    django_resp.set_cookie(
                        cookie_prefix + nv[0].strip(), nv[1].strip(),
                        path='/', samesite='Lax',
                        secure=request.is_secure()
                    )

        return django_resp

    except Exception as e:
        logger.exception("Erro no ProxyEngine acesso_id=%s: %s", acesso_id, e)
        return HttpResponse(
            ProxyEngine.get_error_page(f"Erro interno no proxy: <code>{str(e)}</code>"),
            content_type='text/html', status=500
        )



# ─────────────────────────────────────────────────────────────────────────────
# EDITOR DE TOPOLOGIA DE REDE
# ─────────────────────────────────────────────────────────────────────────────

from .models import TopologiaDiagrama


def _topologia_perm(request, cliente):
    """Verifica permissão para acessar topologia."""
    return _perms.pode_acessar_cliente(request.user, cliente)


@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def topologia_editor(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _topologia_perm(request, cliente):
        return JsonResponse({'error': 'Sem permissao'}, status=403)
    diagrama = TopologiaDiagrama.objects.filter(cliente=cliente).first()
    return render(request, 'topologia_editor.html', {
        'cliente': cliente,
        'diagrama': diagrama,
        'dados_json': diagrama.dados_json if diagrama else '{"nodes":[],"links":[]}',
        'diagrama_id': diagrama.id if diagrama else None,
    })


@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def topologia_drawio(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _topologia_perm(request, cliente):
        return JsonResponse({'error': 'Sem permissao'}, status=403)
    diagrama = TopologiaDiagrama.objects.filter(cliente=cliente).first()
    return render(request, 'topologia_drawio.html', {
        'cliente': cliente,
        'drawio_xml': diagrama.drawio_xml if diagrama else '',
        'diagrama_id': diagrama.id if diagrama else None,
    })


@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def topologia_dados(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _topologia_perm(request, cliente):
        return JsonResponse({'error': 'Sem permissao'}, status=403)
    diagrama = TopologiaDiagrama.objects.filter(cliente=cliente).first()
    if not diagrama:
        return JsonResponse({'nodes': [], 'links': [], 'diagrama_id': None})
    import json as _json
    try:
        dados = _json.loads(diagrama.dados_json)
    except Exception:
        dados = {'nodes': [], 'links': []}
    dados['diagrama_id'] = diagrama.id
    dados['nome'] = diagrama.nome
    dados['drawio_xml'] = diagrama.drawio_xml or ''
    return JsonResponse(dados)


@login_required(login_url='login')
@require_http_methods(['POST'])
@modulo_habilitado_required('topologia')
def topologia_salvar(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _topologia_perm(request, cliente):
        return JsonResponse({'error': 'Sem permissao'}, status=403)
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Body invalido'}, status=400)
    diagrama, _ = TopologiaDiagrama.objects.get_or_create(
        cliente=cliente,
        defaults={'nome': body.get('nome', 'Nova Topologia')}
    )
    if 'nome' in body:
        diagrama.nome = body['nome']
    if 'dados_json' in body:
        v = body['dados_json']
        diagrama.dados_json = v if isinstance(v, str) else json.dumps(v)
    if 'drawio_xml' in body:
        diagrama.drawio_xml = body['drawio_xml']
    diagrama.save()
    return JsonResponse({'ok': True, 'diagrama_id': diagrama.id, 'nome': diagrama.nome})


@login_required(login_url='login')
@modulo_habilitado_required('topologia')
def topologia_hosts(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not _topologia_perm(request, cliente):
        return JsonResponse({'error': 'Sem permissao'}, status=403)
    acessos = Acesso.objects.filter(cliente=cliente).select_related('funcao', 'modelo')
    hosts = []
    for a in acessos:
        funcao_nome = ((a.funcao.descricao or '') if a.funcao else '').lower()
        tipo_lower = (a.tipo or '').lower()
        tipo = 'host'
        mapa = [
            (['cgnat','cg-nat','carrier grade nat'], 'cgnat'),
            (['bras','bng','broadband network'], 'router'),
            (['router','roteador','core','border','borda'], 'router'),
            (['switch l3','sw-l3','camada 3'], 'switch_l3'),
            (['switch','sw-','catalyst','nexus'], 'switch_l2'),
            (['radio','wireless','ubiquiti','mikrotik','ap ','airmax','ltu'], 'radio'),
            (['dwdm','oadm','ots','mstp','transponder'], 'dwdm'),
            (['olt','gpon','xgs','epon'], 'olt'),
            (['onu','ont'], 'onu'),
            (['server','servidor','zabbix','grafana','proxmox'], 'server'),
            (['firewall','utm','fortigate','pfsense','sophos'], 'firewall'),
            (['vm','virtual machine','virtualizado','kvm','qemu','vmware','vps'], 'vm'),
            (['cpe','modem'], 'cpe'),
        ]
        for keywords, dev_tipo in mapa:
            if any(k in funcao_nome or k in tipo_lower for k in keywords):
                tipo = dev_tipo
                break
        hosts.append({
            'id': a.id,
            'label': a.tipo,
            'ip': a.host,
            'porta': a.porta,
            'protocolo': a.protocolo,
            'usuario': a.usuario,
            'tipo': tipo,
            'cliente_id': cliente.id,
            'funcao': (a.funcao.descricao or '') if a.funcao else '',
            'modelo': (a.modelo.nome or '') if a.modelo else '',
        })
    return JsonResponse({'hosts': hosts, 'total': len(hosts)})


def _mascara_para_prefixo(ip, mascara):
    """Converte 'IP + máscara-ou-prefixo' pro formato 'IP/CIDR' usado nos
    campos de IP P2P do editor de topologia. Aceita máscara decimal
    (255.255.255.252, estilo Cisco/Huawei/ZTE) ou prefixo já numérico
    (30, estilo Datacom/Juniper que já vem com "/"). Silencioso em entradas
    inesperadas — melhor não sugerir IP do que sugerir um CIDR errado."""
    mascara = (mascara or '').strip()
    if mascara.isdigit():
        return f'{ip}/{mascara}'
    if re.match(r'^\d{1,3}(\.\d{1,3}){3}$', mascara):
        try:
            prefixo = sum(bin(int(octeto)).count('1') for octeto in mascara.split('.'))
            return f'{ip}/{prefixo}'
        except Exception:
            return ip
    return ip


def _extrair_interfaces_backup(conteudo, fabricante):
    """Extrai as interfaces configuradas a partir do texto de um backup —
    nome, descrição e (quando existir roteamento IP na interface) o IP/CIDR
    configurado nela — de acordo com a sintaxe de cada fabricante. Usado
    para sugerir 'Interface Lado A/B' no editor de topologia: a descrição
    ajuda a identificar qual interface escolher (ex. "P2P-SW-CORE-P6"), e o
    IP pode preencher automaticamente o campo "IP Local/Remoto (P2P)" do
    link quando a interface escolhida tiver um endereço configurado. Cobre
    os três formatos de config usados pelos templates de backup (ver
    docs/backup_automatico.md):
    - MikroTik (`/export`): nomes vêm de `name=` (renomeados — já funcionam
      como descrição embutida, ex. "sfp-sfpplus2 - P2P-SW-CORE-P6") e
      `interface=` (referências, inclusive nomes default nunca renomeados;
      sem descrição própria). IP vem de `/ip address add address=IP/CIDR
      interface=<nome>`.
    - Juniper (`set interfaces ... | display set`): `set interfaces <if>
      [unit N] description <texto>` / `... family inet address IP/CIDR`.
    - Todo o resto (Cisco/Huawei/Datacom/ZTE/HP/Dell/Extreme/genérico —
      `show running-config`/`display current-configuration`): bloco da
      interface (da linha `interface <nome>` até a próxima linha
      `interface ...`) é varrido por uma linha `description <texto>` e uma
      linha `ip address IP MÁSCARA` (Cisco/Huawei/ZTE, máscara decimal) ou
      `ipv4 address IP/CIDR` (Datacom).

    Retorna lista de {'nome': str, 'descricao': str, 'ip': str}, ordenada
    por nome. `ip` vem vazio quando a interface não tem endereço configurado
    (porta L2 pura, trunk, etc.) — nesse caso o frontend não sugere nada.
    """
    fabricante = (fabricante or '').upper()
    interfaces = {}  # nome -> {'descricao': str, 'ip': str}

    def _add(nome, desc='', ip=''):
        if not nome:
            return
        atual = interfaces.setdefault(nome, {'descricao': '', 'ip': ''})
        if desc and not atual['descricao']:
            atual['descricao'] = desc
        if ip and not atual['ip']:
            atual['ip'] = ip

    if fabricante == 'MIKROTIK':
        for m in re.finditer(r'^/interface\b.*?\bname=("([^"]+)"|(\S+))', conteudo, re.MULTILINE):
            _add(m.group(2) or m.group(3))
        for m in re.finditer(r'\binterface=("([^"]+)"|(\S+))', conteudo):
            _add(m.group(2) or m.group(3))
        for m in re.finditer(
            r'^/ip address add .*?\baddress=(\d{1,3}(?:\.\d{1,3}){3}/\d{1,2})\b.*?\binterface=("([^"]+)"|(\S+))',
            conteudo, re.MULTILINE,
        ):
            ip_cidr, nome = m.group(1), (m.group(3) or m.group(4))
            _add(nome, ip=ip_cidr)

    elif fabricante == 'JUNIPER':
        for m in re.finditer(r'^set interfaces (\S+)(?:\s+unit\s+(\d+))?', conteudo, re.MULTILINE):
            base, unit = m.group(1), m.group(2)
            _add(base)
            if unit:
                _add(f'{base}.{unit}')
        for m in re.finditer(r'^set interfaces (\S+) description (.+?)\s*$', conteudo, re.MULTILINE):
            _add(m.group(1), desc=m.group(2).strip())
        for m in re.finditer(r'^set interfaces (\S+) unit (\d+) description (.+?)\s*$', conteudo, re.MULTILINE):
            _add(f'{m.group(1)}.{m.group(2)}', desc=m.group(3).strip())
        for m in re.finditer(
            r'^set interfaces (\S+) unit (\d+) family inet address (\d{1,3}(?:\.\d{1,3}){3}/\d{1,2})',
            conteudo, re.MULTILINE,
        ):
            _add(f'{m.group(1)}.{m.group(2)}', ip=m.group(3))

    else:
        linhas = conteudo.splitlines()
        n = len(linhas)
        i = 0
        while i < n:
            m = re.match(r'^interface\s+(.+?)\s*$', linhas[i])
            if not m:
                i += 1
                continue
            nome = m.group(1).strip()
            # Fim do bloco desta interface = próxima linha "interface <algo>".
            j = i + 1
            while j < n and not re.match(r'^interface\s+\S', linhas[j]):
                j += 1
            # Exclui sub-interface por ONU (ex. ZTE "gpon-onu_1/2/1:5" — uma
            # linha "interface" por ONU registrada na porta GPON). Nunca é o
            # lado de um link de topologia (isso é a porta física do OLT/PON,
            # não a ONU do cliente), e em OLTs grandes ofusca as interfaces
            # físicas relevantes (uplinks, PON) dentro do limite de 500.
            if not nome or re.search(r':\d+$', nome):
                i = j
                continue
            desc, ip = '', ''
            for bloco_linha in linhas[i + 1:j]:
                if not desc:
                    dm = re.match(r'^\s*description\s+(.+?)\s*$', bloco_linha)
                    if dm:
                        desc = dm.group(1).strip()
                if not ip:
                    im = re.match(
                        r'^\s*ip(?:v4)?\s+address\s+(\d{1,3}(?:\.\d{1,3}){3})[/\s]+(\S+)', bloco_linha)
                    if im:
                        ip = _mascara_para_prefixo(im.group(1), im.group(2))
                if desc and ip:
                    break
            _add(nome, desc=desc, ip=ip)
            i = j

    itens = [{'nome': nome, **dados} for nome, dados in interfaces.items()]
    itens.sort(key=lambda it: it['nome'].lower())
    return itens[:500]


@login_required(login_url='login')
@require_http_methods(['GET'])
@modulo_habilitado_required('acessos')
def interfaces_backup_acesso(request, acesso_id):
    """Lista as interfaces (nome, descrição e IP/CIDR quando roteada)
    encontradas no backup mais recente do acesso, para sugestão nos campos
    'Interface Lado A/B' e preenchimento automático do IP P2P do editor de
    topologia. Sem backup disponível (ou arquivo ausente no disco), retorna
    lista vazia — o campo continua editável em texto livre."""
    acesso = get_object_or_404(Acesso, id=acesso_id)

    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    backup = BackupLog.objects.filter(
        acesso=acesso, status__in=['SUCESSO', 'PARCIAL'],
    ).exclude(arquivo_path='').select_related('template').order_by('-data_backup').first()

    if not backup:
        return JsonResponse({'interfaces': [], 'tem_backup': False})

    arquivo_path = os.path.join(settings.MEDIA_ROOT, backup.arquivo_path)
    if not os.path.exists(arquivo_path):
        return JsonResponse({'interfaces': [], 'tem_backup': False})

    LIMITE = 2 * 1024 * 1024  # 2MB — suficiente para qualquer running-config
    with open(arquivo_path, 'r', encoding='utf-8', errors='replace') as f:
        conteudo = f.read(LIMITE)

    fabricante = backup.template.fabricante if backup.template else ''
    interfaces = _extrair_interfaces_backup(conteudo, fabricante)

    return JsonResponse({
        'interfaces': interfaces,
        'tem_backup': True,
        'data_backup': backup.data_backup.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M'),
    })


# ─── L2VPN (VSI/VPLS, VPWS, L2VC) na topologia ────────────────────────────────

from django.core.cache import cache
from .l2vpn_parser import parse_l2vpn, extrair_ips_identidade, resumo_por_tipo
from .l2vpn_actions import (
    L2vpnNaoSuportado,
    VENDORS_SUPORTADOS as _L2VPN_VENDORS,
    comandos_clonar,
    executar as _l2vpn_executar,
    sugerir_id,
    validar_comandos_editados,
    validar_spec,
)

_L2VPN_LIMITE_ARQUIVO = 3 * 1024 * 1024   # running-config completa cabe folgado
_L2VPN_CACHE_TTL = 6 * 60 * 60            # 6h — o backup em si roda 1x/dia
# Quantos backups anteriores tentar quando o mais recente não revela nenhum IP
# de identidade do equipamento. Loopback/LSR-ID praticamente não mudam, e é
# comum a coleta mais nova vir truncada (paginação `---- More ----`) — sem esse
# fallback um host truncado "some" do mapa e os túneis que apontam pra ele
# aparecem como não identificados.
_L2VPN_FALLBACK_BACKUPS = 5


def _backups_recentes(acesso, limite=1):
    """Últimos backups aproveitáveis do acesso (mais novo primeiro), já
    filtrando os que perderam o arquivo no disco."""
    logs = BackupLog.objects.filter(
        acesso=acesso, status__in=['SUCESSO', 'PARCIAL'],
    ).exclude(arquivo_path='').select_related('template').order_by('-data_backup')[:limite * 3]
    encontrados = []
    for log in logs:
        caminho = os.path.join(settings.MEDIA_ROOT, log.arquivo_path)
        if os.path.exists(caminho):
            encontrados.append((log, caminho))
        if len(encontrados) >= limite:
            break
    return encontrados


def _ler_backup(caminho):
    with open(caminho, 'r', encoding='utf-8', errors='replace') as arquivo:
        return arquivo.read(_L2VPN_LIMITE_ARQUIVO)


def _l2vpn_servicos_do_acesso(acesso):
    """(serviços, BackupLog) do backup mais recente do acesso. Cacheado por id
    do BackupLog — o conteúdo de um backup nunca muda depois de gravado, então
    só um backup novo invalida (a chave muda junto)."""
    recentes = _backups_recentes(acesso, limite=1)
    if not recentes:
        return [], None
    log, caminho = recentes[0]
    chave = f'l2vpn:svc:{log.id}'
    servicos = cache.get(chave)
    if servicos is None:
        servicos = parse_l2vpn(_ler_backup(caminho))
        cache.set(chave, servicos, _L2VPN_CACHE_TTL)
    return servicos, log


def _l2vpn_ips_identidade(acesso):
    """{IP: origem} do acesso, cacheado por acesso (e não por backup: aqui o
    que interessa é o valor mais recente conhecido, não um backup específico)."""
    chave = f'l2vpn:ident:acesso:{acesso.id}'
    ips = cache.get(chave)
    if ips is None:
        ips = {}
        for log, caminho in _backups_recentes(acesso, limite=_L2VPN_FALLBACK_BACKUPS):
            ips = extrair_ips_identidade(_ler_backup(caminho))
            if ips:
                break
        cache.set(chave, ips, _L2VPN_CACHE_TTL)
    return ips


def _l2vpn_mapa_identidade(acessos):
    """{IP: {acesso…}} com os IPs pelos quais cada host do cliente é conhecido
    pelos vizinhos MPLS (loopback/LSR-ID/router-id) mais o IP de gerência do
    CRM. É esse mapa que transforma um `peer 198.18.255.2` no host do outro
    lado do túnel."""
    mapa = {}

    def _registrar(ip, acesso, origem):
        if not ip or ip in mapa:
            return
        mapa[ip] = {
            'acesso_id': acesso.id,
            'nome': acesso.tipo,
            'host': acesso.host,
            'porta': acesso.porta,
            'protocolo': acesso.protocolo,
            'origem': origem,
        }

    for acesso in acessos:
        for ip, origem in _l2vpn_ips_identidade(acesso).items():
            _registrar(ip, acesso, origem)

    # O IP de gerência entra por último: quando o equipamento usa o próprio IP
    # de gerência como LSR-ID, o mapa já pegou pelo backup (com a origem certa).
    for acesso in acessos:
        _registrar(acesso.host, acesso, 'IP de gerência (CRM)')

    return mapa


@login_required(login_url='login')
@require_http_methods(['GET'])
@modulo_habilitado_required('acessos')
def l2vpn_backup_acesso(request, acesso_id):
    """Serviços L2VPN (VSI/VPLS, VPWS e L2VC) documentados a partir do backup
    mais recente do host, com cada peer do túnel já resolvido para o host do
    outro lado quando ele também está cadastrado no cliente.

    Alimenta o "Mostrar VSI / L2VPN" do editor de topologia: sem backup, ou
    sem serviço L2 configurado, devolve lista vazia (o modal explica o motivo
    em vez de sumir)."""
    acesso = get_object_or_404(Acesso.objects.select_related('cliente'), id=acesso_id)

    if not _perms.pode_acessar_cliente(request.user, acesso.cliente):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    servicos, log = _l2vpn_servicos_do_acesso(acesso)
    if log is None:
        return JsonResponse({
            'tem_backup': False, 'servicos': [], 'resumo': resumo_por_tipo([]),
            'host': {'id': acesso.id, 'nome': acesso.tipo, 'ip': acesso.host},
        })

    irmaos = list(Acesso.objects.filter(cliente=acesso.cliente).exclude(id=acesso.id))
    mapa = _l2vpn_mapa_identidade(irmaos)

    # Anota cada peer com o host do outro lado (quando identificado). Copia o
    # serviço para não gravar o resultado da resolução no valor cacheado.
    anotados = []
    for servico in servicos:
        servico = dict(servico)
        servico['peers'] = [dict(peer, destino=mapa.get(peer['ip'])) for peer in servico['peers']]
        anotados.append(servico)

    identidade_local = _l2vpn_ips_identidade(acesso)

    return JsonResponse({
        'tem_backup': True,
        'servicos': anotados,
        'resumo': resumo_por_tipo(anotados),
        'host': {
            'id': acesso.id, 'nome': acesso.tipo, 'ip': acesso.host,
            'ips_identidade': identidade_local,
        },
        'data_backup': log.data_backup.astimezone(
            timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M'),
        'arquivo': os.path.basename(log.arquivo_path),
        # Sugestão de id livre e se dá pra clonar neste equipamento — a UI usa
        # pra pré-preencher o formulário e pra esconder o botão onde não dá.
        'id_sugerido': sugerir_id(anotados),
        'pode_clonar': bool(anotados) and anotados[0].get('vendor') in _L2VPN_VENDORS,
    })


@login_required(login_url='login')
@require_http_methods(['POST'])
@modulo_habilitado_required('topologia')
def l2vpn_clonar_acesso(request, acesso_id):
    """Clona um serviço L2VPN do host: monta a config do serviço novo a partir
    de um existente e (fora do preview) aplica no equipamento.

    body: {"origem_idx": int, "spec": {...}, "preview": bool, "comandos": [...]}

    Mesmo contrato da automação BGP (`bgp_views.bgp_executar_acao`):
    `preview=true` só devolve os comandos gerados, sem tocar no equipamento —
    é o que preenche o textarea editável do modal. `preview=false` executa de
    verdade e grava `AcaoL2vpn`; se vier `comandos`, usa exatamente o texto
    revisado pelo operador em vez de gerar de novo.

    Restrito a backoffice: criar pseudowire em equipamento de produção é
    engenharia de rede, não função de portal de cliente.
    """
    acesso = get_object_or_404(Acesso.objects.select_related('cliente'), id=acesso_id)

    if not (_perms.is_backoffice(request.user)
            and _perms.ferramenta_habilitada(request.user, 'topologia')
            and _perms.pode_acessar_cliente(request.user, acesso.cliente)):
        return JsonResponse({'error': 'Sem permissão'}, status=403)

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    servicos, log = _l2vpn_servicos_do_acesso(acesso)
    if log is None or not servicos:
        return JsonResponse({'error': 'Sem serviços L2VPN no backup deste host.'}, status=404)

    try:
        origem_idx = int(body.get('origem_idx'))
        origem = servicos[origem_idx]
    except (TypeError, ValueError, IndexError):
        return JsonResponse({'error': 'Serviço de origem inválido.'}, status=400)

    try:
        spec = validar_spec(body.get('spec') or {}, servicos, origem)
        comandos_gerados = comandos_clonar(spec)
    except L2vpnNaoSuportado as e:
        return JsonResponse({'error': str(e)}, status=422)

    if bool(body.get('preview', True)):
        # Preview sempre gera do zero — é o texto inicial do textarea, nunca
        # deve refletir uma edição anterior.
        return JsonResponse({'comandos': comandos_gerados, 'vendor': spec['vendor']})

    comandos_editados = body.get('comandos')
    if comandos_editados is not None:
        try:
            comandos = validar_comandos_editados(comandos_editados)
        except L2vpnNaoSuportado as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        comandos = comandos_gerados

    output, status = _l2vpn_executar(acesso, spec['vendor'], comandos)

    AcaoL2vpn.objects.create(
        acesso=acesso, usuario=request.user,
        origem=f'{origem.get("tecnologia", "")} {origem.get("id", "")} {origem.get("nome", "")}'.strip(),
        alvo=spec['nome'], servico_id=str(spec['id']), vendor=spec['vendor'],
        comandos='\n'.join(comandos), output=output, status=status,
    )

    if status == 'sucesso':
        # O serviço novo só vai aparecer na listagem depois do próximo backup —
        # o cache guarda o parse do backup atual, que não muda. Deixa explícito
        # na resposta pra UI avisar em vez de o operador achar que falhou.
        logger.info(f'L2VPN clonado em {acesso}: {spec["nome"]} (id {spec["id"]}) por {request.user}')

    return JsonResponse({
        'status': status, 'output': output, 'comandos': comandos,
        'nome': spec['nome'], 'id': spec['id'],
    })


def _exec_migration_topologia(request):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return HttpResponse('Proibido', status=403)
    try:
        from django.db import connection as _conn
        with _conn.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS clientes_topologiadiagrama (
                    id BIGSERIAL PRIMARY KEY,
                    nome VARCHAR(255) NOT NULL DEFAULT 'Nova Topologia',
                    dados_json TEXT NOT NULL DEFAULT '{"nodes":[],"links":[]}',
                    drawio_xml TEXT NOT NULL DEFAULT '',
                    atualizado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    criado_em TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    cliente_id BIGINT NOT NULL REFERENCES clientes_cliente(id) ON DELETE CASCADE
                );
                INSERT INTO django_migrations (app, name, applied)
                VALUES ('clientes', '0040_topologia_diagrama', NOW())
                ON CONFLICT (app, name) DO NOTHING;
            """)
        return HttpResponse('Migration aplicada com sucesso!', status=200)
    except Exception as e:
        return HttpResponse(f'Erro: {e}', status=500)


# =============================================================================
# VPN WireGuard
# =============================================================================
from .models import VPNWireGuard, VPNServidorConfig
from . import vpn_manager as wgm
import json as _json

def _get_servidor_config():
    """Retorna config do servidor, criando se não existir."""
    cfg = VPNServidorConfig.objects.first()
    if not cfg:
        priv, pub = wgm.gerar_par_chaves()
        cfg = VPNServidorConfig.objects.create(
            servidor_private_key=priv,
            servidor_public_key=pub,
            servidor_endpoint='179.48.68.73',
            servidor_porta=51820,
        )
    return cfg


@login_required
@require_http_methods(["GET"])
@modulo_habilitado_required('tuneis')
def vpn_wg_listar(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    vpns    = VPNWireGuard.objects.filter(cliente=cliente)
    cfg     = _get_servidor_config()

    # Status em tempo real
    peers_status = {}
    try:
        peers_status = wgm.get_peers_status()
    except Exception:
        pass

    vpns_data = []
    for v in vpns:
        st = peers_status.get(v.cliente_public_key, {})
        vpns_data.append({
            'id':             v.id,
            'nome':           v.nome,
            'vpn_ip':         v.vpn_ip,
            'redes':          v.redes_lista(),
            'ativo':          v.ativo,
            'peer_no_servidor': v.peer_no_servidor,
            'conectado':      st.get('conectado', False),
            'last_handshake': st.get('last_handshake', 0),
            'rx':             wgm.formatar_bytes(st.get('rx_bytes', 0)),
            'tx':             wgm.formatar_bytes(st.get('tx_bytes', 0)),
            'criado_em':      v.criado_em.strftime('%d/%m/%Y %H:%M'),
        })

    return JsonResponse({'vpns': vpns_data, 'servidor_endpoint': cfg.servidor_endpoint})


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_wg_criar(request, cliente_id):
    """
    Cria uma VPN WireGuard em uma interface ISOLADA dedicada (wg5, wg6, ...)
    — cada cliente tem sua própria interface/porta/sub-rede, nunca
    compartilhando rotas de kernel com outro cliente. Isso evita a classe de
    bug em que excluir a VPN de UM cliente apagava rotas de OUTRO (incidente
    Conecta ISP, 2026-06-14 — ver docs/vpn_wireguard.md).
    """
    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        body = _json.loads(request.body)
        nome          = body.get('nome', 'VPN MikroTik').strip() or 'VPN MikroTik'
        redes_raw     = body.get('redes_privadas', '').strip()

        cfg           = _get_servidor_config()
        priv, pub     = wgm.gerar_par_chaves()
        psk           = wgm.gerar_preshared_key()

        interface, porta, subnet_n = wgm.alocar_proxima_interface()
        vpn_ip            = f'10.{subnet_n}.0.2'
        servidor_ip_local = f'10.{subnet_n}.0.1'

        vpn = VPNWireGuard.objects.create(
            cliente=cliente,
            nome=nome,
            cliente_private_key=priv,
            cliente_public_key=pub,
            preshared_key=psk,
            vpn_ip=vpn_ip,
            redes_privadas=redes_raw,
            ativo=True,
            interface_nome=interface,
            servidor_ip_local=servidor_ip_local,
        )

        # Criar interface dedicada e adicionar o peer só nela
        try:
            wgm.criar_interface_isolada(interface, porta, subnet_n, cfg.servidor_private_key)
            wgm.adicionar_peer_isolado(interface, pub, psk, vpn_ip, vpn.redes_lista())
            vpn.peer_no_servidor = True
            vpn.save()
        except Exception as e:
            logger.warning(f'Peer não adicionado em {interface}: {e}')

        return JsonResponse({'ok': True, 'vpn_id': vpn.id, 'vpn_ip': vpn_ip})

    except Exception as e:
        logger.error(f'vpn_wg_criar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
@modulo_habilitado_required('tuneis')
def vpn_wg_script(request, vpn_id):
    vpn = get_object_or_404(VPNWireGuard, id=vpn_id)
    cfg = _get_servidor_config()
    script = wgm.gerar_script_mikrotik(vpn, cfg)
    return JsonResponse({'ok': True, 'script': script, 'nome': vpn.nome})


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_wg_deletar(request, vpn_id):
    vpn = get_object_or_404(VPNWireGuard, id=vpn_id)
    try:
        if vpn.peer_no_servidor:
            if wgm.vpn_e_isolada(vpn.vpn_ip):
                wgm.remover_interface_isolada(vpn.interface_nome)
            else:
                wgm.remover_peer(vpn.cliente_public_key, vpn.redes_lista())
                wgm.salvar_config_persistente()
        vpn.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        logger.error(f'vpn_wg_deletar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
@modulo_habilitado_required('tuneis')
def vpn_wg_status(request, cliente_id):
    vpns = VPNWireGuard.objects.filter(cliente_id=cliente_id, ativo=True)
    peers = {}
    try:
        peers = wgm.get_peers_status()
    except Exception:
        pass

    result = []
    for v in vpns:
        st = peers.get(v.cliente_public_key, {})
        result.append({
            'id':        v.id,
            'vpn_ip':    v.vpn_ip,
            'conectado': st.get('conectado', False),
            'last_handshake': st.get('last_handshake', 0),
        })
    return JsonResponse({'vpns': result})


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_wg_reativar_peer(request, vpn_id):
    """Re-adiciona peer ao servidor (útil após reboot ou falha pontual)."""
    vpn = get_object_or_404(VPNWireGuard, id=vpn_id)
    cfg = _get_servidor_config()
    try:
        if wgm.vpn_e_isolada(vpn.vpn_ip):
            porta    = wgm.ISOLATED_BASE_PORT + (wgm.interface_subnet_n(vpn.interface_nome) - wgm.ISOLATED_SUBNET_BASE)
            subnet_n = wgm.interface_subnet_n(vpn.interface_nome)
            wgm.criar_interface_isolada(vpn.interface_nome, porta, subnet_n, cfg.servidor_private_key)
            wgm.adicionar_peer_isolado(vpn.interface_nome, vpn.cliente_public_key,
                                       vpn.preshared_key, vpn.vpn_ip, vpn.redes_lista())
        else:
            wgm.adicionar_peer(vpn.cliente_public_key, vpn.preshared_key,
                               vpn.vpn_ip, vpn.redes_lista())
            wgm.salvar_config_persistente()
        vpn.peer_no_servidor = True
        vpn.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_wg_editar(request, vpn_id):
    """Atualiza nome e redes privadas de uma VPN WireGuard."""
    vpn = get_object_or_404(VPNWireGuard, id=vpn_id)
    try:
        body = _json.loads(request.body)
        nome_novo   = body.get('nome', '').strip() or vpn.nome
        redes_novas = body.get('redes_privadas', '').strip()

        redes_antigas = vpn.redes_lista()

        vpn.nome           = nome_novo
        vpn.redes_privadas = redes_novas
        vpn.save()

        if vpn.peer_no_servidor:
            # Remover rotas antigas e re-adicionar peer com novas redes
            if wgm.vpn_e_isolada(vpn.vpn_ip):
                wgm.adicionar_peer_isolado(vpn.interface_nome, vpn.cliente_public_key,
                                           vpn.preshared_key, vpn.vpn_ip, vpn.redes_lista())
            else:
                wgm.remover_peer(vpn.cliente_public_key, redes_antigas)
                wgm.adicionar_peer(vpn.cliente_public_key, vpn.preshared_key,
                                   vpn.vpn_ip, vpn.redes_lista())
                wgm.salvar_config_persistente()

        return JsonResponse({'ok': True})
    except Exception as e:
        logger.error(f'vpn_wg_editar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


# ══════════════════════════════════════════════════════════════════════════════
# Túnel OpenVPN (aba Túneis) — servidor único da CRM, isolado por CN + CCD.
# NÃO CONFUNDIR com a seção "OpenVPN — Configuração automatizada em MikroTik"
# logo abaixo: aquela configura o MikroTik do CLIENTE como servidor OpenVPN
# para acesso remoto do NOC; esta aqui é o terceiro tipo de túnel da aba
# Túneis (ao lado de SSH e WireGuard) — a CRM É o servidor, o MikroTik do
# cliente é o client, com rota isolada nativamente via client-config-dir.
# ══════════════════════════════════════════════════════════════════════════════

from .models import VPNOpenVPN
from . import openvpn_tunnel_manager as ovpnm


@login_required
@require_http_methods(["GET"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_listar(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    vpns = VPNOpenVPN.objects.filter(cliente=cliente)
    vpns_data = [{
        'id':          v.id,
        'nome':        v.nome,
        'vpn_ip':      v.vpn_ip,
        'redes':       v.redes_lista(),
        'ativo':       v.ativo,
        'cert_emitido': v.cert_emitido,
        'criado_em':   v.criado_em.strftime('%d/%m/%Y %H:%M'),
    } for v in vpns]
    return JsonResponse({'vpns': vpns_data, 'endpoint': ovpnm.OVPN_ENDPOINT_HOST})


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_criar(request, cliente_id):
    """
    Cria um túnel OpenVPN em uma instância DEDICADA (porta/interface/sub-rede
    próprias) — mesmo padrão isolado já usado pelo WireGuard
    (VPNWireGuard.interface_nome). Evita que dois clientes com as mesmas
    redes "alcançáveis" (o padrão CGNAT+RFC1918) tenham tráfego roteado pro
    cliente errado quando ambos estão conectados ao mesmo tempo.
    """
    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        body = _json.loads(request.body)
        nome      = body.get('nome', 'VPN MikroTik').strip() or 'VPN MikroTik'
        redes_raw = body.get('redes_privadas', '').strip() or '\n'.join(ovpnm.REDES_PADRAO)

        common_name = ovpnm.gerar_common_name(cliente)
        interface_nome, porta, subnet_n = ovpnm.alocar_proxima_instancia()
        vpn_ip = ovpnm._client_ip(subnet_n)

        vpn = VPNOpenVPN.objects.create(
            cliente=cliente,
            nome=nome,
            common_name=common_name,
            redes_privadas=redes_raw,
            vpn_ip=vpn_ip,
            porta=porta,
            interface_nome=interface_nome,
            subnet_n=subnet_n,
        )

        try:
            ovpnm.emitir_certificado_cliente(common_name)
            ovpnm.criar_instancia_servidor(vpn)
        except Exception:
            # Qualquer falha no meio do provisionamento (cert, instância) não
            # deve deixar um registro "pela metade" no banco — apaga o que já
            # foi gerado e propaga o erro original.
            for ext in ('key', 'crt'):
                path = f'{ovpnm.PKI_DIR}/clients/{common_name}.{ext}'
                if os.path.exists(path):
                    os.remove(path)
            ovpnm.remover_instancia_servidor(vpn)
            vpn.delete()
            raise

        vpn.cert_emitido = True
        vpn.save(update_fields=['cert_emitido'])

        return JsonResponse({'ok': True, 'vpn_id': vpn.id, 'vpn_ip': vpn_ip})
    except Exception as e:
        logger.error(f'vpn_ovpn_criar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_editar(request, vpn_id):
    vpn = get_object_or_404(VPNOpenVPN, id=vpn_id)
    try:
        body = _json.loads(request.body)
        vpn.nome           = body.get('nome', '').strip() or vpn.nome
        vpn.redes_privadas = body.get('redes_privadas', '').strip()
        vpn.save()

        # Reescreve e reinicia SÓ a instância deste cliente — outros túneis
        # rodam em processos separados, não são afetados.
        ovpnm.atualizar_redes_instancia(vpn)

        return JsonResponse({'ok': True})
    except Exception as e:
        logger.error(f'vpn_ovpn_editar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_deletar(request, vpn_id):
    vpn = get_object_or_404(VPNOpenVPN, id=vpn_id)
    try:
        ovpnm.remover_instancia_servidor(vpn)
        ovpnm.revogar_certificado(vpn.common_name)
        vpn.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        logger.error(f'vpn_ovpn_deletar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_reativar(request, vpn_id):
    """Re-emite certificado e recria a instância (útil se os arquivos em disco
    tiverem sido perdidos ou o serviço tiver sido derrubado manualmente)."""
    vpn = get_object_or_404(VPNOpenVPN, id=vpn_id)
    try:
        ovpnm.emitir_certificado_cliente(vpn.common_name)
        ovpnm.criar_instancia_servidor(vpn)
        vpn.cert_emitido = True
        vpn.ativo = True
        vpn.save(update_fields=['cert_emitido', 'ativo'])
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_bootstrap(request, vpn_id):
    """One-liner de bootstrap (fetch + import) para colar no terminal do Mikrotik."""
    vpn = get_object_or_404(VPNOpenVPN, id=vpn_id)
    return JsonResponse({'ok': True, 'script': ovpnm.gerar_oneliner_bootstrap(vpn.token)})


@login_required
@require_http_methods(["POST"])
@modulo_habilitado_required('tuneis')
def vpn_ovpn_regenerar_token(request, vpn_id):
    """Invalida o link de bootstrap antigo e gera um novo (botão 'Novo')."""
    import secrets
    vpn = get_object_or_404(VPNOpenVPN, id=vpn_id)
    vpn.token = secrets.token_urlsafe(32)
    vpn.save(update_fields=['token'])
    return JsonResponse({'ok': True, 'script': ovpnm.gerar_oneliner_bootstrap(vpn.token)})


# ── Endpoints públicos (sem login) — o próprio Mikrotik do cliente busca ────
# estes arquivos via /tool fetch. Protegidos só pelo token (32 bytes
# aleatórios, igual ao padrão já usado em firmware_download/
# FirmwareCompartilhamento.token) — nunca ficam listados/indexados, e
# "Novo" no botão do bootstrap invalida o token antigo imediatamente.

@require_http_methods(["GET"])
def vpn_ovpn_setup_rsc(request, token):
    vpn = get_object_or_404(VPNOpenVPN, token=token, ativo=True)
    ros_version = request.GET.get('v', '')
    script = ovpnm.gerar_setup_rsc(vpn, ros_version=ros_version)
    return HttpResponse(script, content_type='text/plain; charset=utf-8')


@require_http_methods(["GET"])
def vpn_ovpn_setup_arquivo(request, token, nome_arquivo):
    vpn = get_object_or_404(VPNOpenVPN, token=token, ativo=True)
    if nome_arquivo == 'ca.crt':
        with open(ovpnm.CA_CRT) as f:
            conteudo = f.read()
    elif nome_arquivo in ('client.crt', 'client.key'):
        key_pem, crt_pem = ovpnm.ler_certificado_cliente(vpn.common_name)
        if key_pem is None:
            raise Http404('Certificado não encontrado')
        conteudo = crt_pem if nome_arquivo == 'client.crt' else key_pem
    else:
        raise Http404('Arquivo desconhecido')
    return HttpResponse(conteudo, content_type='text/plain; charset=utf-8')


# ══════════════════════════════════════════════════════════════════════════════
# OpenVPN — Configuração automatizada em MikroTik
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@modulo_habilitado_required('vpn')
def openvpn_listar(request, cliente_id):
    """Lista as configurações OpenVPN do cliente incluindo usuários adicionais."""
    from .models import OpenVPNConfig
    cliente = get_object_or_404(Cliente, id=cliente_id)
    qs = OpenVPNConfig.objects.filter(cliente=cliente).select_related('acesso').prefetch_related('usuarios')
    data = []
    for c in qs:
        usuarios = [
            {
                'id':       u.id,
                'nome':     u.nome,
                'username': u.username,
                'status':   u.status,
                'erro_msg': u.erro_msg,
                'tem_arquivo': bool(u.ovpn_path),
            }
            for u in c.usuarios.all()
        ]
        data.append({
            'id':          c.id,
            'nome_vpn':    c.nome_vpn,
            'ip_publico':  c.ip_publico,
            'porta':       c.porta,
            'ros_version': c.ros_version,
            'status':      c.status,
            'erro_msg':    c.erro_msg,
            'acesso_host': c.acesso.host if c.acesso else '',
            'acesso_tipo': c.acesso.tipo if c.acesso else '',
            'vpn_username': c.vpn_username,
            'vpn_password': c.vpn_password,
            'tem_arquivo': bool(c.ovpn_path),
            'criado_em':   c.criado_em.strftime('%d/%m/%Y %H:%M'),
            'usuarios':    usuarios,
        })
    return JsonResponse({'ok': True, 'configs': data})


@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('vpn')
def openvpn_criar(request, cliente_id):
    """Cria uma nova configuração OpenVPN e inicia a execução em background."""
    import threading
    from .models import OpenVPNConfig
    from .openvpn_manager import gerar_senha, executar_config_openvpn

    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        body       = json.loads(request.body)
        acesso_id  = body.get('acesso_id')
        ip_publico = body.get('ip_publico', '').strip()
        porta      = int(body.get('porta', 61194))
        nome_vpn   = body.get('nome_vpn', '').strip()
        ros_ver    = body.get('ros_version', '7')

        if not ip_publico or not nome_vpn or not acesso_id:
            return JsonResponse({'ok': False, 'erro': 'ip_publico, nome_vpn e acesso_id são obrigatórios'}, status=400)

        acesso = get_object_or_404(Acesso, id=acesso_id, cliente=cliente)

        config = OpenVPNConfig.objects.create(
            cliente         = cliente,
            acesso          = acesso,
            nome_vpn        = nome_vpn,
            ip_publico      = ip_publico,
            porta           = porta,
            ros_version     = ros_ver,
            vpn_pool        = body.get('vpn_pool', '192.168.250.128-192.168.250.254'),
            vpn_local_ip    = body.get('vpn_local_ip', '192.168.250.1'),
            vpn_username    = body.get('vpn_username') or nome_vpn,
            vpn_password    = body.get('vpn_password') or gerar_senha(14),
            cert_passphrase = body.get('cert_passphrase') or gerar_senha(10),
            rate_limit      = body.get('rate_limit', '50M/50M'),
            status          = 'configurando',
        )

        # Executa em background para não bloquear o request
        t = threading.Thread(target=executar_config_openvpn, args=(config.id,), daemon=True)
        t.start()

        return JsonResponse({'ok': True, 'id': config.id})
    except Exception as e:
        logger.error(f'openvpn_criar: {e}')
        return JsonResponse({'ok': False, 'erro': str(e)}, status=400)


@login_required
@modulo_habilitado_required('vpn')
def openvpn_status(request, config_id):
    """Retorna o status atual de uma configuração OpenVPN (usado para polling)."""
    from .models import OpenVPNConfig
    c = get_object_or_404(OpenVPNConfig, id=config_id)
    return JsonResponse({
        'ok':        True,
        'status':    c.status,
        'erro_msg':  c.erro_msg,
        'tem_arquivo': bool(c.ovpn_path),
    })


@login_required
@modulo_habilitado_required('vpn')
def openvpn_download(request, config_id):
    """Faz o download do arquivo .ovpn gerado."""
    from .models import OpenVPNConfig
    c = get_object_or_404(OpenVPNConfig, id=config_id)
    if not c.ovpn_path:
        return JsonResponse({'ok': False, 'erro': 'Arquivo ainda não gerado'}, status=404)
    caminho = os.path.join(settings.MEDIA_ROOT, c.ovpn_path)
    if not os.path.exists(caminho):
        return JsonResponse({'ok': False, 'erro': 'Arquivo não encontrado no servidor'}, status=404)
    return FileResponse(
        open(caminho, 'rb'),
        as_attachment=True,
        filename=f'{c.nome_vpn}.ovpn',
    )


@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('vpn')
def openvpn_deletar(request, config_id):
    """Remove uma configuração OpenVPN e seu arquivo .ovpn."""
    from .models import OpenVPNConfig
    c = get_object_or_404(OpenVPNConfig, id=config_id)
    if c.ovpn_path:
        try:
            caminho = os.path.join(settings.MEDIA_ROOT, c.ovpn_path)
            if os.path.exists(caminho):
                os.remove(caminho)
        except Exception:
            pass
    c.delete()
    return JsonResponse({'ok': True})


@login_required
@modulo_habilitado_required('vpn')
def openvpn_logs(request, config_id):
    """Retorna os logs de execução de uma configuração OpenVPN."""
    from .models import OpenVPNConfig
    c = get_object_or_404(OpenVPNConfig, id=config_id)
    return JsonResponse({'ok': True, 'logs': c.logs or '(sem logs)'})


@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('vpn')
def openvpn_reexecutar(request, config_id):
    """Re-executa a configuração OpenVPN (útil quando houve erro)."""
    import threading
    from .models import OpenVPNConfig
    from .openvpn_manager import executar_config_openvpn
    c = get_object_or_404(OpenVPNConfig, id=config_id)
    c.status   = 'configurando'
    c.erro_msg = ''
    c.save(update_fields=['status', 'erro_msg'])
    t = threading.Thread(target=executar_config_openvpn, args=(c.id,), daemon=True)
    t.start()
    return JsonResponse({'ok': True})


# ── OpenVPN — Usuários adicionais ─────────────────────────────────────────────

@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('vpn')
def openvpn_usuario_criar(request, config_id):
    import threading
    import json
    from .models import OpenVPNConfig, OpenVPNUsuario
    from .openvpn_manager import adicionar_usuario_openvpn, gerar_senha

    config = get_object_or_404(OpenVPNConfig, id=config_id)
    try:
        data     = json.loads(request.body)
        nome     = data.get('nome', '').strip()
        username = data.get('username', '').strip() or nome
        if not nome:
            return JsonResponse({'ok': False, 'erro': 'Nome é obrigatório'})

        u = OpenVPNUsuario.objects.create(
            config=config, nome=nome,
            username=username, password=gerar_senha(16),
        )
        t = threading.Thread(target=adicionar_usuario_openvpn, args=(u.id,), daemon=True)
        t.start()
        return JsonResponse({'ok': True, 'id': u.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'erro': str(e)})


@login_required
@modulo_habilitado_required('vpn')
def openvpn_usuario_status(request, usuario_id):
    from .models import OpenVPNUsuario
    u = get_object_or_404(OpenVPNUsuario, id=usuario_id)
    return JsonResponse({
        'status': u.status, 'erro_msg': u.erro_msg,
        'tem_arquivo': bool(u.ovpn_path),
    })


@login_required
@modulo_habilitado_required('vpn')
def openvpn_usuario_download(request, usuario_id):
    from .models import OpenVPNUsuario
    from django.http import FileResponse
    u = get_object_or_404(OpenVPNUsuario, id=usuario_id)
    if not u.ovpn_path:
        return JsonResponse({'erro': 'Arquivo não disponível'}, status=404)
    path = os.path.join(settings.MEDIA_ROOT, u.ovpn_path)
    return FileResponse(open(path, 'rb'), as_attachment=True,
                        filename=f'{u.nome}.ovpn')


@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('vpn')
def openvpn_usuario_deletar(request, usuario_id):
    from .models import OpenVPNUsuario
    u = get_object_or_404(OpenVPNUsuario, id=usuario_id)
    if u.ovpn_path:
        try:
            os.remove(os.path.join(settings.MEDIA_ROOT, u.ovpn_path))
        except Exception:
            pass
    u.delete()
    return JsonResponse({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
# IRR Config — Atualização de objetos IRR no TC (bgp.net.br) via API HTTP
# ═══════════════════════════════════════════════════════════════════════════════

IRR_TC_API_URL = 'https://bgp.net.br/v1/submit/'


def _irr_fmt_asn(asn):
    """Normaliza um ASN para o formato RPSL exigido (prefixo 'AS'), já que o
    formulário aceita tanto '16735' quanto 'AS16735' mas o IRR só aceita a
    segunda forma em `members:`/`import:`/`export:`."""
    s = str(asn or '').strip().upper()
    if s and not s.startswith('AS'):
        s = f'AS{s}'
    return s


def _irr_slug_as_name(name):
    """Deriva um identificador RPSL válido (letras/dígitos/hífen, sem espaços
    ou acentos) a partir de um nome livre (ex: razão social), já que campos
    como `as-name:` e o nome do as-set principal (AS-<nome>) rejeitam
    qualquer caractere fora desse conjunto."""
    import re
    import unicodedata
    ascii_only = unicodedata.normalize('NFKD', str(name or '')).encode('ascii', 'ignore').decode('ascii')
    slug = re.sub(r'[^A-Za-z0-9]+', '-', ascii_only).strip('-').upper()
    return slug


def _irr_normalizar_rota(item):
    """Aceita tanto o formato legado (string = apenas o prefixo) quanto o
    formato atual (dict com prefix/descr/member_of) e devolve sempre um dict."""
    if isinstance(item, dict):
        return {
            'prefix':    (item.get('prefix') or '').strip(),
            'descr':     (item.get('descr') or '').strip(),
            'member_of': (item.get('member_of') or '').strip(),
        }
    return {'prefix': str(item).strip(), 'descr': '', 'member_of': ''}


def _irr_gerar_objetos(cfg):
    """Gera a lista de objetos RPSL (um item por objeto) a partir de um IRRConfig.

    Cada item é o texto completo de um objeto (person, mntner, route-set, route,
    route6, as-set, aut-num) — sem a pseudo-linha `password:`, já que na API essa
    credencial vai à parte, no campo raiz `passwords`.
    """
    from datetime import date
    hoje = date.today().strftime('%Y%m%d')
    asn     = cfg.asn
    as_full = f'AS{asn}'
    mntner  = f'MAINT-AS{asn}'
    rs      = f'{as_full}:RS-ROUTES'
    email   = cfg.email_contato

    partes = []

    # ── Person ───────────────────────────────────────────────────────────────
    partes.append(
        f'person:  {cfg.person_name}\n'
        f'address: {cfg.address}\n'
        f'phone:   {cfg.phone}\n'
        f'e-mail:  {email}\n'
        f'nic-hdl: {cfg.nic_hdl}\n'
        f'mnt-by:  {mntner}\n'
        f'changed: {email} {hoje}\n'
        f'source:  TC\n'
    )

    # ── Mntner ────────────────────────────────────────────────────────────────
    # Gera hash bcrypt automaticamente a partir da senha plaintext
    import bcrypt as _bcrypt
    _bcrypt_hash = _bcrypt.hashpw(cfg.irr_password.encode(), _bcrypt.gensalt()).decode()
    auth_line = f'auth:    BCRYPT-PW {_bcrypt_hash}\n'
    partes.append(
        f'mntner:  {mntner}\n'
        f'descr:   {cfg.empresa_descr}\n'
        f'admin-c: {cfg.nic_hdl}\n'
        f'tech-c:  {cfg.nic_hdl}\n'
        f'upd-to:  {email}\n'
        f'mnt-nfy: {email}\n'
        + auth_line +
        f'mnt-by:  {mntner}\n'
        f'changed: {email} {hoje}\n'
        f'source:  TC\n'
    )

    # ── Route-set ─────────────────────────────────────────────────────────────
    mp_members_lines = ''
    for m in (cfg.route_set_members or []):
        mp_members_lines += f'mp-members: {m}\n'
    partes.append(
        f'route-set: {rs}\n'
        f'descr:     {cfg.empresa_descr}\n'
        + mp_members_lines +
        f'admin-c:   {cfg.nic_hdl}\n'
        f'tech-c:    {cfg.nic_hdl}\n'
        f'notify:    {email}\n'
        f'mnt-by:    {mntner}\n'
        f'changed:   {email} {hoje}\n'
        f'source:    TC\n'
    )

    # ── Geo lines helper ─────────────────────────────────────────────────────
    def geo_lines():
        linhas = ''
        if cfg.geo_pais:        linhas += f'geoidx: {cfg.geo_pais}\n'
        if cfg.geo_pais_alpha3: linhas += f'geoidx: {cfg.geo_pais_alpha3}\n'
        if cfg.geo_pais_num:    linhas += f'geoidx: {cfg.geo_pais_num}\n'
        if cfg.geo_estado:      linhas += f'geoidx: {cfg.geo_estado}\n'
        if cfg.geo_cidade:      linhas += f'geoidx: {cfg.geo_cidade}\n'
        return linhas

    # ── route objects ─────────────────────────────────────────────────────────
    # descr/member-of são por rota, com fallback para o padrão global
    # (empresa_descr / AS{asn}:RS-ROUTES) quando o item não define os seus.
    for item in (cfg.ipv4_rotas or []):
        r = _irr_normalizar_rota(item)
        if not r['prefix']:
            continue
        partes.append(
            f'route:  {r["prefix"]}\n'
            f'descr:  {r["descr"] or cfg.empresa_descr}\n'
            f'origin: {as_full}\n'
            f'member-of: {r["member_of"] or rs}\n'
            f'notify: {email}\n'
            + geo_lines() +
            f'mnt-by: {mntner}\n'
            f'changed: {email} {hoje}\n'
            f'source: TC\n'
        )

    # ── route6 objects ────────────────────────────────────────────────────────
    for item in (cfg.ipv6_rotas or []):
        r = _irr_normalizar_rota(item)
        if not r['prefix']:
            continue
        partes.append(
            f'route6: {r["prefix"]}\n'
            f'descr:  {r["descr"] or cfg.empresa_descr}\n'
            f'origin: {as_full}\n'
            f'member-of: {r["member_of"] or rs}\n'
            f'notify: {email}\n'
            + geo_lines() +
            f'mnt-by: {mntner}\n'
            f'changed: {email} {hoje}\n'
            f'source: TC\n'
        )

    # ── AS-set UPSTREAMS ──────────────────────────────────────────────────────
    up_members = ''
    for u in (cfg.upstream_asns or []):
        nome_comment = f'  # {u["nome"]}' if u.get('nome') else ''
        up_members += f'members: {_irr_fmt_asn(u["asn"])}{nome_comment}\n'
    partes.append(
        f'as-set: {as_full}:AS-UPSTREAMS\n'
        f'descr:  as-set containing {as_full} upstream providers\n'
        + up_members +
        f'admin-c: {cfg.nic_hdl}\n'
        f'tech-c:  {cfg.nic_hdl}\n'
        f'notify:  {email}\n'
        f'mnt-by:  {mntner}\n'
        f'changed: {email} {hoje}\n'
        f'source:  TC\n'
    )

    # ── AS-set CUSTOMERS ──────────────────────────────────────────────────────
    # `members` é opcional no as-set (RFC 2622) — sem clientes, a linha é
    # simplesmente omitida em vez de usar um valor inválido como placeholder.
    cust_members = ''
    for c in (cfg.customer_asns or []):
        if c.get('asn'):
            nome_comment = f'  # {c["nome"]}' if c.get('nome') else ''
            cust_members += f'members: {_irr_fmt_asn(c["asn"])}{nome_comment}\n'
    partes.append(
        f'as-set: {as_full}:AS-CUSTOMERS\n'
        f'descr:  as-set containing {as_full} and its downstream customers\n'
        + cust_members +
        f'admin-c: {cfg.nic_hdl}\n'
        f'tech-c:  {cfg.nic_hdl}\n'
        f'notify:  {email}\n'
        f'mnt-by:  {mntner}\n'
        f'changed: {email} {hoje}\n'
        f'source:  TC\n'
    )

    # ── AS-set principal ──────────────────────────────────────────────────────
    as_name_slug = _irr_slug_as_name(cfg.as_name)
    partes.append(
        f'as-set: {as_full}:AS-{as_name_slug}\n'
        f'descr:  {cfg.empresa_descr} - ANNOUNCEMENTS\n'
        f'members: {as_full}\n'
        f'members: {as_full}:AS-CUSTOMERS\n'
        f'admin-c: {cfg.nic_hdl}\n'
        f'tech-c:  {cfg.nic_hdl}\n'
        f'notify:  {email}\n'
        f'mnt-by:  {mntner}\n'
        f'changed: {email} {hoje}\n'
        f'source:  TC\n'
    )

    # ── aut-num ───────────────────────────────────────────────────────────────
    import_lines = ''
    export_lines = ''
    for u in (cfg.upstream_asns or []):
        import_lines += f'import:   from {_irr_fmt_asn(u["asn"])}  accept ANY\n'
    for u in (cfg.upstream_asns or []):
        export_lines += f'export:   to {_irr_fmt_asn(u["asn"])}  announce {as_full}:AS-{as_name_slug}\n'

    ix_lines = ''
    if cfg.ix_members:
        ix_lines = 'remarks:        ==========================================================\n'
        ix_lines += 'remarks:        Participante IX:\n'
        ix_lines += 'remarks:        ...\n'
        for ix in cfg.ix_members:
            ix_lines += f'member-of:      {ix}\n'
        ix_lines += 'remarks:        ...\n'
        ix_lines += 'remarks:        ==========================================================\n'

    abuse_email  = cfg.email_abuse or email
    website_line = f'remarks:        Website....................: {cfg.website}\n' if cfg.website else ''

    partes.append(
        f'aut-num:        {as_full}\n'
        f'as-name:        {as_name_slug}\n'
        f'descr:          {cfg.empresa_descr}\n'
        + ix_lines +
        f'remarks:        ==========================================================\n'
        + import_lines
        + export_lines +
        f'remarks:        ==========================================================\n'
        f'remarks:        Abuse/UCE..................: {abuse_email}\n'
        f'remarks:        Network....................: {abuse_email}\n'
        f'remarks:        Peering....................: https://{as_full}.peeringdb.com/\n'
        + website_line +
        f'remarks:        ==========================================================\n'
        f'admin-c:        {cfg.nic_hdl}\n'
        f'tech-c:         {cfg.nic_hdl}\n'
        f'mnt-by:         {mntner}\n'
        f'changed:        {email} {hoje}\n'
        f'source:         TC\n'
    )

    return partes


def _irr_gerar_corpo(cfg):
    """Gera o corpo textual completo (password: + objetos RPSL) — usado apenas
    no preview e como referência para envio manual por e-mail, se necessário."""
    objetos = _irr_gerar_objetos(cfg)
    return f'password: {cfg.irr_password}\n\n' + '\n\n'.join(objetos)


@login_required
@modulo_habilitado_required('rpki_irr')
def irr_config_get(request, cliente_id):
    """Retorna a configuração IRR do cliente (ou objeto vazio se não existir)."""
    from .models import IRRConfig
    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        cfg = cliente.irr_config
        data = {
            'ok': True,
            'existe': True,
            'asn': cfg.asn,
            'as_name': cfg.as_name,
            'empresa_descr': cfg.empresa_descr,
            'nic_hdl': cfg.nic_hdl,
            'irr_password': cfg.irr_password,
            'auth_bcrypt': cfg.auth_bcrypt,
            'api_key': cfg.api_key,
            'email_contato': cfg.email_contato,
            'email_abuse': cfg.email_abuse,
            'website': cfg.website,
            'person_name': cfg.person_name,
            'address': cfg.address,
            'phone': cfg.phone,
            'ipv4_rotas': cfg.ipv4_rotas,
            'ipv6_rotas': cfg.ipv6_rotas,
            'route_set_members': cfg.route_set_members,
            'upstream_asns': cfg.upstream_asns,
            'customer_asns': cfg.customer_asns,
            'ix_members': cfg.ix_members,
            'geo_pais': cfg.geo_pais,
            'geo_pais_alpha3': cfg.geo_pais_alpha3,
            'geo_pais_num': cfg.geo_pais_num,
            'geo_estado': cfg.geo_estado,
            'geo_cidade': cfg.geo_cidade,
        }
    except IRRConfig.DoesNotExist:
        # Tenta obter ASN a partir dos blocos IP já cadastrados
        from .models import BlocoIP
        bloco = (BlocoIP.objects
                 .filter(cliente=cliente)
                 .exclude(asn__isnull=True)
                 .exclude(asn='')
                 .order_by('id')
                 .first())
        asn_sugerido = ''
        if bloco and bloco.asn:
            asn_sugerido = bloco.asn.upper().lstrip('AS').strip()
        data = {'ok': True, 'existe': False, 'asn_sugerido': asn_sugerido}
    return JsonResponse(data)


@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('rpki_irr')
def irr_config_salvar(request, cliente_id):
    """Cria ou atualiza a configuração IRR do cliente."""
    from .models import IRRConfig
    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'ok': False, 'erro': 'JSON inválido'}, status=400)

    cfg, _ = IRRConfig.objects.get_or_create(cliente=cliente)

    campos_simples = [
        'asn','as_name','empresa_descr','nic_hdl','irr_password','auth_bcrypt','api_key',
        'email_contato','email_abuse','website','person_name','address','phone',
        'geo_pais','geo_pais_alpha3','geo_pais_num','geo_estado','geo_cidade',
    ]
    for campo in campos_simples:
        if campo in body:
            setattr(cfg, campo, body[campo])

    campos_json = ['ipv4_rotas','ipv6_rotas','route_set_members','upstream_asns','customer_asns','ix_members']
    for campo in campos_json:
        if campo in body:
            setattr(cfg, campo, body[campo])

    cfg.save()
    return JsonResponse({'ok': True})


@login_required
@modulo_habilitado_required('rpki_irr')
def irr_preview(request, cliente_id):
    """Retorna o preview do e-mail IRR gerado a partir da config salva."""
    from .models import IRRConfig
    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        cfg = cliente.irr_config
    except IRRConfig.DoesNotExist:
        return JsonResponse({'ok': False, 'erro': 'Configuração IRR não encontrada. Salve os dados primeiro.'}, status=404)
    corpo = _irr_gerar_corpo(cfg)
    return JsonResponse({
        'ok': True, 'corpo': corpo,
        'assunto': 'IRR Route Update',
        'destino': 'auto-dbm@bgp.net.br',
        'api_endpoint': IRR_TC_API_URL,
    })


@login_required
@require_http_methods(['POST'])
@modulo_habilitado_required('rpki_irr')
def irr_enviar(request, cliente_id):
    """Envia a atualização de objetos IRR ao TC via API HTTP (bgp.net.br/v1/submit).

    Substitui o antigo envio por e-mail para auto-dbm@bgp.net.br: a API aceita o
    mesmo texto RPSL por objeto e a mesma senha do mntner (campo `passwords`),
    mas responde de forma síncrona com o resultado por objeto — não é mais
    necessário verificar uma caixa de e-mail depois.
    """
    from .models import IRRConfig

    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        cfg = cliente.irr_config
    except IRRConfig.DoesNotExist:
        return JsonResponse({'ok': False, 'erro': 'Configuração IRR não encontrada.'}, status=404)

    if not cfg.irr_password and not cfg.api_key:
        return JsonResponse({'ok': False, 'erro': 'Informe a senha do mntner (ou API key) antes de enviar.'}, status=400)

    objetos = _irr_gerar_objetos(cfg)
    payload = {'objects': [{'object_text': obj} for obj in objetos]}
    if cfg.irr_password:
        payload['passwords'] = [cfg.irr_password]
    if cfg.api_key:
        payload['api_keys'] = [cfg.api_key]

    try:
        # Submissões válidas (objetos reais, senha correta) podem levar bem mais
        # tempo que payloads inválidos — o TC responde rejeições quase na hora,
        # mas criações/alterações reais parecem ficar perto do teto de 120s do
        # worker gunicorn. 100s dá folga suficiente sem estourar esse teto.
        resp = requests.post(IRR_TC_API_URL, json=payload, timeout=100)
    except requests.RequestException as e:
        return JsonResponse({'ok': False, 'erro': f'Falha ao conectar à API do TC: {e}'}, status=502)

    if resp.status_code == 400:
        return JsonResponse({'ok': False, 'erro': f'JSON rejeitado pela API do TC: {resp.text[:500]}'}, status=400)

    try:
        resultado = resp.json()
    except ValueError:
        return JsonResponse({
            'ok': False,
            'erro': f'Resposta inesperada da API do TC (HTTP {resp.status_code}): {resp.text[:500]}',
        }, status=502)

    resumo   = resultado.get('summary', {}) or {}
    objs_out = resultado.get('objects', []) or []

    aceitos, rejeitados, erros = [], [], []
    for o in objs_out:
        rotulo = f"{o.get('object_class', '?')}: {o.get('rpsl_pk', '')}"
        if o.get('successful'):
            aceitos.append(rotulo)
        else:
            rejeitados.append(rotulo)
            for msg_erro in (o.get('error_messages') or []):
                erros.append(f'{rotulo} — {msg_erro}')

    total_falhas = resumo.get('failed', len(rejeitados))
    if total_falhas == 0:
        status_geral = 'sucesso'
    elif aceitos:
        status_geral = 'parcial'
    else:
        status_geral = 'erro'

    return JsonResponse({
        'ok':           True,
        'status_geral': status_geral,
        'resumo':       resumo,
        'aceitos':      aceitos,
        'rejeitados':   rejeitados,
        'erros':        erros,
        'objetos':      objs_out,
        'mensagem':     f"{resumo.get('successful', len(aceitos))} objeto(s) aceito(s), {total_falhas} falha(s).",
    })


@login_required
@modulo_habilitado_required('rpki_irr')
def irr_consultar_whois(request, cliente_id):
    """Consulta o WHOIS do TC/NIC.br para o ASN informado e retorna objetos parsados."""
    import socket
    asn_param = request.GET.get('asn', '').strip().upper().lstrip('AS').strip()
    if not asn_param:
        return JsonResponse({'ok': False, 'erro': 'Informe o número do AS'}, status=400)

    asn_full = f'AS{asn_param}'

    def whois_query(server, query, port=43, timeout=10):
        try:
            with socket.create_connection((server, port), timeout=timeout) as s:
                s.sendall((query + '\r\n').encode())
                resp = b''
                while True:
                    chunk = s.recv(4096)
                    if not chunk:
                        break
                    resp += chunk
            return resp.decode('utf-8', errors='replace')
        except Exception as e:
            return f'%% Erro: {e}'

    # Tenta servidores em ordem
    servidores = ['whois.nic.br', 'irr.nic.br']
    resultados = {}

    for srv in servidores:
        # aut-num
        resp_autnum = whois_query(srv, f'-T aut-num {asn_full}')
        if asn_full in resp_autnum and '%% Not found' not in resp_autnum and 'Erro:' not in resp_autnum:
            resultados['aut_num_raw'] = resp_autnum
            resultados['servidor']    = srv
            break

    if not resultados:
        # Tenta RADB como fallback
        resp_radb = whois_query('whois.radb.net', f'-T aut-num {asn_full}')
        if asn_full in resp_radb and 'Not found' not in resp_radb:
            resultados['aut_num_raw'] = resp_radb
            resultados['servidor']    = 'whois.radb.net'

    # Busca routes
    srv_final = resultados.get('servidor', 'whois.nic.br')
    resp_routes  = whois_query(srv_final, f'-T route -i origin {asn_full}')
    resp_routes6 = whois_query(srv_final, f'-T route6 -i origin {asn_full}')
    resp_mntner  = whois_query(srv_final, f'-T mntner MAINT-{asn_full}')

    def parse_field(texto, campo):
        """Extrai o primeiro valor de um campo RPSL."""
        for linha in texto.splitlines():
            if linha.lower().startswith(campo.lower() + ':'):
                return linha.split(':', 1)[1].strip()
        return ''

    def parse_all_field(texto, campo):
        """Extrai todos os valores de um campo repetido."""
        valores = []
        for linha in texto.splitlines():
            if linha.lower().startswith(campo.lower() + ':'):
                v = linha.split(':', 1)[1].strip()
                if v and not v.startswith('#'):
                    valores.append(v)
        return valores

    def parse_route_objects(texto, tipo='route'):
        """Extrai objetos route/route6 completos (prefix + descr + member-of).
        No retorno do whois cada objeto RPSL é separado por linha em branco,
        então uma nova linha `route:`/`route6:` sempre abre um bloco novo.

        Um mesmo prefixo pode aparecer em mais de um objeto — o registro real
        (source: TC, o mesmo que este sistema gerencia), uma versão
        auto-gerada a partir do RPKI (source: RPKI) e às vezes um terceiro via
        RADB/outro mnt-by. Por isso o resultado é deduplicado por prefixo,
        priorizando sempre o objeto com source TC quando existir."""
        objetos = []
        atual = None
        for linha in texto.splitlines():
            if not linha.strip():
                continue
            if ':' not in linha:
                continue
            campo, valor = linha.split(':', 1)
            campo = campo.strip().lower()
            valor = valor.split('#', 1)[0].strip()  # descarta comentário inline (comum em source/changed)
            if campo == tipo:
                if atual and atual.get('prefix'):
                    objetos.append(atual)
                atual = {'prefix': valor, 'descr': '', 'member_of': '', 'source': ''}
            elif atual is not None and campo == 'descr' and not atual['descr']:
                atual['descr'] = valor
            elif atual is not None and campo == 'member-of' and not atual['member_of']:
                atual['member_of'] = valor
            elif atual is not None and campo == 'source' and not atual['source']:
                atual['source'] = valor
        if atual and atual.get('prefix'):
            objetos.append(atual)

        por_prefixo = {}
        ordem = []
        for obj in objetos:
            p = obj['prefix']
            if p not in por_prefixo:
                por_prefixo[p] = obj
                ordem.append(p)
            elif obj['source'].upper() == 'TC' and por_prefixo[p]['source'].upper() != 'TC':
                por_prefixo[p] = obj
        return [
            {'prefix': p, 'descr': por_prefixo[p]['descr'], 'member_of': por_prefixo[p]['member_of']}
            for p in ordem
        ]

    def parse_remarks_email(texto, label_prefix):
        """Extrai e-mail de linhas remarks no formato 'remarks: Label: email@dominio'."""
        for linha in texto.splitlines():
            l = linha.strip()
            if l.lower().startswith('remarks:'):
                resto = l.split(':', 1)[1].strip()
                if resto.lower().startswith(label_prefix.lower()):
                    m = re.search(r'[\w.+-]+@[\w.-]+\.\w+', resto)
                    if m:
                        return m.group(0)
        return ''

    dados = {
        'ok':      True,
        'servidor': srv_final,
        'asn':     asn_param,
    }

    autnum_raw = resultados.get('aut_num_raw', '')
    if autnum_raw:
        dados['as_name']       = parse_field(autnum_raw, 'as-name')
        dados['descr']         = parse_field(autnum_raw, 'descr')
        changed_parts = parse_field(autnum_raw, 'changed').split()
        dados['email_abuse']   = parse_remarks_email(autnum_raw, 'abuse')
        dados['email_contato'] = (
            parse_field(autnum_raw, 'e-mail')
            or (changed_parts[0] if changed_parts else '')
            or parse_remarks_email(autnum_raw, 'network')
            or dados['email_abuse']
        )
        dados['nic_hdl']       = parse_field(autnum_raw, 'admin-c')
        dados['mntner']        = parse_field(autnum_raw, 'mnt-by')

        # AS-SETs/grupos de anúncio (member-of) — ex: AS271699:AS-ANNOUNCEMENTS, AS-PTTMetro-SP
        dados['ix_members'] = parse_all_field(autnum_raw, 'member-of')

        # import/mp-import → upstream ASNs (registros modernos usam mp-import para dual-stack)
        imports = parse_all_field(autnum_raw, 'import') + parse_all_field(autnum_raw, 'mp-import')
        upstream = []
        vistos = set()
        for imp in imports:
            # "from AS52554 accept ANY" / "from AS269595 action pref = 100; accept ANY" → AS52554
            parts = imp.split()
            if len(parts) >= 2 and parts[0].lower() == 'from':
                asn_up = parts[1].rstrip(';,')
                if asn_up not in vistos:
                    vistos.add(asn_up)
                    upstream.append({'asn': asn_up, 'nome': ''})
        dados['upstream_asns'] = upstream

    # Rotas IPv4 — cada item traz prefix/descr/member_of já preenchidos, se existirem no IRR
    dados['ipv4_rotas'] = parse_route_objects(resp_routes, 'route')

    # Rotas IPv6
    dados['ipv6_rotas'] = parse_route_objects(resp_routes6, 'route6')

    # Auth hash do mntner
    if resp_mntner and 'BCRYPT-PW' in resp_mntner:
        for linha in resp_mntner.splitlines():
            if 'auth' in linha.lower() and 'BCRYPT-PW' in linha:
                dados['auth_bcrypt'] = linha.split(':', 1)[1].strip()
                break

    dados['raw'] = {
        'aut_num':  autnum_raw[:8000] if autnum_raw else '',
        'routes':   resp_routes[:8000],
        'routes6':  resp_routes6[:8000],
        'mntner':   resp_mntner[:4000],
    }

    return JsonResponse(dados)


# ─────────────────────────────────────────────────────────────────────────────
# TROCAR SENHAS EM MASSA
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def trocar_senha_massa(request):
    """Página principal da ferramenta Trocar Senhas em Massa."""
    if not request.user.is_superuser:
        return HttpResponseForbidden('Acesso restrito a administradores.')

    from .models import TrocaSenhaJob, TrocaSenhaAcesso, Cliente as ClienteModel
    from django.db.models import Count, Sum, Max, Q
    import secrets, string

    clientes = ClienteModel.objects.all().order_by('nome_empresa')

    # ── Stats globais ─────────────────────────────────────────────────────────
    total_jobs      = TrocaSenhaJob.objects.count()
    total_sucesso   = TrocaSenhaJob.objects.aggregate(s=Sum('total_sucesso'))['s'] or 0
    total_erro      = TrocaSenhaJob.objects.aggregate(s=Sum('total_erro'))['s'] or 0
    total_removidos = TrocaSenhaAcesso.objects.filter(usuario_removido=True).count()
    clientes_c_jobs = TrocaSenhaJob.objects.values('cliente_id').distinct().count()
    taxa_sucesso    = round(total_sucesso / (total_sucesso + total_erro) * 100) if (total_sucesso + total_erro) > 0 else 0

    # ── Por cliente (últimos jobs) ────────────────────────────────────────────
    ids_com_jobs = set(TrocaSenhaJob.objects.values_list('cliente_id', flat=True).distinct())

    clientes_painel      = []   # processados
    clientes_sem_painel  = []   # não processados
    for c in ClienteModel.objects.prefetch_related('troca_senha_jobs').order_by('nome_empresa'):
        if c.id in ids_com_jobs:
            ultimo_job = c.troca_senha_jobs.first()
            clientes_painel.append({
                'cliente':    c,
                'ultimo_job': ultimo_job,
                'total_jobs': c.troca_senha_jobs.count(),
            })
        else:
            total_ssh = Acesso.objects.filter(cliente=c, protocolo='SSH').count()
            clientes_sem_painel.append({
                'cliente':   c,
                'total_ssh': total_ssh,
            })

    clientes_s_jobs = len(clientes_sem_painel)

    # ── Jobs recentes (todos os clientes) ────────────────────────────────────
    jobs_recentes = TrocaSenhaJob.objects.select_related('cliente', 'criado_por').prefetch_related('itens')[:15]

    # ── Cliente selecionado ───────────────────────────────────────────────────
    cliente_id = request.GET.get('cliente_id')
    cliente_selecionado = None
    jobs_cliente = []
    if cliente_id:
        try:
            cliente_selecionado = ClienteModel.objects.get(pk=cliente_id)
            jobs_cliente = TrocaSenhaJob.objects.filter(
                cliente=cliente_selecionado
            ).prefetch_related('itens')[:20]
        except ClienteModel.DoesNotExist:
            pass

    chars = string.ascii_letters + string.digits + '!@#$%&*'
    senha_sugerida = ''.join(secrets.choice(chars) for _ in range(16))

    return render(request, 'troca_senha_massa.html', {
        'clientes':           clientes,
        'cliente_selecionado': cliente_selecionado,
        'jobs':               jobs_cliente,
        'senha_sugerida':     senha_sugerida,
        'is_admin':           True,
        'is_superuser':       True,
        # dashboard
        'total_jobs':         total_jobs,
        'total_sucesso':      total_sucesso,
        'total_erro':         total_erro,
        'clientes_s_jobs':    clientes_s_jobs,
        'clientes_sem_painel': clientes_sem_painel,
        'total_removidos':    total_removidos,
        'clientes_c_jobs':    clientes_c_jobs,
        'taxa_sucesso':       taxa_sucesso,
        'clientes_painel':    clientes_painel,
        'jobs_recentes':      jobs_recentes,
    })


@login_required
@require_http_methods(['GET'])
def trocar_senha_massa_listar_hosts(request):
    """Retorna JSON com todos os acessos SSH do cliente + vendor detectado."""
    if not request.user.is_superuser:
        return JsonResponse({'erro': 'Acesso restrito.'}, status=403)

    from .models import Cliente as ClienteModel
    from .tasks import _detectar_vendor

    cliente_id = request.GET.get('cliente_id')
    if not cliente_id:
        return JsonResponse({'erro': 'cliente_id obrigatório.'}, status=400)

    try:
        cliente = ClienteModel.objects.get(pk=cliente_id)
    except ClienteModel.DoesNotExist:
        return JsonResponse({'erro': 'Cliente não encontrado.'}, status=404)

    acessos = Acesso.objects.filter(
        cliente=cliente, protocolo='SSH'
    ).select_related('modelo', 'funcao').order_by('tipo', 'host')

    hosts = []
    for a in acessos:
        try:
            vendor = _detectar_vendor(a)
        except Exception:
            vendor = 'desconhecido'
        hosts.append({
            'id':        a.id,
            'tipo':      a.tipo,
            'host':      a.host,
            'porta':     a.porta,
            'usuario':   a.usuario,
            'vendor':    vendor,
            'fabricante': (a.modelo.fabricante if a.modelo else '') or '',
            'funcao':    (a.funcao.descricao if a.funcao else '') or '',
        })

    return JsonResponse({'hosts': hosts, 'total': len(hosts)})


@login_required
@require_http_methods(['POST'])
def trocar_senha_massa_iniciar(request):
    """Inicia um job de troca de senhas em massa para um cliente."""
    if not request.user.is_superuser:
        return JsonResponse({'erro': 'Acesso restrito a administradores.'}, status=403)

    from .models import TrocaSenhaJob, Cliente as ClienteModel
    from .tasks import executar_troca_senhas_massa

    cliente_id   = request.POST.get('cliente_id')
    novo_usuario = request.POST.get('novo_usuario', '').strip()
    nova_senha   = request.POST.get('nova_senha', '').strip()
    acesso_ids   = [int(x) for x in request.POST.getlist('acesso_ids[]') if x.isdigit()]

    if not cliente_id or not novo_usuario or not nova_senha:
        return JsonResponse({'erro': 'Preencha todos os campos.'}, status=400)

    try:
        cliente = ClienteModel.objects.get(pk=cliente_id)
    except ClienteModel.DoesNotExist:
        return JsonResponse({'erro': 'Cliente não encontrado.'}, status=404)

    if acesso_ids:
        # Valida que os IDs pertencem ao cliente
        validos = list(
            Acesso.objects.filter(pk__in=acesso_ids, cliente=cliente, protocolo='SSH')
            .values_list('pk', flat=True)
        )
        if not validos:
            return JsonResponse({'erro': 'Nenhum host SSH válido selecionado.'}, status=400)
        acesso_ids = validos
    else:
        acesso_ids = None  # task vai pegar todos

    job = TrocaSenhaJob.objects.create(
        cliente=cliente,
        criado_por=request.user,
        novo_usuario=novo_usuario,
        nova_senha=nova_senha,
    )

    executar_troca_senhas_massa.delay(job.id, acesso_ids=acesso_ids)

    return JsonResponse({'ok': True, 'job_id': job.id})


@login_required
def trocar_senha_massa_status(request, job_id):
    """Retorna JSON com o status atual de um job."""
    if not request.user.is_superuser:
        return JsonResponse({'erro': 'Acesso restrito.'}, status=403)

    from .models import TrocaSenhaJob

    try:
        job = TrocaSenhaJob.objects.prefetch_related('itens__acesso').get(pk=job_id)
    except TrocaSenhaJob.DoesNotExist:
        return JsonResponse({'erro': 'Job não encontrado.'}, status=404)

    itens = []
    for item in job.itens.all():
        itens.append({
            'id':               item.id,
            'acesso_tipo':      item.acesso.tipo if item.acesso else '(removido)',
            'acesso_host':      item.acesso.host if item.acesso else '',
            'status':           item.status,
            'mensagem':         item.mensagem,
            'usuario_antigo':   item.usuario_antigo,
            'usuario_removido': item.usuario_removido,
            'duracao':          item.duracao_segundos,
        })

    return JsonResponse({
        'job_id':         job.id,
        'status':         job.status,
        'novo_usuario':   job.novo_usuario,
        'total_acessos':  job.total_acessos,
        'total_sucesso':  job.total_sucesso,
        'total_erro':     job.total_erro,
        'criado_em':      job.criado_em.strftime('%d/%m/%Y %H:%M'),
        'concluido_em':   job.concluido_em.strftime('%d/%m/%Y %H:%M') if job.concluido_em else None,
        'itens':          itens,
    })


@login_required
@require_http_methods(['POST'])
def trocar_senha_remover_antigos(request, job_id):
    """Dispara a task de remoção dos usuários antigos para um job concluído."""
    if not request.user.is_superuser:
        return JsonResponse({'erro': 'Acesso restrito.'}, status=403)

    from .models import TrocaSenhaJob
    from .tasks import remover_usuarios_antigos_task

    try:
        job = TrocaSenhaJob.objects.get(pk=job_id)
    except TrocaSenhaJob.DoesNotExist:
        return JsonResponse({'erro': 'Job não encontrado.'}, status=404)

    if job.status != 'CONCLUIDO':
        return JsonResponse({'erro': 'Job ainda não concluído.'}, status=400)

    remover_usuarios_antigos_task.delay(job_id)
    return JsonResponse({'ok': True, 'mensagem': 'Remoção dos usuários antigos iniciada.'})
